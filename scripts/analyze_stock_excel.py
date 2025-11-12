#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script d'analyse du fichier stock_initial_COMPLET.xlsx
"""

from openpyxl import load_workbook
from collections import Counter
import os

file_path = '/home/one/apps/stockex/docx/stock_initial_COMPLET.xlsx'

print("=" * 90)
print(" " * 20 + "ANALYSE: stock_initial_COMPLET.xlsx")
print("=" * 90)

# Taille du fichier
file_size = os.path.getsize(file_path) / (1024 * 1024)
print(f"\n📦 Taille du fichier: {file_size:.2f} Mo")

# Charger le fichier
wb = load_workbook(file_path, read_only=True, data_only=True)
print(f"📋 Feuilles: {wb.sheetnames}")

# Analyser la feuille Stock Initial
ws = wb['Stock Initial']

print(f"\n{'=' * 90}")
print(" " * 30 + "STRUCTURE DU FICHIER")
print(f"{'=' * 90}")

print(f"\n📊 Dimensions: {ws.max_row:,} lignes × {ws.max_column} colonnes")

# En-têtes
print(f"\n📝 Colonnes:")
for col in range(1, ws.max_column + 1):
    header = ws.cell(1, col).value
    print(f"   {col}. {header}")

# Aperçu des données
print(f"\n👁️  Aperçu des 10 premières lignes:")
print(f"{'─' * 90}")
for row in range(2, min(12, ws.max_row + 1)):
    code = ws.cell(row, 1).value or ""
    produit = ws.cell(row, 2).value or ""
    categorie = ws.cell(row, 3).value or ""
    entrepot = ws.cell(row, 5).value or ""
    qte = ws.cell(row, 6).value or ""
    prix = ws.cell(row, 7).value or ""
    print(f"{row-1:3d}. {str(code)[:15]:15s} | {str(produit)[:25]:25s} | {str(categorie)[:20]:20s} | Qté: {str(qte):>6s} | Prix: {str(prix):>10s}")

# Collecte des données
print(f"\n{'=' * 90}")
print(" " * 30 + "STATISTIQUES DÉTAILLÉES")
print(f"{'=' * 90}")

categories = []
entrepots = []
codes_produits = []
quantites = []
prix_list = []
valeurs = []

print(f"\n⏳ Lecture des {ws.max_row - 1:,} lignes en cours...")

for row in range(2, ws.max_row + 1):
    code = ws.cell(row, 1).value
    cat = ws.cell(row, 3).value
    ent = ws.cell(row, 5).value
    qte = ws.cell(row, 6).value
    prix = ws.cell(row, 7).value
    
    if code:
        codes_produits.append(str(code))
    if cat:
        categories.append(str(cat))
    if ent:
        entrepots.append(str(ent))
    if qte is not None:
        try:
            q = float(qte)
            quantites.append(q)
            if prix is not None:
                try:
                    p = float(prix)
                    prix_list.append(p)
                    valeurs.append(q * p)
                except:
                    pass
        except:
            pass

# Afficher les statistiques
print(f"\n{'─' * 90}")
print(f"📈 STATISTIQUES GÉNÉRALES")
print(f"{'─' * 90}")
print(f"   • Total lignes (hors en-tête): {ws.max_row - 1:,}")
print(f"   • Codes produits uniques: {len(set(codes_produits)):,}")
print(f"   • Catégories uniques: {len(set(categories))}")
print(f"   • Entrepôts uniques: {len(set(entrepots))}")

# Catégories
print(f"\n{'─' * 90}")
print(f"📁 CATÉGORIES ({len(set(categories))} catégories)")
print(f"{'─' * 90}")
cat_count = Counter(categories)
for cat, count in cat_count.most_common():
    pct = (count / (ws.max_row - 1)) * 100
    print(f"   • {cat:45s}: {count:5,} lignes ({pct:5.1f}%)")

# Entrepôts
print(f"\n{'─' * 90}")
print(f"🏢 ENTREPÔTS ({len(set(entrepots))} entrepôts)")
print(f"{'─' * 90}")
ent_count = Counter(entrepots)
for ent, count in ent_count.most_common():
    pct = (count / (ws.max_row - 1)) * 100
    print(f"   • {ent:35s}: {count:5,} lignes ({pct:5.1f}%)")

# Quantités
print(f"\n{'─' * 90}")
print(f"📦 QUANTITÉS")
print(f"{'─' * 90}")
print(f"   • Lignes avec quantité: {len(quantites):,}")
print(f"   • Lignes sans quantité: {ws.max_row - 1 - len(quantites):,}")
if quantites:
    print(f"   • Quantité minimale: {min(quantites):,.2f}")
    print(f"   • Quantité maximale: {max(quantites):,.2f}")
    print(f"   • Quantité totale: {sum(quantites):,.2f}")
    print(f"   • Quantité moyenne: {sum(quantites)/len(quantites):,.2f}")

# Prix
print(f"\n{'─' * 90}")
print(f"💰 PRIX UNITAIRES")
print(f"{'─' * 90}")
print(f"   • Lignes avec prix: {len(prix_list):,}")
print(f"   • Lignes sans prix: {ws.max_row - 1 - len(prix_list):,}")
if prix_list:
    print(f"   • Prix minimal: {min(prix_list):,.2f} FCFA")
    print(f"   • Prix maximal: {max(prix_list):,.2f} FCFA")
    print(f"   • Prix moyen: {sum(prix_list)/len(prix_list):,.2f} FCFA")

# Valorisation
print(f"\n{'─' * 90}")
print(f"💵 VALORISATION (Quantité × Prix)")
print(f"{'─' * 90}")
if valeurs:
    print(f"   • Lignes valorisées: {len(valeurs):,}")
    print(f"   • Valorisation totale: {sum(valeurs):,.2f} FCFA")
    print(f"   • Valorisation minimale: {min(valeurs):,.2f} FCFA")
    print(f"   • Valorisation maximale: {max(valeurs):,.2f} FCFA")
    print(f"   • Valorisation moyenne: {sum(valeurs)/len(valeurs):,.2f} FCFA")

wb.close()

print(f"\n{'=' * 90}")
print(" " * 35 + "FIN DE L'ANALYSE")
print(f"{'=' * 90}\n")

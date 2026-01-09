from openpyxl import load_workbook
from collections import Counter
import os
import sys

file_path = '/mnt/extra-addons/stockex/docx/stock_initial_COMPLET.xlsx'

print("=" * 80)
print("ANALYSE COMPLETE: stock_initial_COMPLET.xlsx")
print("=" * 80)

if not os.path.exists(file_path):
    print(f"Fichier non trouve: {file_path}")
    sys.exit(1)

file_size = os.path.getsize(file_path) / (1024 * 1024)
print(f"\nTaille du fichier: {file_size:.2f} Mo")

wb = load_workbook(file_path, read_only=True, data_only=True)
print(f"Nombre de feuilles: {len(wb.sheetnames)}")
print(f"Feuilles: {wb.sheetnames}")

# Analyser la feuille Stock Initial
ws = wb['Stock Initial']

print(f"\n{'=' * 80}")
print(f"FEUILLE: Stock Initial")
print(f"{'=' * 80}")

max_row = ws.max_row
max_col = ws.max_column
print(f"\nDimensions: {max_row} lignes x {max_col} colonnes")

# En-têtes
headers = []
for col in range(1, max_col + 1):
    cell_value = ws.cell(1, col).value
    headers.append(cell_value if cell_value else f"Col_{col}")

print(f"\nColonnes ({len(headers)}):")
for i, header in enumerate(headers, 1):
    print(f"  {i}. {header}")

# Aperçu
print(f"\nApercu des 10 premieres lignes:")
for row_idx in range(2, min(12, max_row + 1)):
    row_data = []
    for col_idx in range(1, max_col + 1):
        val = ws.cell(row_idx, col_idx).value
        row_data.append(str(val)[:20] if val is not None else "")
    print(f"  Ligne {row_idx}: {' | '.join(row_data)}")

# Collecte des données pour statistiques
print(f"\n{'=' * 80}")
print("STATISTIQUES DETAILLEES")
print(f"{'=' * 80}")

categories = []
entrepots = []
codes_produits = []
produits = []
quantites = []
prix = []
valeurs = []

print("\nLecture des donnees en cours...")
for row in range(2, max_row + 1):
    code_produit = ws.cell(row, 1).value
    produit = ws.cell(row, 2).value
    categorie = ws.cell(row, 3).value
    entrepot = ws.cell(row, 5).value
    quantite = ws.cell(row, 6).value
    prix_unit = ws.cell(row, 7).value
    
    if code_produit:
        codes_produits.append(str(code_produit))
    if produit:
        produits.append(str(produit))
    if categorie:
        categories.append(str(categorie))
    if entrepot:
        entrepots.append(str(entrepot))
    if quantite is not None:
        try:
            qty = float(quantite)
            quantites.append(qty)
            if prix_unit is not None:
                try:
                    pu = float(prix_unit)
                    prix.append(pu)
                    valeurs.append(qty * pu)
                except:
                    pass
        except:
            pass

# Afficher les statistiques
print(f"\nStatistiques generales:")
print(f"  - Total lignes (sans en-tete): {max_row - 1}")
print(f"  - Codes produits uniques: {len(set(codes_produits))}")
print(f"  - Noms produits uniques: {len(set(produits))}")
print(f"  - Categories uniques: {len(set(categories))}")
print(f"  - Entrepots uniques: {len(set(entrepots))}")

print(f"\nCategories ({len(set(categories))})")
cat_count = Counter(categories)
for cat, count in cat_count.most_common():
    print(f"  - {cat}: {count} lignes")

print(f"\nEntrepots ({len(set(entrepots))})")
ent_count = Counter(entrepots)
for ent, count in ent_count.most_common():
    print(f"  - {ent}: {count} lignes")

print(f"\nQuantites:")
print(f"  - Lignes avec quantite: {len(quantites)}")
print(f"  - Lignes sans quantite: {max_row - 1 - len(quantites)}")
if quantites:
    print(f"  - Quantite min: {min(quantites):.2f}")
    print(f"  - Quantite max: {max(quantites):.2f}")
    print(f"  - Quantite totale: {sum(quantites):.2f}")
    print(f"  - Quantite moyenne: {sum(quantites)/len(quantites):.2f}")

print(f"\nPrix unitaires:")
print(f"  - Lignes avec prix: {len(prix)}")
print(f"  - Lignes sans prix: {max_row - 1 - len(prix)}")
if prix:
    print(f"  - Prix min: {min(prix):.2f} FCFA")
    print(f"  - Prix max: {max(prix):.2f} FCFA")
    print(f"  - Prix moyen: {sum(prix)/len(prix):.2f} FCFA")

print(f"\nValorisation:")
if valeurs:
    print(f"  - Lignes valorisees: {len(valeurs)}")
    print(f"  - Valorisation totale: {sum(valeurs):,.2f} FCFA")
    print(f"  - Valorisation min: {min(valeurs):.2f} FCFA")
    print(f"  - Valorisation max: {max(valeurs):.2f} FCFA")
    print(f"  - Valorisation moyenne: {sum(valeurs)/len(valeurs):.2f} FCFA")

wb.close()

print(f"\n{'=' * 80}")
print("FIN DE L'ANALYSE")
print(f"{'=' * 80}")

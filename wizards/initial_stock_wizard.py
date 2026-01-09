# -*- coding: utf-8 -*-

import logging
from odoo import models, fields, api
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)


class InitialStockWizard(models.TransientModel):
    """Assistant pour créer le stock initial dans une base vide."""
    _name = 'stockex.initial.stock.wizard'
    _description = 'Assistant Stock Initial'
    
    name = fields.Char(
        string='Nom de l\'Inventaire Initial',
        default='Stock Initial',
        required=True
    )
    
    date = fields.Date(
        string='Date du Stock Initial',
        required=True,
        default=fields.Date.today,
        help='Date de référence pour le stock initial (ex: date de mise en service)'
    )
    
    location_id = fields.Many2one(
        comodel_name='stock.warehouse',
        string='Entrepôt par Défaut',
        required=False,
        help='Entrepôt par défaut si non spécifié dans le fichier Excel. Si le fichier contient une colonne ENTREPOT, elle sera prioritaire.'
    )
    
    create_warehouses = fields.Boolean(
        string='Créer les Entrepôts Manquants',
        default=True,
        help='Si coché, crée automatiquement les entrepôts non trouvés'
    )
    
    import_file = fields.Binary(
        string='Fichier d\'Import (Excel)',
        help='Fichier Excel avec colonnes: CODE PRODUIT, QUANTITE, PRIX UNITAIRE'
    )
    
    filename = fields.Char(string='Nom du Fichier')
    
    create_products = fields.Boolean(
        string='Créer les Produits Manquants',
        default=True,
        help='Si coché, crée automatiquement les produits non trouvés'
    )
    
    create_categories = fields.Boolean(
        string='Créer les Catégories Manquantes',
        default=True,
        help='Si coché, crée automatiquement les catégories de produits non trouvées'
    )
    
    company_id = fields.Many2one(
        comodel_name='res.company',
        string='Société',
        default=lambda self: self.env.company,
        required=True
    )
    
    force_reset = fields.Boolean(
        string='⚠️ Réinitialiser tous les stocks',
        default=False,
        help='ATTENTION : Supprime TOUS les mouvements de stock et remet toutes les quantités à 0 avant l\'import. '
             'Cette action est IRRÉVERSIBLE. À utiliser uniquement pour une réinitialisation complète de la base.'
    )
    
    incremental_update = fields.Boolean(
        string='🔄 Ajouter au stock existant (mise à jour incrémentale)',
        default=False,
        help='⚠️ ATTENTION : Cette option permet d\'ajouter des quantités aux stocks existants.\n'
             '• Les produits existants verront leurs quantités AJOUTÉES (pas remplacées)\n'
             '• Les nouveaux produits seront créés\n'
             '\n'
             '⚡ Cas d\'usage typique :\n'
             '  - Import partiel d\'un entrepôt supplémentaire (ex: NKONGSAMBA après DLA)\n'
             '  - Complément d\'un import précédent\n'
             '\n'
             '🔒 Si DÉCOCHÉ (recommandé) : Vérifie qu\'aucun stock n\'existe déjà pour éviter les doublons.'
    )
    
    # Nouveau champ pour l'option d'enregistrement des mouvements
    create_stock_moves = fields.Boolean(
        string='📦 Enregistrer les mouvements de stock',
        default=False,
        help='Si coché, crée des mouvements de stock lors de l\'import du stock initial.\n'
             '⚠️ Cela ralentira l\'import mais permettra une traçabilité complète des mouvements.\n'
             'Par défaut, les quantités sont mises à jour directement sans mouvements.'
    )
    
    # Champs pour la prévisualisation
    state = fields.Selection([
        ('step1', 'Import'),
        ('step2', 'Prévisualisation'),
    ], string='Étape', default='step1')
    
    preview_data = fields.Text(string='Données de prévisualisation', readonly=True)
    preview_summary = fields.Html(string='Résumé', readonly=True)
    lines_count = fields.Integer(string='Nombre de lignes', readonly=True)
    warehouses_preview = fields.Char(string='Entrepôts détectés', readonly=True)
    categories_preview = fields.Char(string='Catégories détectées', readonly=True)
    
    # Champ de progression
    progress = fields.Float(string='Progression', readonly=True, default=0.0)
    progress_message = fields.Char(string='Message de progression', readonly=True)
    
    def action_preview(self):
        """Prévisualiser les données avant import."""
        self.ensure_one()
        
        if not self.import_file:
            raise UserError("⚠️ Veuillez sélectionner un fichier Excel.")
        
        # Parser le fichier
        lines = self._parse_excel_file()
        
        if not lines:
            raise UserError("⚠️ Le fichier Excel ne contient aucune donnée.")
        
        # Analyser les données
        warehouses = set()
        categories = set()
        products_count = 0
        total_quantity = 0
        total_cost = 0
        
        for line_data in lines:
            product_code = str(line_data.get('CODE PRODUIT', '')).strip()
            if product_code:
                products_count += 1
                
            warehouse_name = str(line_data.get('ENTREPOT', '') or line_data.get('EMPLACEMENT', '')).strip()
            if warehouse_name:
                warehouses.add(warehouse_name)
            
            category_name = str(line_data.get('CATEGORIE', '')).strip()
            if category_name:
                categories.add(category_name)
            
            # Convertir en float avec gestion des valeurs vides/nulles
            try:
                quantity_str = line_data.get('QUANTITE', '')
                quantity = float(quantity_str) if quantity_str and str(quantity_str).strip() else 0.0
            except (ValueError, TypeError):
                quantity = 0.0
            
            try:
                price_str = line_data.get('PRIX UNITAIRE', '')
                price = float(price_str) if price_str and str(price_str).strip() else 0.0
            except (ValueError, TypeError):
                price = 0.0
            
            total_quantity += quantity
            total_cost += quantity * price
        
        # Générer le résumé HTML
        summary_html = f"""
        <div style="font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; max-width: 900px; margin: 0 auto;">
            <!-- En-tête -->
            <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 30px; border-radius: 12px 12px 0 0; text-align: center; box-shadow: 0 4px 6px rgba(0,0,0,0.1);">
                <h2 style="margin: 0 0 10px 0; font-size: 28px; font-weight: 600;">
                    <span style="font-size: 32px; margin-right: 10px;">📊</span>
                    Prévisualisation de l'Import
                </h2>
                <p style="margin: 0; font-size: 14px; opacity: 0.9;">Vérifiez les données avant de créer l'inventaire</p>
            </div>
            
            <!-- Statistiques principales -->
            <div style="display: grid; grid-template-columns: repeat(4, 1fr); gap: 20px; padding: 30px; background: #f8f9fa;">
                <div style="background: white; padding: 20px; border-radius: 10px; box-shadow: 0 2px 8px rgba(0,0,0,0.08); border-left: 4px solid #667eea;">
                    <div style="font-size: 32px; color: #667eea; margin-bottom: 8px;">📄</div>
                    <div style="font-size: 28px; font-weight: bold; color: #2d3748; margin-bottom: 4px;">{len(lines):,}</div>
                    <div style="font-size: 11px; color: #718096; text-transform: uppercase; letter-spacing: 0.5px;">Lignes totales</div>
                </div>
                
                <div style="background: white; padding: 20px; border-radius: 10px; box-shadow: 0 2px 8px rgba(0,0,0,0.08); border-left: 4px solid #48bb78;">
                    <div style="font-size: 32px; color: #48bb78; margin-bottom: 8px;">📦</div>
                    <div style="font-size: 28px; font-weight: bold; color: #2d3748; margin-bottom: 4px;">{products_count:,}</div>
                    <div style="font-size: 11px; color: #718096; text-transform: uppercase; letter-spacing: 0.5px;">Produits distincts</div>
                </div>
                
                <div style="background: white; padding: 20px; border-radius: 10px; box-shadow: 0 2px 8px rgba(0,0,0,0.08); border-left: 4px solid #ed8936;">
                    <div style="font-size: 32px; color: #ed8936; margin-bottom: 8px;">📊</div>
                    <div style="font-size: 28px; font-weight: bold; color: #2d3748; margin-bottom: 4px;">{total_quantity:,.0f}</div>
                    <div style="font-size: 11px; color: #718096; text-transform: uppercase; letter-spacing: 0.5px;">Quantité totale</div>
                </div>
                
                <div style="background: white; padding: 20px; border-radius: 10px; box-shadow: 0 2px 8px rgba(0,0,0,0.08); border-left: 4px solid #9f7aea;">
                    <div style="font-size: 32px; color: #9f7aea; margin-bottom: 8px;">💰</div>
                    <div style="font-size: 24px; font-weight: bold; color: #2d3748; margin-bottom: 4px;">{total_cost:,.0f}</div>
                    <div style="font-size: 11px; color: #718096; text-transform: uppercase; letter-spacing: 0.5px;">Coût total (FCFA)</div>
                </div>
            </div>
            
            <!-- Détails des entrepôts et catégories -->
            <div style="padding: 0 30px 30px 30px; background: #f8f9fa;">
                <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 20px;">
                    <!-- Entrepôts -->
                    <div style="background: white; padding: 25px; border-radius: 10px; box-shadow: 0 2px 8px rgba(0,0,0,0.08);">
                        <div style="display: flex; align-items: center; margin-bottom: 15px;">
                            <span style="font-size: 24px; margin-right: 10px;">🏭</span>
                            <div>
                                <div style="font-size: 18px; font-weight: 600; color: #2d3748;">Entrepôts</div>
                                <div style="font-size: 24px; font-weight: bold; color: #667eea;">{len(warehouses)}</div>
                            </div>
                        </div>
                        <div style="max-height: 200px; overflow-y: auto;">
                            {''.join([f'<div style="padding: 8px 12px; margin: 5px 0; background: #f7fafc; border-radius: 6px; font-size: 13px; color: #4a5568; border-left: 3px solid #667eea;">{wh}</div>' for wh in list(warehouses)[:10]])}
                            {f'<div style="padding: 8px; text-align: center; color: #a0aec0; font-size: 12px; font-style: italic;">... et {len(warehouses) - 10} autre(s)</div>' if len(warehouses) > 10 else ''}
                        </div>
                    </div>
                    
                    <!-- Catégories -->
                    <div style="background: white; padding: 25px; border-radius: 10px; box-shadow: 0 2px 8px rgba(0,0,0,0.08);">
                        <div style="display: flex; align-items: center; margin-bottom: 15px;">
                            <span style="font-size: 24px; margin-right: 10px;">📁</span>
                            <div>
                                <div style="font-size: 18px; font-weight: 600; color: #2d3748;">Catégories</div>
                                <div style="font-size: 24px; font-weight: bold; color: #48bb78;">{len(categories)}</div>
                            </div>
                        </div>
                        <div style="max-height: 200px; overflow-y: auto;">
                            {''.join([f'<div style="padding: 8px 12px; margin: 5px 0; background: #f0fff4; border-radius: 6px; font-size: 13px; color: #4a5568; border-left: 3px solid #48bb78;">{cat}</div>' for cat in list(categories)[:10]])}
                            {f'<div style="padding: 8px; text-align: center; color: #a0aec0; font-size: 12px; font-style: italic;">... et {len(categories) - 10} autre(s)</div>' if len(categories) > 10 else ''}
                        </div>
                    </div>
                </div>
            </div>
            
            <!-- Aperçu des premières lignes -->
            <div style="padding: 0 30px 30px 30px; background: #f8f9fa;">
                <div style="background: white; padding: 25px; border-radius: 10px; box-shadow: 0 2px 8px rgba(0,0,0,0.08);">
                    <div style="display: flex; align-items: center; margin-bottom: 20px;">
                        <span style="font-size: 24px; margin-right: 10px;">👁️</span>
                        <div style="font-size: 18px; font-weight: 600; color: #2d3748;">Aperçu des 5 premières lignes</div>
                    </div>
                    <div style="overflow-x: auto;">
                        <table style="width: 100%; border-collapse: collapse; font-size: 13px;">
                            <thead>
                                <tr style="background: #f7fafc; border-bottom: 2px solid #e2e8f0;">
                                    <th style="padding: 12px; text-align: left; font-weight: 600; color: #4a5568;">#</th>
                                    <th style="padding: 12px; text-align: left; font-weight: 600; color: #4a5568;">Code</th>
                                    <th style="padding: 12px; text-align: left; font-weight: 600; color: #4a5568;">Produit</th>
                                    <th style="padding: 12px; text-align: left; font-weight: 600; color: #4a5568;">Entrepôt</th>
                                    <th style="padding: 12px; text-align: right; font-weight: 600; color: #4a5568;">Quantité</th>
                                    <th style="padding: 12px; text-align: right; font-weight: 600; color: #4a5568;">Prix U.</th>
                                </tr>
                            </thead>
                            <tbody>
                                {''.join([f'''<tr style="border-bottom: 1px solid #f7fafc;">
                                    <td style="padding: 10px; color: #a0aec0;">{idx + 1}</td>
                                    <td style="padding: 10px; color: #2d3748; font-family: monospace; font-size: 12px;">{str(line.get('CODE PRODUIT', ''))[:15]}</td>
                                    <td style="padding: 10px; color: #4a5568;">{str(line.get('PRODUIT', '') or line.get('NOM PRODUIT', ''))[:30]}</td>
                                    <td style="padding: 10px;">
                                        <span style="background: #edf2f7; padding: 4px 8px; border-radius: 4px; font-size: 11px; color: #667eea;">{str(line.get('ENTREPOT', '') or line.get('EMPLACEMENT', ''))[:20]}</span>
                                    </td>
                                    <td style="padding: 10px; text-align: right; font-weight: 500; color: #2d3748;">{float(line.get('QUANTITE') or 0) if line.get('QUANTITE') and str(line.get('QUANTITE')).strip() else 0:,.0f}</td>
                                    <td style="padding: 10px; text-align: right; font-weight: 500; color: #2d3748;">{float(line.get('PRIX UNITAIRE') or 0) if line.get('PRIX UNITAIRE') and str(line.get('PRIX UNITAIRE')).strip() else 0:,.0f}</td>
                                </tr>''' for idx, line in enumerate(lines[:5])])}
                            </tbody>
                        </table>
                    </div>
                </div>
            </div>
            
            <!-- Message d'information -->
            <div style="padding: 0 30px 30px 30px; background: #f8f9fa; border-radius: 0 0 12px 12px;">
                <div style="background: linear-gradient(135deg, #667eea15 0%, #764ba215 100%); padding: 20px; border-radius: 10px; border: 2px solid #667eea30;">
                    <div style="display: flex; align-items: start;">
                        <span style="font-size: 24px; margin-right: 15px;">ℹ️</span>
                        <div>
                            <div style="font-weight: 600; color: #2d3748; margin-bottom: 5px;">Prêt à importer</div>
                            <div style="color: #4a5568; font-size: 14px; line-height: 1.6;">
                                Les données ci-dessus sont une prévisualisation de votre fichier Excel.<br/>
                                Cliquez sur <strong>"✅ Confirmer l'import"</strong> pour créer l'inventaire initial avec ces données.
                            </div>
                        </div>
                    </div>
                </div>
            </div>
        </div>
        """
        
        # Mettre à jour les champs de prévisualisation
        self.write({
            'state': 'step2',
            'preview_summary': summary_html,
            'lines_count': len(lines),
            'warehouses_preview': ', '.join(list(warehouses)[:5]),
            'categories_preview': ', '.join(list(categories)[:5]),
        })
        
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'stockex.initial.stock.wizard',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }
    
    def action_back(self):
        """Retour à l'étape 1."""
        self.write({'state': 'step1'})
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'stockex.initial.stock.wizard',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }
    
    def _send_notifications(self, created_count, message):
        """Envoie les notifications Email, WhatsApp et Telegram."""
        try:
            # Récupérer les paramètres
            ICP = self.env['ir.config_parameter'].sudo()
            
            # Email
            if ICP.get_param('stockex.notify_by_email', 'False') == 'True':
                email_list = ICP.get_param('stockex.notification_emails', '')
                if email_list:
                    self._send_email(email_list, created_count, message)
            
            # WhatsApp  
            if ICP.get_param('stockex.notify_by_whatsapp', 'False') == 'True':
                numbers_list = ICP.get_param('stockex.whatsapp_numbers', '')
                api_url = ICP.get_param('stockex.whatsapp_api_url', '')
                if numbers_list and api_url:
                    self._send_whatsapp(numbers_list, api_url, created_count, message)
            
            # Telegram
            if ICP.get_param('stockex.notify_by_telegram', 'False') == 'True':
                bot_token = ICP.get_param('stockex.telegram_bot_token', '')
                chat_ids = ICP.get_param('stockex.telegram_chat_ids', '')
                if bot_token and chat_ids:
                    self._send_telegram(bot_token, chat_ids, created_count, message)
                    
        except Exception as e:
            _logger.error(f"❌ Erreur notifications : {str(e)}")
    
    def _send_email(self, email_list, created_count, message):
        """Envoie une notification par email."""
        emails = [email.strip() for email in email_list.split(',') if email.strip()]
        if not emails:
            return
        
        subject = f"✅ Stock Initial Créé - {self.name}"
        body = f"""
        <div style="font-family: Arial, sans-serif;">
            <h2>✅ Stock Initial Créé</h2>
            <p><strong>Nom :</strong> {self.name}</p>
            <p><strong>Date :</strong> {self.date}</p>
            <p><strong>Enregistrements créés :</strong> {created_count}</p>
            <pre>{message}</pre>
        </div>
        """
        
        try:
            mail_values = {
                'subject': subject,
                'body_html': body,
                'email_to': ', '.join(emails),
                'email_from': self.env.user.company_id.email or 'noreply@odoo.com',
            }
            mail = self.env['mail.mail'].sudo().create(mail_values)
            mail.send()
            _logger.info(f"📧 Email envoyé à : {', '.join(emails)}")
        except Exception as e:
            _logger.error(f"❌ Erreur envoi email : {str(e)}")
    
    def _send_whatsapp(self, numbers_list, api_url, created_count, message):
        """Envoie une notification WhatsApp."""
        try:
            import requests
            numbers = [num.strip() for num in numbers_list.split(',') if num.strip()]
            
            whatsapp_message = f"✅ *Stock Initial Créé*\n\n📦 {created_count} enregistrement(s)\n📅 Date : {self.date}"
            
            for number in numbers:
                try:
                    payload = {'to': number, 'message': whatsapp_message}
                    response = requests.post(api_url, json=payload, timeout=10)
                    if response.status_code in [200, 201]:
                        _logger.info(f"💬 WhatsApp envoyé à : {number}")
                except Exception as e:
                    _logger.error(f"❌ Erreur WhatsApp {number} : {str(e)}")
        except ImportError:
            _logger.error("❌ Module 'requests' non installé")
        except Exception as e:
            _logger.error(f"❌ Erreur WhatsApp : {str(e)}")
    
    def _send_telegram(self, bot_token, chat_ids_list, created_count, message):
        """Envoie une notification Telegram."""
        try:
            import requests
            chat_ids = [cid.strip() for cid in chat_ids_list.split(',') if cid.strip()]
            
            telegram_message = f"✅ *Stock Initial Créé*\n\n📦 {created_count} enregistrement(s)\n📅 Date : {self.date}"
            api_url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
            
            for chat_id in chat_ids:
                try:
                    payload = {'chat_id': chat_id, 'text': telegram_message, 'parse_mode': 'Markdown'}
                    response = requests.post(api_url, json=payload, timeout=10)
                    if response.status_code == 200:
                        _logger.info(f"📱 Telegram envoyé à : {chat_id}")
                except Exception as e:
                    _logger.error(f"❌ Erreur Telegram {chat_id} : {str(e)}")
        except ImportError:
            _logger.error("❌ Module 'requests' non installé")
        except Exception as e:
            _logger.error(f"❌ Erreur Telegram : {str(e)}")
    
    def _reset_all_stocks(self):
        """Transfère le stock existant vers un emplacement technique 'Ancien stock ENEO'."""
        self.ensure_one()
        
        _logger.warning("🔄 RÉINITIALISATION DES STOCKS : Transfert vers 'Ancien stock ENEO'")
        
        # 1. Créer ou récupérer l'emplacement technique "Ancien stock ENEO"
        old_stock_location = self.env['stock.location'].search([
            ('name', '=', 'Ancien stock ENEO'),
            ('usage', '=', 'inventory'),
            ('company_id', '=', self.company_id.id)
        ], limit=1)
        
        if not old_stock_location:
            old_stock_location = self.env['stock.location'].create({
                'name': 'Ancien stock ENEO',
                'usage': 'inventory',
                'company_id': self.company_id.id,
                'location_id': self.env.ref('stock.stock_location_locations').id,
            })
            _logger.info(f"✅ Emplacement technique créé: {old_stock_location.name}")
        
        # 2. Récupérer tous les quants avec stock > 0
        stock_quants = self.env['stock.quant'].search([
            ('quantity', '>', 0),
            ('location_id.usage', '=', 'internal'),
            ('company_id', '=', self.company_id.id)
        ])
        
        quants_count = len(stock_quants)
        _logger.warning(f"📦 {quants_count} quants à transférer vers 'Ancien stock ENEO'")
        
        if quants_count == 0:
            _logger.info("✅ Aucun stock existant à transférer")
            return {
                'quants_transferred': 0,
                'pickings_created': 0
            }
        
        # 3. Créer des pickings de transfert interne par lot
        pickings_created = 0
        picking_type = self.env['stock.picking.type'].search([
            ('code', '=', 'internal'),
            ('company_id', '=', self.company_id.id)
        ], limit=1)
        
        if not picking_type:
            raise UserError("Type de picking interne introuvable. Vérifiez la configuration Odoo.")
        
        # Regrouper par emplacement source pour optimiser
        quants_by_location = {}
        for quant in stock_quants:
            loc_id = quant.location_id.id
            if loc_id not in quants_by_location:
                quants_by_location[loc_id] = []
            quants_by_location[loc_id].append(quant)
        
        # Créer un picking par emplacement source
        for location_id, quants_group in quants_by_location.items():
            location = self.env['stock.location'].browse(location_id)
            
            picking_vals = {
                'picking_type_id': picking_type.id,
                'location_id': location.id,
                'location_dest_id': old_stock_location.id,
                'origin': f'Réinitialisation stock {self.date}',
                'company_id': self.company_id.id,
            }
            
            picking = self.env['stock.picking'].create(picking_vals)
            
            # Créer les mouvements pour chaque quant
            for quant in quants_group:
                move_vals = {
                    'name': f'Transfert ancien stock: {quant.product_id.name}',
                    'product_id': quant.product_id.id,
                    'product_uom_qty': quant.quantity,
                    'product_uom': quant.product_id.uom_id.id,
                    'picking_id': picking.id,
                    'location_id': location.id,
                    'location_dest_id': old_stock_location.id,
                    'company_id': self.company_id.id,
                }
                self.env['stock.move'].create(move_vals)
            
            # Confirmer et valider le picking
            picking.action_confirm()
            picking.action_assign()
            
            # Valider avec les quantités disponibles
            for move in picking.move_ids:
                move.quantity = move.product_uom_qty
            
            picking.button_validate()
            pickings_created += 1
            
            _logger.info(f"✅ Picking {picking.name} créé et validé: {len(quants_group)} produits transférés")
        
        # Commit pour libérer la mémoire
        self.env.cr.commit()
        
        _logger.warning(f"✅ Réinitialisation terminée: {quants_count} quants transférés via {pickings_created} pickings")
        
        return {
            'quants_transferred': quants_count,
            'pickings_created': pickings_created
        }

    def action_create_initial_stock(self):
        """Crée le stock initial : via mouvements (Option A) ou quants directs (ancien mode)."""
        self.ensure_one()
        
        # Si réinitialisation forcée demandée
        if self.force_reset:
            existing_quants = self.env['stock.quant'].search([
                ('company_id', '=', self.company_id.id),
            ])
            existing_moves = self.env['stock.move'].search([
                ('company_id', '=', self.company_id.id),
            ])
            
            if existing_quants or existing_moves:
                _logger.warning(
                    f"⚠️ RÉINITIALISATION FORCÉE : "
                    f"{len(existing_quants)} quants et {len(existing_moves)} mouvements seront transférés"
                )
                
                reset_result = self._reset_all_stocks()
                
                _logger.warning(
                    f"✅ Réinitialisation terminée : "
                    f"{reset_result['quants_transferred']} quants transférés via "
                    f"{reset_result['pickings_created']} pickings"
                )
        else:
            # Vérifier qu'il n'y a pas déjà de stock
            existing_quants = self.env['stock.quant'].search([
                ('quantity', '>', 0),
                ('location_id.usage', '=', 'internal'),
                ('company_id', '=', self.company_id.id),
            ])
            
            if existing_quants and not self.incremental_update:
                raise UserError(
                    f"⚠️ ATTENTION : {len(existing_quants)} enregistrement(s) de stock déjà existant(s).\n\n"
                    f"Cette fonction est destinée aux bases de données VIDES.\n\n"
                    f"Options :\n"
                    f"1. Cochez '🔄 Mise à jour incrémentale' pour continuer l'import et compléter les données\n"
                    f"2. Utilisez un inventaire normal pour mettre à jour le stock existant\n"
                    f"3. Cochez '⚠️ Réinitialiser tous les stocks' pour forcer une réinitialisation complète"
                )
            
            # Si mode incrémental, afficher un message d'information
            if existing_quants and self.incremental_update:
                _logger.warning(
                    f"🔄 MODE INCRÉMENTAL : {len(existing_quants)} quants existants détectés. "
                    f"L'import continuera en mode mise à jour."
                )
        
        if not self.import_file:
            raise UserError(
                "⚠️ Veuillez fournir un fichier Excel pour le stock initial.\n\n"
                "Le stock initial nécessite des données d'import."
            )
        
        # Parser les lignes
        lines = self._parse_excel_file()
        
        # La méthode _create_initial_stock_quants gère automatiquement
        # l'option create_stock_moves en interne
        _logger.info(f"📦 MODE: {'Avec mouvements de stock' if self.create_stock_moves else 'Quants directs'}")
        created_count = self._create_initial_stock_quants(lines)
        
        if created_count == 0:
            raise UserError(
                "⚠️ Aucun stock n'a pu être créé.\n\n"
                "Vérifiez que votre fichier Excel contient des données valides.\n"
                "Consultez les logs Odoo pour plus de détails."
            )
        
        # Créer un message récapitulatif
        message = f"✅ Stock initial créé avec succès !\n\n"
        message += f"• {created_count} enregistrement(s) de stock créé(s)\n"
        if self.force_reset:
            message += f"• Stocks préalablement réinitialisés (transférés vers 'Ancien stock ENEO')\n"
        if self.create_stock_moves:
            message += f"• Mode : Pickings/Mouvements Odoo (valorisation complète)\n"
        else:
            message += f"• Mode : Quants directs (rapide)\n"
        message += f"• Date : {self.date}\n"
        
        # Envoyer les notifications
        try:
            self._send_notifications(created_count, message)
        except Exception as notif_error:
            _logger.error(f"❌ Erreur lors de l'envoi des notifications : {str(notif_error)}")
        
        # Afficher un message de succès dans les logs et via notification
        _logger.info(message)
        
        # Retourner une action qui affiche un message ET ferme le wizard
        return {
            'type': 'ir.actions.act_window_close',
            'infos': {
                'type': 'success',
                'title': '✅ Stock Initial Créé avec Succès',
                'message': message,
            }
        }
    
    def _parse_excel_file(self):
        """Parse le fichier Excel du stock initial."""
        try:
            from openpyxl import load_workbook
            import base64
            from io import BytesIO
        except ImportError:
            raise UserError(
                "Le module 'openpyxl' n'est pas installé.\n"
                "Installez-le avec: pip install openpyxl"
            )
        
        file_content = base64.b64decode(self.import_file)
        excel_file = BytesIO(file_content)
        wb = load_workbook(excel_file, read_only=True, data_only=True)
        
        # Charger la feuille 'Stock Initial' en priorité, sinon la première feuille active
        if 'Stock Initial' in wb.sheetnames:
            ws = wb['Stock Initial']
            _logger.info("📄 Chargement de la feuille 'Stock Initial'")
        else:
            ws = wb.active
            _logger.info(f"📄 Chargement de la feuille active '{ws.title}'")
        
        # Lire les en-têtes
        headers = [cell.value for cell in ws[1]]
        _logger.info(f"📄 En-têtes Excel: {headers}")
        
        # Lire les données
        lines = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            
            row_dict = {}
            for idx, value in enumerate(row):
                if idx < len(headers):
                    row_dict[headers[idx]] = value if value is not None else ''
            
            lines.append(row_dict)
        
        wb.close()
        _logger.info(f"📄 {len(lines)} lignes lues depuis Excel")
        _logger.info(f"📄 En-têtes détectés: {headers}")
        if len(lines) > 0:
            _logger.info(f"📄 Première ligne: {lines[0]}")
            _logger.info(f"📄 Clés disponibles: {list(lines[0].keys())}")
        else:
            _logger.warning("⚠️ Aucune ligne de données trouvée dans le fichier Excel")
        return lines
    
    def _get_or_create_warehouse(self, warehouse_name):
        """
Recherche ou crée un entrepôt.
        """
        if not warehouse_name:
            # Utiliser l'entrepôt par défaut du formulaire ou le premier
            return self.location_id if self.location_id else self.env['stock.warehouse'].search([], limit=1)
        
        warehouse_name = str(warehouse_name).strip()
        
        try:
            # Savepoint pour isoler cette opération
            savepoint_name = f"warehouse_{id(self)}_{hash(warehouse_name)}"
            self.env.cr.execute(f'SAVEPOINT "{savepoint_name}"')
            
            # Rechercher par nom
            warehouse = self.env['stock.warehouse'].search([
                ('name', '=', warehouse_name),
                ('company_id', 'in', [self.company_id.id, False])
            ], limit=1)
            
            # Si trouvé, libérer le savepoint et retourner
            if warehouse:
                self.env.cr.execute(f'RELEASE SAVEPOINT "{savepoint_name}"')
                return warehouse
            
            # Si pas trouvé et qu'on autorise la création
            if self.create_warehouses:
                # Créer un code unique pour l'entrepôt
                # Stratégie de génération de code intelligente
                import re
                
                # Nettoyer et prendre les éléments significatifs
                parts = re.findall(r'[A-Z0-9]+', warehouse_name.upper())
                
                if len(parts) >= 2:
                    # Si plusieurs parties, combiner intelligemment
                    # Ex: BASSA-KITS-COMP → BASSAKITSCO ou BASSA-KITS
                    code_base = ''.join(parts)[:16]
                else:
                    # Sinon prendre le nom nettoyé
                    code_base = re.sub(r'[^A-Z0-9]', '', warehouse_name.upper())[:16]
                
                # Si le code est trop court, le compléter
                if len(code_base) < 3:
                    code_base = (code_base + str(abs(hash(warehouse_name)))[:8]).ljust(5, 'X')
                
                # Chercher un code disponible avec stratégie améliorée
                counter = 0
                test_code = code_base[:5]  # Commencer avec 5 caractères (minimum Odoo)
                
                while counter < 100:
                    existing = self.env['stock.warehouse'].search([
                        ('code', '=', test_code),
                        ('company_id', '=', self.company_id.id)
                    ], limit=1)
                    
                    if not existing:
                        break
                    
                    counter += 1
                    
                    # Stratégies progressives pour éviter les doublons
                    if counter == 1 and len(code_base) > 5:
                        # Essayer avec le code complet (jusqu'à 16 caractères)
                        test_code = code_base[:min(16, len(code_base))]
                    elif counter < 10:
                        # Ajouter un suffixe numérique court
                        suffix_len = len(str(counter))
                        test_code = f"{code_base[:min(16-suffix_len, len(code_base))]}{counter}"
                    else:
                        # Utiliser un hash unique
                        unique_hash = str(abs(hash(f"{warehouse_name}{counter}")))[:3]
                        test_code = f"{code_base[:min(13, len(code_base))]}{unique_hash}"
                
                # Si toujours pas trouvé après 100 tentatives
                if counter >= 100:
                    import random
                    # Dernier fallback: suffixe aléatoire 3 caractères et quelques essais supplémentaires
                    for _ in range(10):
                        rand = ''.join(random.choice('ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789') for _ in range(3))
                        test_code = f"{code_base[:min(13, len(code_base))]}{rand}"
                        existing = self.env['stock.warehouse'].search([
                            ('code', '=', test_code),
                            ('company_id', '=', self.company_id.id)
                        ], limit=1)
                        if not existing:
                            break
                    else:
                        _logger.error(f"❌ Impossible de générer un code unique pour '{warehouse_name}' après 100 tentatives")
                        self.env.cr.execute(f'ROLLBACK TO SAVEPOINT "{savepoint_name}"')
                        return None
                
                warehouse_vals = {
                    'name': warehouse_name,
                    'code': test_code,
                    'company_id': self.company_id.id,
                }
                
                warehouse = self.env['stock.warehouse'].create(warehouse_vals)
                _logger.info(f"✅ Entrepôt créé: {warehouse.name} (Code: {warehouse.code})")
            
            # Libérer le savepoint
            self.env.cr.execute(f'RELEASE SAVEPOINT "{savepoint_name}"')
            return warehouse
            
        except Exception as e:
            _logger.error(f"❌ Erreur lors de la gestion de l'entrepôt '{warehouse_name}': {str(e)}")
            # Rollback au savepoint uniquement
            try:
                self.env.cr.execute(f'ROLLBACK TO SAVEPOINT "{savepoint_name}"')
            except:
                pass
            return None
    
    def _get_or_create_category(self, category_name, category_code=None):
        """Recherche ou crée une catégorie de produit."""
        if not category_name:
            # Retourner la catégorie par défaut 'All'
            return self.env.ref('product.product_category_all', raise_if_not_found=False)
        
        category_name = str(category_name).strip()
        
        # Rechercher par nom d'abord
        category = self.env['product.category'].search([
            ('name', '=', category_name)
        ], limit=1)
        
        # Si pas trouvée et qu'on autorise la création
        if not category and self.create_categories:
            category_vals = {
                'name': category_name,
            }
            
            # Ajouter le code de catégorie si fourni
            if category_code:
                category_code = str(category_code).strip()
                if category_code:
                    # Vérifier si le code existe déjà
                    existing_with_code = self.env['product.category'].search([
                        ('property_account_income_categ_id', '=', category_code)
                    ], limit=1)
                    
                    if not existing_with_code:
                        # Stockage du code dans le champ de référence
                        category_vals['complete_name'] = f"[{category_code}] {category_name}"
            
            category = self.env['product.category'].create(category_vals)
            _logger.info(f"✅ Catégorie créée: {category.name} (Code: {category_code or 'N/A'})")
        
        return category
    
    def _get_uom_by_name(self, uom_name):
        """Recherche une unité de mesure par son nom.
        
        Supporte les noms courants en anglais et français.
        Retourne l'unité trouvée ou None.
        """
        if not uom_name:
            return None
        
        uom_name_clean = str(uom_name).strip().upper()
        
        # Mapping des noms courants vers les noms Odoo standards
        uom_mapping = {
            'PC': 'Units',      # Pièce
            'PIECE': 'Units',
            'UNIT': 'Units',
            'UNITS': 'Units',
            'U': 'Units',
            'PCS': 'Units',
            'KG': 'kg',         # Kilogramme
            'KILOGRAMME': 'kg',
            'KILO': 'kg',
            'G': 'g',           # Gramme
            'GRAMME': 'g',
            'L': 'L',           # Litre
            'LITRE': 'L',
            'M': 'm',           # Mètre
            'METRE': 'm',
            'METER': 'm',
            'CM': 'cm',         # Centimètre
            'M2': 'm²',        # Mètre carré
            'M3': 'm³',        # Mètre cube
            'H': 'Hour(s)',     # Heure
            'HEURE': 'Hour(s)',
            'HOUR': 'Hour(s)',
            'DAY': 'Day(s)',    # Jour
            'JOUR': 'Day(s)',
            'BOX': 'Units',     # Boîte (traitée comme unité)
            'BOITE': 'Units',
            'CARTON': 'Units',
        }
        
        # Utiliser le mapping si disponible
        odoo_name = uom_mapping.get(uom_name_clean, uom_name_clean)
        
        # Rechercher l'unité dans Odoo
        # 1. Recherche exacte par nom
        uom = self.env['uom.uom'].search([
            ('name', '=', odoo_name)
        ], limit=1)
        
        if uom:
            return uom
        
        # 2. Recherche insensible à la casse
        uom = self.env['uom.uom'].search([
            ('name', 'ilike', odoo_name)
        ], limit=1)
        
        if uom:
            return uom
        
        # 3. Recherche par nom original
        uom = self.env['uom.uom'].search([
            '|',
            ('name', '=', uom_name),
            ('name', 'ilike', uom_name)
        ], limit=1)
        
        if not uom:
            _logger.warning(f"⚠️ Unité de mesure '{uom_name}' non trouvée, utilisation de l'unité par défaut")
        
        return uom
    
    def _create_initial_stock_quants(self, lines):
        """Crée les quants de stock initial directement (sans inventaire)."""
        created_count = 0
        updated_count = 0
        errors = []
        created_categories = set()
        created_warehouses = set()
        created_products = set()
        total_lines = len(lines)
        
        # Cache pour améliorer les performances
        warehouse_cache = {}
        category_cache = {}
        product_cache = {}
        
        _logger.info(f"🔍 Début création stock initial: {total_lines} lignes à traiter")
        if self.incremental_update:
            _logger.info("🔄 Mode incrémental activé : mise à jour des stocks existants autorisée")
        
        # Récupérer l'emplacement d'inventaire virtuel pour les mouvements
        inventory_loc = None
        if self.create_stock_moves:
            inventory_loc = self.env.ref('stock.location_inventory', raise_if_not_found=False)
            if not inventory_loc:
                # Créer l'emplacement d'inventaire virtuel s'il n'existe pas
                inventory_loc = self.env['stock.location'].create({
                    'name': 'Inventory adjustment',
                    'usage': 'inventory',
                    'company_id': self.company_id.id,
                })
            _logger.info("📦 Mode mouvements de stock activé : les mouvements seront créés")
        
        for i, line_data in enumerate(lines):
            # Commit intermédiaire tous les 500 lignes pour éviter les timeouts
            if i > 0 and i % 500 == 0:
                # Sauvegarder la progression
                progress_percent = (i / total_lines) * 100
                _logger.info(f"💾 Commit intermédiaire: {i}/{total_lines} ({progress_percent:.1f}%) - {created_count} stocks créés")
                # Commit pour libérer la mémoire
                self.env.cr.commit()
                _logger.info(f"✅ Commit réussi, continuation de l'import...")
            
            # Afficher la progression (tous les 100 lignes)
            if i % 100 == 0:
                progress_percent = (i / total_lines) * 100
                _logger.info(f"📊 Progression: {i + 1}/{total_lines} ({progress_percent:.1f}%) - {created_count} stocks créés")
            
            try:
                product_code = str(line_data.get('CODE PRODUIT', '')).strip()
                product_name = str(line_data.get('PRODUIT', '') or line_data.get('NOM PRODUIT', '')).strip()
                category_name = str(line_data.get('CATEGORIE', '')).strip() if line_data.get('CATEGORIE') else None
                category_code = str(line_data.get('CODE CATEGORIE', '')).strip() if line_data.get('CODE CATEGORIE') else None
                warehouse_name = str(line_data.get('ENTREPOT', '') or line_data.get('EMPLACEMENT', '')).strip() if line_data.get('ENTREPOT') or line_data.get('EMPLACEMENT') else None
                uom_name = str(line_data.get('UDM', '') or line_data.get('UNITE', '') or line_data.get('UM', '')).strip() if (line_data.get('UDM') or line_data.get('UNITE') or line_data.get('UM')) else None
                
                # Convertir en float avec gestion des valeurs vides/nulles
                try:
                    quantity_str = line_data.get('QUANTITE', '')
                    quantity = float(quantity_str) if quantity_str and str(quantity_str).strip() else 0.0
                except (ValueError, TypeError):
                    quantity = 0.0
                
                try:
                    price_str = line_data.get('PRIX UNITAIRE', '')
                    price = float(price_str) if price_str and str(price_str).strip() else 0.0
                except (ValueError, TypeError):
                    price = 0.0
                
                if not product_code:
                    _logger.warning(f"⚠️ Ligne {i+2}: CODE PRODUIT vide, ignorée")
                    errors.append(f"Ligne {i+2}: CODE PRODUIT vide")
                    continue
                
                if quantity <= 0:
                    _logger.warning(f"⚠️ Ligne {i+2}: Quantité nulle ou négative ({quantity}), ignorée")
                    errors.append(f"Ligne {i+2}: Quantité invalide ({quantity})")
                    continue
                
                # Gérer l'entrepôt (avec cache)
                if warehouse_name in warehouse_cache:
                    warehouse = warehouse_cache[warehouse_name]
                else:
                    warehouse = self._get_or_create_warehouse(warehouse_name)
                    if warehouse:
                        warehouse_cache[warehouse_name] = warehouse
                
                if not warehouse:
                    errors.append(f"Ligne {i+2}: Entrepôt '{warehouse_name}' non trouvé et création désactivée")
                    continue
                
                if warehouse.name not in created_warehouses:
                    created_warehouses.add(warehouse.name)
                
                # Utiliser l'emplacement stock de l'entrepôt
                location = warehouse.lot_stock_id
                
                # Vérifier que l'emplacement est de type internal
                if not location or location.usage != 'internal':
                    _logger.error(f"❌ Ligne {i+2}: Emplacement invalide (type: {location.usage if location else 'None'})")
                    errors.append(f"Ligne {i+2}: Emplacement invalide pour entrepôt '{warehouse.name}'")
                    continue
                
                # Gérer la catégorie (avec cache)
                category = None
                if category_name:
                    cache_key = f"{category_name}|{category_code}"
                    if cache_key in category_cache:
                        category = category_cache[cache_key]
                    else:
                        category = self._get_or_create_category(category_name, category_code)
                        if category:
                            category_cache[cache_key] = category
                    
                    if category and category.name not in created_categories:
                        created_categories.add(category.name)
                
                # Rechercher ou créer le produit (avec cache)
                if product_code in product_cache:
                    product = product_cache[product_code]
                else:
                    product = self.env['product.product'].search([
                        ('default_code', '=', product_code)
                    ], limit=1)
                    if product:
                        product_cache[product_code] = product
                
                if not product:
                    if not self.create_products:
                        errors.append(f"Ligne {i+2}: Produit '{product_code}' non trouvé et création désactivée")
                        continue
                    
                    # Créer le nouveau produit (sans prix pour éviter les erreurs de valorisation)
                    product_vals = {
                        'name': product_name or product_code,
                        'default_code': product_code,
                        # Le prix sera ajouté après création (ligne 1062-1065)
                    }
                    # Déterminer le champ de type selon la version (detailed_type vs type)
                    # detailed_type volontairement non défini pour éviter les erreurs de validation
                    
                    if category:
                        product_vals['categ_id'] = category.id
                    
                    # Gérer l'unité de mesure si spécifiée
                    if uom_name:
                        uom = self._get_uom_by_name(uom_name)
                        if uom:
                            product_vals['uom_id'] = uom.id
                    
                    # Créer le template produit, puis récupérer la variante
                    tmpl_vals = dict(product_vals)
                    # default_code appartient à la variante, pas au template
                    default_code_val = tmpl_vals.pop('default_code', None)
                    try:
                        tmpl = self.env['product.template'].create(tmpl_vals)
                    except Exception as e:
                        if 'product.template.type' in str(e):
                            tmpl_vals.pop('detailed_type', None)
                            tmpl_vals.pop('type', None)
                            tmpl = self.env['product.template'].create(tmpl_vals)
                        else:
                            raise
                    product = tmpl.product_variant_id
                    if default_code_val:
                        product.write({'default_code': default_code_val})
                    product_cache[product_code] = product
                    created_products.add(product_code)
                
                elif category:
                    # Mettre à jour la catégorie du produit existant si spécifiée
                    if product.categ_id.id != category.id:
                        product.write({'categ_id': category.id})
                
                # Forcer type produit = consu et suivre l'inventaire
                pp_fields = self.env['product.product']._fields
                try:
                    pp_write = {}
                    if 'type' in pp_fields and getattr(product, 'type', None) != 'consu':
                        pp_write['type'] = 'consu'
                    if 'is_storable' in pp_fields and getattr(product, 'is_storable', None) is False:
                        pp_write['is_storable'] = True
                    if pp_write:
                        product.sudo().write(pp_write)
                except Exception:
                    pass
                
                # Mettre à jour les prix (coûtant ET vente) si fourni
                if price > 0:
                    tmpl_record = product.product_tmpl_id
                    if tmpl_record:
                        tmpl_record.write({
                            'standard_price': price,  # Prix de revient (coût)
                            'list_price': price,       # Prix de vente (affiché)
                        })
                
                # Convertir en stockable si quantité > 0
                tmpl = product.product_tmpl_id
                if quantity and quantity > 0 and tmpl:
                    tmpl_fields = self.env['product.template']._fields
                    try:
                        to_write = {}
                        if 'detailed_type' in tmpl_fields and getattr(tmpl, 'detailed_type', None) != 'product':
                            to_write['detailed_type'] = 'consu'
                        if 'type' in tmpl_fields and getattr(tmpl, 'type', None) != 'product':
                            to_write['type'] = 'consu'
                        if to_write:
                            tmpl.sudo().write(to_write)
                            # Assurer la propagation au variant si nécessaire
                            try:
                                product_fields = self.env['product.product']._fields
                                if 'type' in product_fields and getattr(product, 'type', None) != 'product':
                                    product.sudo().write({'type': 'consu'})
                                if 'is_storable' in product_fields and getattr(product, 'is_storable', None) is False:
                                    product.sudo().write({'is_storable': True})
                            except Exception:
                                pass
                    except Exception:
                        pass
                # Garde: convertir non-stockables ignorés
                tmpl2 = product.product_tmpl_id
                if tmpl2:
                    tmpl_fields2 = self.env['product.template']._fields
                    current_type = getattr(tmpl2, 'detailed_type', None) if 'detailed_type' in tmpl_fields2 else getattr(tmpl2, 'type', None)
                    is_storable_flag = getattr(product, 'is_storable', None)
                    allowed = (current_type == 'product') or (current_type == 'consu' and bool(is_storable_flag))
                    if not allowed:
                        errors.append(f"Ligne {i+2}: Produit '{product_code}' non stockable, ignoré")
                        _logger.error(f"❌ Erreur ligne {i+2}: Produit non stockable (type={current_type}, is_storable={is_storable_flag})")
                        continue
                # Créer ou mettre à jour le quant
                quant = self.env['stock.quant'].search([
                    ('product_id', '=', product.id),
                    ('location_id', '=', location.id),
                    ('company_id', '=', self.company_id.id),
                ], limit=1)
                
                if quant:
                    # Le quant existe déjà, mettre à jour directement la quantité (sans action_apply_inventory)
                    old_quantity = quant.quantity
                    quant.write({'quantity': quantity})
                    updated_count += 1
                    
                    # Créer un mouvement de stock si l'option est activée
                    if self.create_stock_moves and inventory_loc:
                        difference = quantity - old_quantity
                        if difference != 0:
                            self._create_stock_move(product, location, inventory_loc, abs(difference), difference > 0)
                else:
                    # Créer un nouveau quant directement (sans action_apply_inventory)
                    quant = self.env['stock.quant'].create({
                        'product_id': product.id,
                        'location_id': location.id,
                        'company_id': self.company_id.id,
                        'quantity': quantity,
                    })
                    
                    # Créer un mouvement de stock si l'option est activée
                    if self.create_stock_moves and inventory_loc and quantity > 0:
                        self._create_stock_move(product, location, inventory_loc, quantity, True)
                
                created_count += 1
                
            except Exception as e:
                errors.append(f"Ligne {i+2}: {str(e)}")
                _logger.error(f"❌ Erreur ligne {i+2}: {str(e)}")
        
        _logger.info(f"✅ Import terminé: {created_count} stock(s) traité(s) ({created_count - updated_count} créés, {updated_count} mis à jour)")
        
        # Progression à 100% - Commit final unique
        self.write({
            'progress': 100.0,
            'progress_message': f'Import terminé : {created_count} stock(s) traité(s)'
        })
        # Un seul commit à la fin pour éviter de fermer le curseur en cours de traitement
        self.env.cr.commit()
        _logger.info("💾 Commit final effectué")
        
        # Message récapitulatif dans les logs
        message = f"✅ {created_count} stock(s) traité(s)"
        if updated_count > 0:
            message += f" (🔄 {updated_count} mis à jour, ➕ {created_count - updated_count} créés)"
        if created_products:
            message += f"\n📦 {len(created_products)} produit(s) créé(s)"
        if created_warehouses:
            message += f"\n🏭 {len(created_warehouses)} entrepôt(s): {', '.join(sorted(created_warehouses))}"
        if created_categories:
            message += f"\n📁 {len(created_categories)} catégorie(s): {', '.join(sorted(created_categories))}"
        if errors:
            message += f"\n⚠️ {len(errors)} ligne(s) ignorée(s)"
            for error in errors[:20]:
                _logger.warning(f"  - {error}")
        
        _logger.info(message)
        
        return created_count
    
    def _create_stock_move(self, product, location, inventory_loc, quantity, is_incoming):
        """Crée un mouvement de stock pour l'import initial.
        
        Args:
            product: Le produit concerné
            location: L'emplacement de stock
            inventory_loc: L'emplacement d'inventaire virtuel
            quantity: La quantité à déplacer
            is_incoming: True si c'est une entrée de stock, False si c'est une sortie
        """
        try:
            StockMove = self.env['stock.move']
            
            # Créer un stock.move pour l'ajustement
            move_vals = {
                'name': f'Stock Initial - {product.display_name}',
                'product_id': product.id,
                'product_uom': product.uom_id.id,
                'product_uom_qty': quantity,
                'company_id': self.company_id.id,
                'date': self.date or fields.Datetime.now(),
                'origin': f'Stock Initial {self.name}',
                'reference': f'Ajustement stock initial {self.name}',
            }
            
            # Si c'est une entrée : depuis inventory vers location
            # Si c'est une sortie : depuis location vers inventory
            if is_incoming:
                move_vals.update({
                    'location_id': inventory_loc.id,
                    'location_dest_id': location.id,
                })
            else:
                move_vals.update({
                    'location_id': location.id,
                    'location_dest_id': inventory_loc.id,
                })
            
            # Créer et valider le mouvement
            move = StockMove.create(move_vals)
            move._action_confirm()
            move._action_assign()
            move._action_done()
            
            _logger.debug(
                f"📦 Mouvement de stock créé: {quantity} x {product.default_code} "
                f"{'📥' if is_incoming else '📤'} {location.name}"
            )
            
        except Exception as e:
            _logger.error(f"❌ Erreur création mouvement stock pour {product.default_code}: {str(e)}")

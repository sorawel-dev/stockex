# -*- coding: utf-8 -*-

from odoo import models, fields, api
from odoo.exceptions import UserError
import io
import base64
from datetime import datetime


class WarehouseValuationExportWizard(models.TransientModel):
    _name = 'stockex.warehouse.valuation.export.wizard'
    _description = 'Assistant Export Valorisation par Entrepôt'

    # Champ requis pour le widget badge
    color = fields.Integer(string='Couleur', default=0)

    # Filtres
    warehouse_ids = fields.Many2many(
        'stock.warehouse',
        string='Entrepôts',
        help='Laisser vide pour exporter tous les entrepôts'
    )
    date_from = fields.Date(
        string='Date Début',
        help='Filtrer les inventaires à partir de cette date'
    )
    date_to = fields.Date(
        string='Date Fin',
        help='Filtrer les inventaires jusqu\'à cette date'
    )
    category_ids = fields.Many2many(
        'product.category',
        string='Catégories',
        help='Filtrer par catégories de produits'
    )
    
    # Options d'export
    export_type = fields.Selection([
        ('summary', 'Résumé par Entrepôt'),
        ('detailed', 'Détaillé par Produit'),
        ('both', 'Les Deux (Résumé + Détail)')
    ], string='Type d\'Export', default='summary', required=True)
    
    include_vsd = fields.Boolean(
        string='Inclure VSD (Valeur avec Décote)',
        default=True,
        help='Ajouter la colonne Valeur avec Décote selon rotation'
    )
    
    include_ecarts = fields.Boolean(
        string='Inclure les Écarts',
        default=True,
        help='Ajouter les colonnes d\'écarts Inventorié - Stock Réel'
    )
    
    # Fichier généré
    export_file = fields.Binary(
        string='Fichier Excel',
        readonly=True
    )
    export_filename = fields.Char(
        string='Nom du Fichier',
        readonly=True
    )
    state = fields.Selection([
        ('draft', 'Configuration'),
        ('done', 'Terminé')
    ], default='draft')

    @api.model
    def default_get(self, fields_list):
        """Initialiser avec les filtres du dashboard si disponibles."""
        res = super().default_get(fields_list)
        
        # Récupérer le contexte du dashboard
        if self.env.context.get('active_model') == 'stockex.inventory.summary':
            dashboard = self.env['stockex.inventory.summary'].browse(
                self.env.context.get('active_id')
            )
            if dashboard:
                if dashboard.warehouse_ids:
                    res['warehouse_ids'] = [(6, 0, dashboard.warehouse_ids.ids)]
                if dashboard.date_from:
                    res['date_from'] = dashboard.date_from
                if dashboard.date_to:
                    res['date_to'] = dashboard.date_to
                if dashboard.category_ids:
                    res['category_ids'] = [(6, 0, dashboard.category_ids.ids)]
        
        return res

    def action_export(self):
        """Générer le fichier Excel."""
        self.ensure_one()
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise UserError(
                "Le module openpyxl n'est pas installé. "
                "Veuillez l'installer avec : pip install openpyxl"
            )
        
        # Créer le workbook
        wb = Workbook()
        wb.remove(wb.active)  # Supprimer la feuille par défaut
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        
        # Récupérer les données
        warehouses = self.warehouse_ids if self.warehouse_ids else self.env['stock.warehouse'].search([])
        
        # Export résumé
        if self.export_type in ['summary', 'both']:
            ws_summary = wb.create_sheet("Résumé par Entrepôt")
            self._export_summary_sheet(ws_summary, warehouses, header_font, header_fill, total_fill, border, center_align, right_align)
        
        # Export détaillé
        if self.export_type in ['detailed', 'both']:
            for warehouse in warehouses:
                ws_detail = wb.create_sheet(warehouse.name[:31])  # Limite Excel: 31 caractères
                self._export_detailed_sheet(ws_detail, warehouse, header_font, header_fill, border, center_align, right_align)
        
        # Sauvegarder
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Nom du fichier
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"Valorisation_Entrepots_{timestamp}.xlsx"
        
        # Mettre à jour le wizard
        self.write({
            'export_file': base64.b64encode(output.read()),
            'export_filename': filename,
            'state': 'done'
        })
        
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'view_mode': 'form',
            'res_id': self.id,
            'views': [(False, 'form')],
            'target': 'new',
        }

    def _export_summary_sheet(self, ws, warehouses, header_font, header_fill, total_fill, border, center_align, right_align):
        """Exporter la feuille résumé."""
        # En-tête
        headers = ['Entrepôt', 'Quantité Inventoriée', 'Valeur Inventoriée', 'Stock Initial']
        col_idx = 1
        
        if self.include_vsd:
            headers.append('VSD (avec Décote)')
            col_idx += 1
        
        headers.extend(['% du Total'])
        
        if self.include_ecarts:
            headers.append('Écart (Inventorié - Initial)')
        
        # Écrire les en-têtes
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        
        # Données
        dashboard = self.env['stockex.inventory.summary'].search([], limit=1)
        if not dashboard:
            dashboard = self.env['stockex.inventory.summary'].create({})
        
        # Récupérer les données par entrepôt
        row = 2
        total_qty = 0
        total_value = 0
        total_initial = 0
        total_vsd = 0
        
        for warehouse in warehouses:
            # Calculer les valeurs pour cet entrepôt
            inv_data = self._get_warehouse_data(warehouse, dashboard)
            
            col = 1
            ws.cell(row=row, column=col, value=warehouse.name).border = border
            col += 1
            ws.cell(row=row, column=col, value=inv_data['qty']).border = border
            ws.cell(row=row, column=col).number_format = '#,##0'
            ws.cell(row=row, column=col).alignment = right_align
            col += 1
            ws.cell(row=row, column=col, value=inv_data['value']).border = border
            ws.cell(row=row, column=col).number_format = '#,##0'
            ws.cell(row=row, column=col).alignment = right_align
            col += 1
            ws.cell(row=row, column=col, value=inv_data['initial']).border = border
            ws.cell(row=row, column=col).number_format = '#,##0'
            ws.cell(row=row, column=col).alignment = right_align
            col += 1
            
            if self.include_vsd:
                ws.cell(row=row, column=col, value=inv_data['vsd']).border = border
                ws.cell(row=row, column=col).number_format = '#,##0'
                ws.cell(row=row, column=col).alignment = right_align
                col += 1
            
            percent = (inv_data['value'] / total_value * 100) if total_value > 0 else 0
            ws.cell(row=row, column=col, value=percent).border = border
            ws.cell(row=row, column=col).number_format = '0.00'
            ws.cell(row=row, column=col).alignment = right_align
            col += 1
            
            if self.include_ecarts:
                ecart = inv_data['value'] - inv_data['initial']
                ws.cell(row=row, column=col, value=ecart).border = border
                ws.cell(row=row, column=col).number_format = '#,##0'
                ws.cell(row=row, column=col).alignment = right_align
            
            total_qty += inv_data['qty']
            total_value += inv_data['value']
            total_initial += inv_data['initial']
            total_vsd += inv_data['vsd']
            row += 1
        
        # Ligne Total
        col = 1
        cell = ws.cell(row=row, column=col, value='TOTAL')
        cell.font = Font(bold=True)
        cell.fill = total_fill
        cell.border = border
        col += 1
        
        cell = ws.cell(row=row, column=col, value=total_qty)
        cell.font = Font(bold=True)
        cell.fill = total_fill
        cell.border = border
        cell.number_format = '#,##0'
        cell.alignment = right_align
        col += 1
        
        cell = ws.cell(row=row, column=col, value=total_value)
        cell.font = Font(bold=True)
        cell.fill = total_fill
        cell.border = border
        cell.number_format = '#,##0'
        cell.alignment = right_align
        col += 1
        
        cell = ws.cell(row=row, column=col, value=total_initial)
        cell.font = Font(bold=True)
        cell.fill = total_fill
        cell.border = border
        cell.number_format = '#,##0'
        cell.alignment = right_align
        col += 1
        
        if self.include_vsd:
            cell = ws.cell(row=row, column=col, value=total_vsd)
            cell.font = Font(bold=True)
            cell.fill = total_fill
            cell.border = border
            cell.number_format = '#,##0'
            cell.alignment = right_align
            col += 1
        
        cell = ws.cell(row=row, column=col, value=100.00)
        cell.font = Font(bold=True)
        cell.fill = total_fill
        cell.border = border
        cell.number_format = '0.00'
        cell.alignment = right_align
        col += 1
        
        if self.include_ecarts:
            total_ecart = total_value - total_initial
            cell = ws.cell(row=row, column=col, value=total_ecart)
            cell.font = Font(bold=True)
            cell.fill = total_fill
            cell.border = border
            cell.number_format = '#,##0'
            cell.alignment = right_align
        
        # Ajuster les largeurs de colonnes
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

    def _export_detailed_sheet(self, ws, warehouse, header_font, header_fill, border, center_align, right_align):
        """Exporter la feuille détaillée pour un entrepôt."""
        # En-tête
        headers = ['Produit', 'Référence', 'Catégorie', 'Quantité', 'Prix Unitaire', 'Valeur Totale']
        
        if self.include_vsd:
            headers.append('VSD Unitaire')
            headers.append('VSD Total')
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        
        # Récupérer les lignes d'inventaire pour cet entrepôt
        inv_lines = self._get_warehouse_inventory_lines(warehouse)
        
        row = 2
        for line in inv_lines:
            col = 1
            ws.cell(row=row, column=col, value=line.product_id.name).border = border
            col += 1
            ws.cell(row=row, column=col, value=line.product_id.default_code or '').border = border
            col += 1
            ws.cell(row=row, column=col, value=line.product_id.categ_id.name).border = border
            col += 1
            ws.cell(row=row, column=col, value=line.product_qty).border = border
            ws.cell(row=row, column=col).number_format = '#,##0'
            ws.cell(row=row, column=col).alignment = right_align
            col += 1
            
            price = line.standard_price
            ws.cell(row=row, column=col, value=price).border = border
            ws.cell(row=row, column=col).number_format = '#,##0'
            ws.cell(row=row, column=col).alignment = right_align
            col += 1
            
            value = line.product_qty * price
            ws.cell(row=row, column=col, value=value).border = border
            ws.cell(row=row, column=col).number_format = '#,##0'
            ws.cell(row=row, column=col).alignment = right_align
            col += 1
            
            if self.include_vsd:
                # Calculer VSD pour ce produit
                vsd_price = price * 0.6  # Placeholder - à adapter selon la logique de décote
                ws.cell(row=row, column=col, value=vsd_price).border = border
                ws.cell(row=row, column=col).number_format = '#,##0'
                ws.cell(row=row, column=col).alignment = right_align
                col += 1
                
                vsd_value = line.product_qty * vsd_price
                ws.cell(row=row, column=col, value=vsd_value).border = border
                ws.cell(row=row, column=col).number_format = '#,##0'
                ws.cell(row=row, column=col).alignment = right_align
            
            row += 1
        
        # Ajuster les largeurs
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

    def _get_warehouse_data(self, warehouse, dashboard):
        """Récupérer les données d'un entrepôt."""
        # Placeholder - à implémenter selon votre logique
        return {
            'qty': 0.0,
            'value': 0.0,
            'initial': 0.0,
            'vsd': 0.0
        }

    def _get_warehouse_inventory_lines(self, warehouse):
        """Récupérer les lignes d'inventaire d'un entrepôt."""
        locations = self.env['stock.location'].search([
            ('warehouse_id', '=', warehouse.id),
            ('usage', '=', 'internal')
        ])
        
        domain = [('location_id', 'in', locations.ids)]
        
        if self.date_from:
            domain.append(('inventory_id.date', '>=', self.date_from))
        if self.date_to:
            domain.append(('inventory_id.date', '<=', self.date_to))
        if self.category_ids:
            domain.append(('product_id.categ_id', 'in', self.category_ids.ids))
        
        return self.env['stockex.stock.inventory.line'].search(domain)

    def action_download(self):
        """Télécharger le fichier."""
        self.ensure_one()
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content?model={self._name}&id={self.id}&field=export_file&filename={self.export_filename}&download=true',
            'target': 'self',
        }

    def action_back(self):
        """Retour à la configuration."""
        self.write({'state': 'draft'})
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'view_mode': 'form',
            'res_id': self.id,
            'views': [(False, 'form')],
            'target': 'new',
        }

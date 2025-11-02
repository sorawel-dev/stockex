# -*- coding: utf-8 -*-
from odoo import models, fields, api
from odoo.exceptions import UserError
import base64
import io
import logging

_logger = logging.getLogger(__name__)


class ImportExcelWizard(models.TransientModel):
    _name = 'stockex.import.excel.wizard'
    _description = 'Assistant d\'Import d\'Inventaire Excel'

    name = fields.Char(
        string='Nom de l\'inventaire',
        required=True,
        default='Import Excel',
        help='Nom descriptif pour identifier cet inventaire'
    )
    date = fields.Date(
        string='Date',
        required=True,
        default=fields.Date.today,
        help='Date de réalisation de l\'inventaire'
    )
    file = fields.Binary(
        string='Fichier Excel',
        required=True,
        help='Fichier Excel (.xlsx) contenant les données d\'inventaire'
    )
    filename = fields.Char(
        string='Nom du fichier',
        help='Nom du fichier Excel importé'
    )
    sheet_name = fields.Char(
        string='Feuille à importer',
        help='Nom de la feuille Excel à importer. Laissez vide pour utiliser la première feuille disponible.'
    )
    auto_detect_sheet = fields.Boolean(
        string='Détecter automatiquement',
        default=True,
        help='Utilise automatiquement la première feuille du fichier Excel'
    )
    create_missing_products = fields.Boolean(
        string='Créer les produits manquants',
        default=lambda self: self.env['ir.config_parameter'].sudo().get_param('stockex.excel_create_products', 'True') == 'True',
        help='Si coché, les produits non trouvés seront créés automatiquement'
    )
    create_missing_locations = fields.Boolean(
        string='Créer les emplacements manquants',
        default=lambda self: self.env['ir.config_parameter'].sudo().get_param('stockex.excel_create_locations', 'True') == 'True',
        help='Si coché, les emplacements non trouvés seront créés automatiquement'
    )
    update_product_prices = fields.Boolean(
        string='Mettre à jour les prix des produits',
        default=lambda self: self.env['ir.config_parameter'].sudo().get_param('stockex.excel_update_prices', 'False') == 'True',
        help='Si coché, les prix des produits seront mis à jour avec les valeurs du fichier'
    )
    import_geolocation = fields.Boolean(
        string='Importer la géolocalisation',
        default=lambda self: self.env['ir.config_parameter'].sudo().get_param('stockex.excel_import_geolocation', 'True') == 'True',
        help='Si coché, les coordonnées GPS seront importées si disponibles dans le fichier'
    )
    company_id = fields.Many2one(
        comodel_name='res.company',
        string='Société',
        default=lambda self: self.env.company,
        required=True
    )
    
    def _parse_excel(self):
        """Parse le fichier Excel et retourne les lignes."""
        self.ensure_one()
        
        if not self.file:
            raise UserError("Veuillez sélectionner un fichier Excel.")
        
        try:
            # Importer openpyxl
            try:
                from openpyxl import load_workbook
            except ImportError:
                raise UserError(
                    "Le module 'openpyxl' n'est pas installé.\n"
                    "Installez-le avec: pip install openpyxl"
                )
            
            # Décoder le fichier
            file_content = base64.b64decode(self.file)
            excel_file = io.BytesIO(file_content)
            
            # Charger le workbook
            wb = load_workbook(excel_file, read_only=True, data_only=True)
            
            # Déterminer quelle feuille utiliser
            if self.auto_detect_sheet or not self.sheet_name:
                # Utiliser la première feuille disponible
                sheet_to_use = wb.sheetnames[0]
                _logger.info(f"📄 Utilisation automatique de la feuille: {sheet_to_use}")
            else:
                # Vérifier que la feuille spécifiée existe
                if self.sheet_name not in wb.sheetnames:
                    raise UserError(
                        f"La feuille '{self.sheet_name}' n'existe pas dans le fichier.\n"
                        f"Feuilles disponibles : {', '.join(wb.sheetnames)}\n\n"
                        f"💡 Conseil: Cochez 'Détecter automatiquement' pour utiliser la première feuille."
                    )
                sheet_to_use = self.sheet_name
            
            ws = wb[sheet_to_use]
            _logger.info(f"✅ Lecture de la feuille: {sheet_to_use}")
            
            # Lire les en-têtes (première ligne)
            headers = []
            for cell in ws[1]:
                headers.append(cell.value)
            
            # Lire les données
            lines = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):  # Ignorer les lignes vides
                    continue
                
                row_dict = {}
                for idx, value in enumerate(row):
                    if idx < len(headers):
                        row_dict[headers[idx]] = value if value is not None else ''
                
                lines.append(row_dict)
            
            wb.close()
            return lines
            
        except Exception as e:
            raise UserError(f"Erreur lors de la lecture du fichier Excel : {str(e)}")

    def action_preview(self):
        """Prévisualise l'import sans créer de données."""
        self.ensure_one()
        
        lines = self._parse_excel()
        
        if not lines:
            raise UserError("Le fichier Excel ne contient aucune ligne de données.")
        
        # Analyser les données
        products_count = len(set(line.get('CODE PRODUIT', '') for line in lines if line.get('CODE PRODUIT')))
        locations_count = len(set(line.get('CODE ENTREPOT', '') for line in lines if line.get('CODE ENTREPOT')))
        
        message = f"📊 Aperçu de l'import\n\n"
        message += f"Feuille : {self.sheet_name}\n"
        message += f"Total de lignes : {len(lines)}\n"
        message += f"Produits uniques : {products_count}\n"
        message += f"Emplacements uniques : {locations_count}\n\n"
        
        # Afficher les 5 premières lignes
        message += "Aperçu des 5 premières lignes :\n"
        for i, line in enumerate(lines[:5], 1):
            message += f"\n{i}. {line.get('CODE PRODUIT', '')} - {line.get('PRODUIT', '')}\n"
            message += f"   Emplacement : {line.get('ENTREPOT', '')}\n"
            message += f"   Quantité : {line.get('QUANTITE', '')}\n"
        
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': 'Aperçu de l\'import',
                'message': message,
                'type': 'info',
                'sticky': True,
            }
        }

    def action_import(self):
        """Importe les données et crée l'inventaire."""
        self.ensure_one()
        
        lines = self._parse_excel()
        
        if not lines:
            raise UserError("Le fichier Excel ne contient aucune ligne de données.")
        
        # Générer un nom unique avec timestamp si nécessaire
        from datetime import datetime
        inventory_name = self.name
        existing = self.env['stockex.stock.inventory'].search([
            ('name', '=', inventory_name),
            ('company_id', '=', self.company_id.id)
        ])
        if existing:
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            inventory_name = f"{self.name} ({timestamp})"
        
        # Créer l'inventaire
        inventory = self.env['stockex.stock.inventory'].create({
            'name': inventory_name,
            'date': self.date,
            'company_id': self.company_id.id,
            'user_id': self.env.user.id,
            'state': 'draft',
            'description': f'Import Excel - {self.filename}'
        })
        
        # Caches pour éviter les recherches répétées
        products_cache = {}
        warehouses_cache = {}
        parent_warehouses_cache = {}
        categories_cache = {}
        
        imported = 0
        skipped = 0
        errors_detail = []
        
        for i, line in enumerate(lines):
            try:
                # Extraire les données
                product_code = str(line.get('CODE PRODUIT', '')).strip()
                product_name = str(line.get('PRODUIT', '')).strip()
                location_code = str(line.get('CODE ENTREPOT', '')).strip()
                location_name = str(line.get('ENTREPOT', '')).strip()
                parent_location_code = str(line.get('CODE ENTREPOT PARENT', '')).strip()
                parent_location_name = str(line.get('ENTREPOT PARENT', '')).strip()
                category_code = str(line.get('CODE CATEGORIE', '')).strip()
                category_name = str(line.get('CATEGORIE', '')).strip()
                
                # Nettoyer la quantité (format anglais: virgule=milliers, point=décimales)
                quantity_raw = (line.get('QUANTITE') or line.get('QTE') or 
                               line.get('Quantite') or line.get('Quantité') or 
                               line.get('QUANTITY') or line.get('Qte') or '0')
                quantity_str = str(quantity_raw).strip()
                
                # Enlever espaces et virgules (milliers), garder le point (décimal)
                quantity_str = quantity_str.replace(' ', '').replace(',', '')
                try:
                    quantity = float(quantity_str) if quantity_str and quantity_str not in ['', '-', 'None', '0'] else 0.0
                except Exception as e:
                    _logger.warning(f"Erreur parsing quantité ligne {i+2}: '{quantity_raw}' -> '{quantity_str}' : {e}")
                    quantity = 0.0
                
                # Nettoyer le prix (format anglais: virgule=milliers, point=décimales)
                price_str = str(line.get('COUT UNITAIRE', '0') or line.get('COUT', '0'))
                price_str = price_str.replace(' ', '').replace(',', '')  # Enlever virgules (milliers)
                try:
                    standard_price = float(price_str) if price_str and price_str not in ['', '-', 'None'] else 0.0
                except:
                    standard_price = 0.0
                
                if not product_code or not location_code:
                    skipped += 1
                    continue
                
                # Rechercher ou créer la catégorie de produit
                category_id = None
                if category_code and category_code not in categories_cache:
                    category = self.env['product.category'].search([
                        '|',
                        ('name', '=', category_name),
                        ('name', '=', category_code)
                    ], limit=1)
                    
                    if not category and self.create_missing_products:
                        category = self.env['product.category'].create({
                            'name': category_name or category_code,
                        })
                    
                    categories_cache[category_code] = category.id if category else False
                
                category_id = categories_cache.get(category_code) if category_code else None
                
                # Rechercher ou créer l'entrepôt parent
                parent_warehouse_id = None
                if parent_location_code and parent_location_code not in parent_warehouses_cache:
                    parent_warehouse = self.env['stock.warehouse'].search([
                        '|',
                        ('code', '=', parent_location_code),
                        ('name', '=', str(line.get('ENTREPOT PARENT', '')).strip())
                    ], limit=1)
                    
                    if not parent_warehouse and self.create_missing_locations:
                        # Créer l'entrepôt parent
                        parent_warehouse = self.env['stock.warehouse'].create({
                            'name': parent_location_name or parent_location_code,
                            'code': parent_location_code[:5].upper(),  # Code court pour l'entrepôt
                            'company_id': self.company_id.id,
                        })
                    
                    parent_warehouses_cache[parent_location_code] = parent_warehouse.id if parent_warehouse else False
                
                parent_warehouse_id = parent_warehouses_cache.get(parent_location_code)
                
                # Rechercher ou créer l'entrepôt
                if location_code not in warehouses_cache:
                    warehouse = self.env['stock.warehouse'].search([
                        '|',
                        ('code', '=', location_code),
                        ('name', '=', str(line.get('ENTREPOT', '')).strip())
                    ], limit=1)
                    
                    if not warehouse and self.create_missing_locations:
                        # Préparer les valeurs pour l'entrepôt
                        warehouse_vals = {
                            'name': location_name or location_code,
                            'code': location_code[:5].upper(),  # Code court pour l'entrepôt
                            'company_id': self.company_id.id,
                            'parent_id': parent_warehouse_id if parent_warehouse_id else False,
                        }
                        
                        # Ajouter la géolocalisation si demandé
                        if self.import_geolocation:
                            try:
                                lat = line.get('LATITUDE', '') or line.get('Latitude', '')
                                lon = line.get('LONGITUDE', '') or line.get('Longitude', '')
                                if lat and lon:
                                    warehouse_vals['latitude'] = float(str(lat).replace(',', '.'))
                                    warehouse_vals['longitude'] = float(str(lon).replace(',', '.'))
                                
                                # Autres infos de géolocalisation
                                if line.get('VILLE') or line.get('Ville'):
                                    warehouse_vals['city'] = str(line.get('VILLE') or line.get('Ville', '')).strip()
                                if line.get('ADRESSE') or line.get('Adresse'):
                                    warehouse_vals['address'] = str(line.get('ADRESSE') or line.get('Adresse', '')).strip()
                                if line.get('TELEPHONE') or line.get('Téléphone'):
                                    warehouse_vals['phone'] = str(line.get('TELEPHONE') or line.get('Téléphone', '')).strip()
                                if line.get('EMAIL') or line.get('Email'):
                                    warehouse_vals['email'] = str(line.get('EMAIL') or line.get('Email', '')).strip()
                            except Exception as e:
                                _logger.warning(f"Erreur géolocalisation ligne {i+2}: {e}")
                        
                        warehouse = self.env['stock.warehouse'].create(warehouse_vals)
                    
                    # Stocker l'emplacement de stock de l'entrepôt (lot_stock_id)
                    warehouses_cache[location_code] = warehouse.lot_stock_id.id if warehouse else False
                
                location_id = warehouses_cache.get(location_code)
                if not location_id:
                    skipped += 1
                    errors_detail.append(f"Ligne {i+2}: Entrepôt '{location_code}' non trouvé")
                    continue
                
                # Rechercher ou créer le produit
                if product_code not in products_cache:
                    product = self.env['product.product'].search([
                        ('default_code', '=', product_code)
                    ], limit=1)
                    
                    if not product and self.create_missing_products:
                        uom_name = str(line.get('UDM', 'PC')).strip()
                        
                        # Rechercher l'UdM
                        uom = self.env['uom.uom'].search([
                            ('name', 'ilike', uom_name)
                        ], limit=1)
                        
                        if not uom:
                            uom = self.env.ref('uom.product_uom_unit')
                        
                        product = self.env['product.product'].create({
                            'name': product_name or product_code,
                            'default_code': product_code,
                            'type': 'product',
                            'categ_id': category_id if category_id else self.env.ref('product.product_category_all').id,
                            'uom_id': uom.id,
                            'uom_po_id': uom.id,
                            'standard_price': standard_price,
                        })
                    
                    products_cache[product_code] = product.id if product else False
                
                product_id = products_cache.get(product_code)
                if not product_id:
                    skipped += 1
                    errors_detail.append(f"Ligne {i+2}: Produit '{product_code}' non trouvé")
                    continue
                
                # Mettre à jour le prix si demandé
                if self.update_product_prices and standard_price > 0:
                    product = self.env['product.product'].browse(product_id)
                    if product.standard_price != standard_price:
                        product.write({'standard_price': standard_price})
                
                # Log pour debug (peut être enlevé après)
                _logger.info(f"Import ligne {i+2}: Produit={product_code}, Qté={quantity}, Prix={standard_price}")
                
                # Créer la ligne d'inventaire avec le prix unitaire
                self.env['stockex.stock.inventory.line'].create({
                    'inventory_id': inventory.id,
                    'product_id': product_id,
                    'location_id': location_id,
                    'product_qty': quantity,
                    'standard_price': standard_price,
                })
                
                imported += 1
                
                # Commit par batch tous les 500 enregistrements
                if imported % 500 == 0:
                    self.env.cr.commit()
                    
            except Exception as e:
                skipped += 1
                errors_detail.append(f"Ligne {i+2}: {str(e)}")
                _logger.error(f"Erreur ligne {i+2}: {str(e)}")
        
        # Message de confirmation
        message = f"Import terminé avec succès !\n\n"
        message += f"✅ Lignes importées : {imported}\n"
        message += f"⚠️  Lignes ignorées : {skipped}\n"
        if errors_detail[:5]:
            message += f"\nPremières erreurs :\n" + "\n".join(errors_detail[:5])
        
        inventory.write({
            'description': inventory.description + f"\n\n{message}"
        })
        
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'stockex.stock.inventory',
            'res_id': inventory.id,
            'view_mode': 'form',
            'target': 'current',
            'context': {'form_view_initial_mode': 'edit'}
        }

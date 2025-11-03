# -*- coding: utf-8 -*-

import base64
import io
import logging

from odoo import models, fields, api
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)


class ImportFlexibleInventoryWizard(models.TransientModel):
    _name = 'stockex.import.flexible.inventory.wizard'
    _description = 'Import Flexible d\'Inventaire (Excel/CSV)'

    name = fields.Char(
        string='Nom de l\'inventaire',
        required=True,
        default=lambda self: f'Inventaire {fields.Date.today().strftime("%d/%m/%Y")}',
    )
    date = fields.Date(
        string='Date',
        required=True,
        default=fields.Date.today,
    )
    file = fields.Binary(
        string='Fichier Excel ou CSV',
        required=True,
        help='Fichier contenant les données d\'inventaire'
    )
    filename = fields.Char(string='Nom du fichier')
    
    # Options d'import
    create_products = fields.Boolean(
        string='Créer les produits manquants',
        default=True,
    )
    create_categories = fields.Boolean(
        string='Créer les catégories manquantes',
        default=True,
    )
    create_warehouses = fields.Boolean(
        string='Créer les entrepôts manquants',
        default=True,
    )
    create_locations = fields.Boolean(
        string='Créer les emplacements enfants manquants',
        default=True,
        help='Crée les sous-magasins, rayons, casiers sous l\'entrepôt parent',
    )
    update_prices = fields.Boolean(
        string='Mettre à jour les prix',
        default=False,
        help='Si coché, met à jour le prix de vente des produits avec les valeurs du fichier',
    )
    default_price = fields.Float(
        string='Prix par défaut',
        default=0.0,
        help='Prix utilisé si colonne PRIX UNITAIRE absente',
    )
    
    # Option multi-entrepôts
    multi_warehouse_mode = fields.Selection([
        ('global', 'Inventaire Global (tous les entrepôts ensemble)'),
        ('split', 'Inventaires Séparés (un par entrepôt)'),
    ], string='Mode Multi-Entrepôts', default='global', required=True,
       help='Comment gérer les données provenant de plusieurs entrepôts')
    
    # Prévisualisation
    state = fields.Selection([
        ('draft', 'Configuration'),
        ('preview', 'Prévisualisation'),
    ], default='draft')
    
    preview_summary = fields.Html(string='Résumé', readonly=True)
    preview_data = fields.Text(string='Données', readonly=True)  # JSON des données parsées

    def _get_column_mapping(self, headers):
        """Détecte automatiquement les colonnes du fichier.
        
        Returns: dict avec les mappings {colonne_standard: nom_dans_fichier}
        """
        mapping = {}
        headers_upper = {h.upper().strip(): h for h in headers if h}
        
        # CODE PRODUIT
        for variant in ['CODE PRODUIT', 'MATERIAL', 'CODE D\'ARTICLE', 'CODE', 'PRODUCT_CODE']:
            if variant in headers_upper:
                mapping['code'] = headers_upper[variant]
                break
        
        # NOM PRODUIT
        for variant in ['PRODUIT', 'NOM PRODUIT', 'MATERIAL DESCRIPTION', 'NOM DU MATERIEL', 'PRODUCT_NAME', 'DESIGNATION']:
            if variant in headers_upper:
                mapping['name'] = headers_upper[variant]
                break
        
        # QUANTITE
        for variant in ['QUANTITE', 'QUANTITÉ', 'QTE', 'QUANTITY', 'QTY']:
            if variant in headers_upper:
                mapping['quantity'] = headers_upper[variant]
                break
        
        # PRIX UNITAIRE (optionnel)
        for variant in ['PRIX UNITAIRE', 'PRIX', 'PRICE', 'STANDARD_PRICE', 'UNIT_PRICE']:
            if variant in headers_upper:
                mapping['price'] = headers_upper[variant]
                break
        
        # CATEGORIE (optionnel)
        for variant in ['CATEGORIE', 'CATÉGORIE', 'CATEGORY', 'TYPE D\'ARTICLE', 'TYPE']:
            if variant in headers_upper:
                mapping['category'] = headers_upper[variant]
                break
        
        # ENTREPOT (optionnel) - magasin principal uniquement
        for variant in ['MAGASIN', 'WAREHOUSE', 'ENTREPOT', 'ENTREPÔT']:
            if variant in headers_upper:
                mapping['warehouse'] = headers_upper[variant]
                break
        
        # SOUS-MAGASIN / EMPLACEMENT ENFANT (optionnel)
        for variant in ['EMPLACEMENT', 'NOM DU SOUS-MAGASIN', 'SOUS-MAGASIN', 'SUB-LOCATION', 'SUBLOCATION', 'LOCATION']:
            if variant in headers_upper:
                mapping['sub_location'] = headers_upper[variant]
                break
        
        # RANGER / RAYON (optionnel)
        for variant in ['RANGER', 'RAYON', 'AISLE', 'ROW']:
            if variant in headers_upper:
                mapping['aisle'] = headers_upper[variant]
                break
        
        # CASIER / ETAGERE (optionnel)
        for variant in ['CASIER', 'ÉTAGÈRE', 'ETAGERE', 'RACK', 'SHELF']:
            if variant in headers_upper:
                mapping['rack'] = headers_upper[variant]
                break
        
        return mapping

    def _parse_excel_file(self, preview_mode=False):
        """Parse un fichier Excel et retourne les données.
        
        Args:
            preview_mode: Si True, limite la lecture pour économiser la mémoire
        """
        try:
            import openpyxl
        except ImportError:
            raise UserError("La bibliothèque openpyxl n'est pas installée.")
        
        if not self.file:
            raise UserError("Veuillez sélectionner un fichier.")
        
        try:
            # Décoder le fichier
            file_data = base64.b64decode(self.file)
            file_obj = io.BytesIO(file_data)
            
            # Ouvrir le workbook en mode read_only pour économiser la mémoire
            wb = openpyxl.load_workbook(file_obj, data_only=True, read_only=True)
            ws = wb.active
            
            # Lire les en-têtes
            headers = []
            first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = list(first_row)
            
            # Détecter les colonnes
            mapping = self._get_column_mapping(headers)
            
            # Identifier les colonnes utiles à garder
            useful_cols = set()
            for key, col_name in mapping.items():
                if col_name in headers:
                    useful_cols.add(headers.index(col_name))
            
            # Lire les données (optimisé)
            data = []
            max_rows = 100 if preview_mode else None  # Limiter en mode preview
            
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
                if preview_mode and idx > max_rows:
                    break
                
                # Ne garder que les colonnes utiles
                row_data = {}
                for col_idx, value in enumerate(row):
                    if col_idx < len(headers) and (not preview_mode or col_idx in useful_cols):
                        header = headers[col_idx]
                        if header:
                            row_data[header] = value
                
                data.append(row_data)
            
            # Compter le nombre total de lignes si en mode preview
            total_lines = len(data)
            if preview_mode:
                # Estimation du nombre total
                total_lines = ws.max_row - 1 if ws.max_row else len(data)
            
            wb.close()
            
            return {
                'headers': headers,
                'mapping': mapping,
                'data': data,
                'total_lines': total_lines,
            }
            
        except MemoryError:
            raise UserError("Le fichier est trop volumineux. Veuillez réduire sa taille ou contacter l'administrateur.")
        except Exception as e:
            raise UserError(f"Erreur lors de la lecture du fichier : {str(e)}")

    def action_preview(self):
        """Prévisualise les données avant import."""
        self.ensure_one()
        
        # Parser le fichier en mode preview (limité pour économiser la mémoire)
        parsed = self._parse_excel_file(preview_mode=True)
        mapping = parsed['mapping']
        data = parsed['data']
        
        # Vérifier les colonnes obligatoires
        if 'code' not in mapping or 'quantity' not in mapping:
            raise UserError(
                "Colonnes obligatoires manquantes !\n\n"
                f"Colonnes détectées : {', '.join(parsed['headers'][:10])}\n\n"
                "Colonnes requises :\n"
                "- CODE PRODUIT (ou Material, Code d'article)\n"
                "- QUANTITE (ou Quantité, Quantity)\n\n"
                "Colonnes optionnelles :\n"
                "- PRODUIT (nom)\n"
                "- PRIX UNITAIRE (sinon prix = 0)\n"
                "- CATEGORIE\n"
                "- ENTREPOT (ou Magasin)"
            )
        
        # Analyser les données et nettoyer pour JSON
        products_count = 0
        warehouses = set()
        sub_locations = set()
        categories = set()
        total_qty = 0
        cleaned_data = []
        
        for line in data:
            # Nettoyer la ligne pour JSON (convertir datetime, etc.)
            cleaned_line = {}
            for key, value in line.items():
                if value is None:
                    cleaned_line[key] = None
                elif isinstance(value, (int, float, str, bool)):
                    cleaned_line[key] = value
                else:
                    # Convertir tout le reste en string
                    cleaned_line[key] = str(value)
            cleaned_data.append(cleaned_line)
            
            code = line.get(mapping['code'])
            if code:
                products_count += 1
            
            qty = line.get(mapping.get('quantity'))
            if qty:
                try:
                    total_qty += float(qty)
                except:
                    pass
            
            # Compter les entrepôts (magasins principaux)
            if 'warehouse' in mapping:
                wh = line.get(mapping['warehouse'])
                if wh:
                    warehouses.add(str(wh))
            
            # Compter les emplacements enfants (sous-magasins)
            if 'sub_location' in mapping:
                sub_loc = line.get(mapping['sub_location'])
                if sub_loc:
                    sub_locations.add(str(sub_loc))
            
            if 'category' in mapping:
                cat = line.get(mapping['category'])
                if cat:
                    categories.add(str(cat))
        
        # Générer le résumé
        summary = f"""
        <div style="font-family: Arial, sans-serif;">
            <h3 style="color: #00A09D;">📊 Résumé de l'Import</h3>
            
            <div style="background: #f0f9ff; padding: 15px; border-radius: 8px; margin: 10px 0;">
                <h4>📁 Fichier</h4>
                <p><strong>{self.filename}</strong></p>
                <p>Total lignes : <strong>{parsed['total_lines']}</strong></p>
                <p style="font-size: 0.9em; color: #666;"><em>Prévisualisation sur {len(data)} premières lignes</em></p>
            </div>
            
            <div style="background: #f0fdf4; padding: 15px; border-radius: 8px; margin: 10px 0;">
                <h4>✅ Colonnes Détectées</h4>
                <ul>
                    <li>Code Produit : <strong>{mapping.get('code', '❌ NON TROUVÉ')}</strong></li>
                    <li>Nom Produit : <strong>{mapping.get('name', 'Non spécifié')}</strong></li>
                    <li>Quantité : <strong>{mapping.get('quantity', '❌ NON TROUVÉ')}</strong></li>
                    <li>Prix : <strong>{mapping.get('price', f'Par défaut: {self.default_price}')}</strong></li>
                    <li>Catégorie : <strong>{mapping.get('category', 'Non spécifié')}</strong></li>
                    <li>Entrepôt : <strong>{mapping.get('warehouse', 'Non spécifié')}</strong></li>
                </ul>
            </div>
            
            <div style="background: #fefce8; padding: 15px; border-radius: 8px; margin: 10px 0;">
                <h4>📊 Statistiques</h4>
                <ul>
                    <li>Produits : <strong>{products_count}</strong></li>
                    <li>Quantité totale : <strong>{total_qty:,.0f}</strong></li>
                    {f'<li>Entrepôts : <strong>{len(warehouses)}</strong> ({", ".join(sorted(warehouses))})</li>' if warehouses else ''}
                    {f'<li>Emplacements : <strong>{len(sub_locations)}</strong></li>' if sub_locations else ''}
                    {f'<li>Catégories : <strong>{len(categories)}</strong></li>' if categories else ''}
                </ul>
            </div>
            
            {f'''
            <div style="background: #fef3c7; padding: 15px; border-radius: 8px; margin: 10px 0;">
                <h4>🏭 Multi-Entrepôts Détecté</h4>
                <p>Le fichier contient des données pour <strong>{len(warehouses)}</strong> entrepôt(s).</p>
                <p>Mode sélectionné : <strong>{'Inventaire Global' if self.multi_warehouse_mode == 'global' else 'Inventaires Séparés'}</strong></p>
                {f'<p>➡️ Résultat : <strong>{"1 inventaire avec toutes les données" if self.multi_warehouse_mode == "global" else str(len(warehouses)) + " inventaires (un par entrepôt)"}</strong></p>' if warehouses else ''}
            </div>
            ''' if warehouses and len(warehouses) > 1 else ''}
            
            <div style="background: #e0f2fe; padding: 15px; border-radius: 8px; margin: 10px 0;">
                <h4>⚙️ Options d'Import</h4>
                <ul>
                    <li>{'✅' if self.create_products else '❌'} Créer les produits manquants</li>
                    <li>{'✅' if self.create_categories else '❌'} Créer les catégories manquantes</li>
                    <li>{'✅' if self.create_warehouses else '❌'} Créer les entrepôts manquants</li>
                    <li>{'✅' if self.create_locations else '❌'} Créer les emplacements enfants manquants</li>
                    <li>Prix par défaut : <strong>{self.default_price}</strong></li>
                </ul>
            </div>
        </div>
        """
        
        # Sauvegarder pour l'import (uniquement les métadonnées, pas les données complètes)
        import json
        self.write({
            'state': 'preview',
            'preview_summary': summary,
            'preview_data': json.dumps({
                'mapping': mapping,
                'warehouses': list(warehouses),
                'total': len(data),
                # Ne pas stocker les données complètes - elles seront relues du fichier lors de l'import
            }),
        })
        
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'stockex.import.flexible.inventory.wizard',
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
        }

    def action_import(self):
        """Importe les données et crée le(s) inventaire(s)."""
        self.ensure_one()
        
        if self.state != 'preview':
            raise UserError("Veuillez d'abord prévisualiser les données.")
        
        import json
        preview = json.loads(self.preview_data)
        mapping = preview['mapping']
        warehouses = preview.get('warehouses', [])
        
        # Re-parser le fichier complet (car données non stockées en preview)
        _logger.info("Re-parsing du fichier Excel pour l'import...")
        parsed = self._parse_excel_file()
        data = parsed['data']
        _logger.info(f"Fichier parsé: {len(data)} lignes")
        
        created_inventories = []
        
        if self.multi_warehouse_mode == 'global' or not warehouses:
            # Mode global : 1 seul inventaire
            inventory = self._create_inventory_from_data(
                name=self.name,
                data=data,
                mapping=mapping,
            )
            created_inventories.append(inventory)
        else:
            # Mode split : 1 inventaire par entrepôt
            for warehouse_name in warehouses:
                # Filtrer les données pour cet entrepôt
                warehouse_data = [
                    line for line in data
                    if line.get(mapping.get('warehouse')) == warehouse_name
                ]
                
                if warehouse_data:
                    inventory = self._create_inventory_from_data(
                        name=f"{self.name} - {warehouse_name}",
                        data=warehouse_data,
                        mapping=mapping,
                        warehouse_filter=warehouse_name,
                    )
                    created_inventories.append(inventory)
        
        # Afficher le(s) inventaire(s) créé(s)
        if len(created_inventories) == 1:
            return {
                'type': 'ir.actions.act_window',
                'res_model': 'stockex.stock.inventory',
                'view_mode': 'form',
                'res_id': created_inventories[0].id,
                'target': 'current',
            }
        else:
            return {
                'type': 'ir.actions.act_window',
                'res_model': 'stockex.stock.inventory',
                'view_mode': 'tree,form',
                'domain': [('id', 'in', [inv.id for inv in created_inventories])],
                'target': 'current',
            }

    def _get_or_create_child_location(self, parent_location, location_name, location_type='view'):
        """Crée ou récupère un emplacement enfant sous un emplacement parent."""
        if not location_name or not str(location_name).strip():
            return parent_location
        
        location_name = str(location_name).strip()
        
        # Rechercher l'emplacement existant
        child_location = self.env['stock.location'].search([
            ('name', '=', location_name),
            ('location_id', '=', parent_location.id),
        ], limit=1)
        
        if not child_location and self.create_locations:
            # Créer le nouvel emplacement
            child_location = self.env['stock.location'].create({
                'name': location_name,
                'location_id': parent_location.id,
                'usage': 'internal',
                'company_id': parent_location.company_id.id,
            })
            _logger.info(f"✅ Emplacement créé: {location_name} sous {parent_location.complete_name}")
        
        return child_location or parent_location
    
    def _create_inventory_from_data(self, name, data, mapping, warehouse_filter=None):
        """Crée un inventaire à partir des données parsées."""
        # Générer un nom unique
        base_name = name
        counter = 1
        unique_name = base_name
        
        while self.env['stockex.stock.inventory'].search([
            ('name', '=', unique_name),
            ('company_id', '=', self.env.company.id)
        ], limit=1):
            counter += 1
            unique_name = f"{base_name} ({counter})"
            if counter > 100:  # Sécurité
                unique_name = f"{base_name} - {fields.Datetime.now().strftime('%H:%M:%S')}"
                break
        
        # Créer l'inventaire
        inventory = self.env['stockex.stock.inventory'].create({
            'name': unique_name,
            'date': self.date,
            'state': 'draft',
        })
        
        created_count = 0
        errors = []
        
        for i, line in enumerate(data):
            try:
                # Extraire les données
                code = str(line.get(mapping['code'], '')).strip()
                if not code:
                    continue
                
                name_product = line.get(mapping.get('name'), code)
                quantity = float(line.get(mapping.get('quantity'), 0) or 0)
                price = float(line.get(mapping.get('price'), self.default_price) or self.default_price)
                category_name = line.get(mapping.get('category'))
                warehouse_name = line.get(mapping.get('warehouse'), 'Stock')
                
                # Rechercher ou créer le produit
                product = self.env['product.product'].search([
                    ('default_code', '=', code)
                ], limit=1)
                
                # Gérer la catégorie
                category = None
                if category_name and self.create_categories:
                    category = self.env['product.category'].search([
                        ('name', '=', str(category_name))
                    ], limit=1)
                    
                    if not category:
                        category = self.env['product.category'].create({
                            'name': str(category_name)
                        })
                
                if not product and self.create_products:
                    # Type = consu (Biens/Goods) pour suivi par quantité
                    product_vals = {
                        'name': str(name_product),
                        'default_code': code,
                        'type': 'consu',  # Type = Biens/Goods
                        'is_storable': True,  # ✅ Cocher "Suivre l'inventaire"
                        'standard_price': price,
                    }
                    
                    if category:
                        product_vals['categ_id'] = category.id
                    
                    product = self.env['product.product'].create(product_vals)
                    _logger.info(f"✅ Produit créé: {code} - {name_product} (type: consu - Biens/Goods)")
                elif product:
                    # Mettre à jour la catégorie et le prix si demandé
                    update_vals = {}
                    if category and product.categ_id.id != category.id:
                        update_vals['categ_id'] = category.id
                    if self.update_prices and price > 0:
                        update_vals['standard_price'] = price
                    if update_vals:
                        product.write(update_vals)
                        if 'standard_price' in update_vals:
                            _logger.info(f"✅ Prix mis à jour pour {code}: {price}")
                
                if not product:
                    errors.append(f"Ligne {i+2}: Produit '{code}' non trouvé")
                    continue
                
                # Vérifier et convertir en type 'consu' (Biens/Goods) si nécessaire
                # Type = consu (Biens/Goods) → Suivi d'inventaire par quantité
                if product.type != 'consu' or not product.is_storable:
                    _logger.warning(f"⚠️ Produit {code} n'est pas correctement configuré. Type: {product.type}, Suivre inventaire: {product.is_storable}. Correction...")
                    product.write({
                        'type': 'consu',  # Convertir en Biens/Goods
                        'is_storable': True,  # ✅ Cocher "Suivre l'inventaire"
                    })
                    _logger.info(f"✅ Produit {code}: Converti en type Biens/Goods (consu) avec suivi d'inventaire activé")
                
                # Gérer l'entrepôt
                warehouse = self.env['stock.warehouse'].search([
                    '|',
                    ('name', '=', warehouse_name),
                    ('code', '=', warehouse_name[:5].upper())
                ], limit=1)
                
                if not warehouse and self.create_warehouses:
                    try:
                        # Utiliser un savepoint pour éviter d'annuler toute la transaction
                        savepoint_name = f'warehouse_create_{i}'
                        self.env.cr.execute(f'SAVEPOINT {savepoint_name}')
                        
                        warehouse = self.env['stock.warehouse'].create({
                            'name': warehouse_name,
                            'code': warehouse_name[:5].upper(),
                        })
                        
                        self.env.cr.execute(f'RELEASE SAVEPOINT {savepoint_name}')
                        _logger.info(f"✅ Entrepôt créé: {warehouse_name} (code: {warehouse.code})")
                        
                    except Exception as wh_error:
                        # Rollback au savepoint pour continuer
                        self.env.cr.execute(f'ROLLBACK TO SAVEPOINT {savepoint_name}')
                        _logger.warning(f"⚠️ Erreur création entrepôt {warehouse_name}: {wh_error}")
                        
                        # Retry recherche (peut exister avec un code différent)
                        warehouse = self.env['stock.warehouse'].search([
                            '|',
                            ('name', '=', warehouse_name),
                            ('code', 'ilike', warehouse_name[:5])
                        ], limit=1)
                
                if not warehouse:
                    errors.append(f"Ligne {i+2}: Entrepôt '{warehouse_name}' non trouvé")
                    continue
                
                # Point de départ : emplacement stock de l'entrepôt
                location = warehouse.lot_stock_id
                
                # Gérer la hiérarchie d'emplacements enfants si présente
                if 'sub_location' in mapping:
                    sub_location_name = line.get(mapping['sub_location'])
                    if sub_location_name and str(sub_location_name).strip():
                        location = self._get_or_create_child_location(location, sub_location_name)
                
                if 'aisle' in mapping:
                    aisle_name = line.get(mapping['aisle'])
                    if aisle_name and str(aisle_name).strip():
                        location = self._get_or_create_child_location(location, aisle_name)
                
                if 'rack' in mapping:
                    rack_name = line.get(mapping['rack'])
                    if rack_name and str(rack_name).strip():
                        location = self._get_or_create_child_location(location, rack_name)
                
                # Récupérer la quantité théorique depuis les quants
                StockQuant = self.env['stock.quant']
                quant = StockQuant.search([
                    ('product_id', '=', product.id),
                    ('location_id', '=', location.id),
                    ('company_id', '=', self.company_id.id),
                ], limit=1)
                
                theoretical_qty = 0.0
                if quant:
                    theoretical_qty = quant.quantity - quant.reserved_quantity
                
                # Log pour debug
                _logger.info(
                    f"Import ligne {i+2}: Produit={code}, "
                    f"Qté théo={theoretical_qty}, Qté réelle={quantity}, Prix={price}"
                )
                
                # Créer la ligne d'inventaire avec quantité théorique ET prix unitaire
                self.env['stockex.stock.inventory.line'].create({
                    'inventory_id': inventory.id,
                    'product_id': product.id,
                    'location_id': location.id,
                    'product_qty': quantity,
                    'theoretical_qty': theoretical_qty,
                    'standard_price': price,
                })
                created_count += 1
                
            except Exception as e:
                errors.append(f"Ligne {i+2}: {str(e)}")
                _logger.error(f"Erreur ligne {i+2}: {e}")
        
        # Message récapitulatif
        message = f"✅ Import terminé : {created_count} ligne(s) créée(s)"
        if errors:
            message += f"\n⚠️ {len(errors)} erreur(s):\n" + "\n".join(errors[:20])
        
        # Poster le message avec gestion d'erreur transaction
        try:
            inventory.message_post(body=message)
        except Exception as msg_error:
            _logger.warning(f"⚠️ Impossible de poster le message dans le chatter: {msg_error}")
            # En cas d'erreur de transaction, rollback et retry
            try:
                self.env.cr.rollback()
                self.env.cr.commit()
                inventory.invalidate_recordset()
                inventory = self.env['stockex.stock.inventory'].browse(inventory.id)
                inventory.message_post(body=message)
            except Exception as retry_error:
                _logger.error(f"❌ Échec définitif message_post: {retry_error}")
        
        # Envoyer les notifications
        self._send_notifications(inventory, created_count, len(errors))
        
        return inventory
    
    def _send_notifications(self, inventory, created_count, error_count):
        """Envoie les notifications configurées pour l'inventaire créé."""
        try:
            # Récupérer les paramètres de notification
            IrConfigParameter = self.env['ir.config_parameter'].sudo()
            
            notify_email = IrConfigParameter.get_param('stockex.notify_email', default=False)
            notify_whatsapp = IrConfigParameter.get_param('stockex.notify_whatsapp', default=False)
            notify_telegram = IrConfigParameter.get_param('stockex.notify_telegram', default=False)
            
            email_recipients = IrConfigParameter.get_param('stockex.email_recipients', default='')
            whatsapp_number = IrConfigParameter.get_param('stockex.whatsapp_number', default='')
            telegram_chat_id = IrConfigParameter.get_param('stockex.telegram_chat_id', default='')
            telegram_bot_token = IrConfigParameter.get_param('stockex.telegram_bot_token', default='')
            
            # Préparer le message
            notification_title = f"📦 Nouvel Inventaire Importé"
            notification_body = f"""
Inventaire : {inventory.name}
Date : {inventory.date}
Lignes créées : {created_count}
{'⚠️ Erreurs : ' + str(error_count) if error_count > 0 else '✅ Aucune erreur'}

État : Brouillon (à démarrer)
"""
            
            # Envoyer email
            if notify_email and email_recipients:
                self._send_email_notification(
                    email_recipients,
                    notification_title,
                    notification_body,
                    inventory
                )
            
            # Envoyer WhatsApp
            if notify_whatsapp and whatsapp_number:
                self._send_whatsapp_notification(
                    whatsapp_number,
                    notification_title,
                    notification_body
                )
            
            # Envoyer Telegram
            if notify_telegram and telegram_chat_id and telegram_bot_token:
                self._send_telegram_notification(
                    telegram_bot_token,
                    telegram_chat_id,
                    notification_title,
                    notification_body
                )
                
        except Exception as e:
            _logger.warning(f"Erreur lors de l'envoi des notifications : {str(e)}")
    
    def _send_email_notification(self, recipients, title, body, inventory):
        """Envoie une notification par email."""
        try:
            mail_values = {
                'subject': title,
                'body_html': f"<pre>{body}</pre>",
                'email_to': recipients,
                'auto_delete': True,
            }
            mail = self.env['mail.mail'].sudo().create(mail_values)
            mail.send()
            _logger.info(f"📧 Email envoyé à {recipients}")
        except Exception as e:
            _logger.error(f"Erreur envoi email : {str(e)}")
    
    def _send_whatsapp_notification(self, phone_number, title, body):
        """Envoie une notification WhatsApp."""
        try:
            import requests
            message = f"{title}\n\n{body}"
            # TODO: Implémenter l'envoi WhatsApp via API
            _logger.info(f"📱 WhatsApp préparé pour {phone_number}")
        except Exception as e:
            _logger.error(f"Erreur envoi WhatsApp : {str(e)}")
    
    def _send_telegram_notification(self, bot_token, chat_id, title, body):
        """Envoie une notification Telegram."""
        try:
            import requests
            message = f"<b>{title}</b>\n\n<pre>{body}</pre>"
            url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
            data = {
                'chat_id': chat_id,
                'text': message,
                'parse_mode': 'HTML'
            }
            response = requests.post(url, data=data, timeout=10)
            if response.status_code == 200:
                _logger.info(f"📨 Telegram envoyé au chat {chat_id}")
            else:
                _logger.error(f"Erreur Telegram : {response.status_code}")
        except Exception as e:
            _logger.error(f"Erreur envoi Telegram : {str(e)}")

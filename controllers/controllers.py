# -*- coding: utf-8 -*-
from odoo import http
from odoo.http import request, Response
import io
import logging

_logger = logging.getLogger(__name__)


class StockexInventoryController(http.Controller):
    """Contrôleur pour exports et rapports d'inventaire."""
    
    @http.route('/stockex/inventory/warehouse_report/<int:warehouse_id>/excel', type='http', auth='user')
    def warehouse_inventory_excel_report(self, warehouse_id, **kwargs):
        """Exporte le rapport d'inventaire pour un entrepôt en Excel."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from datetime import datetime
            
            warehouse = request.env['stock.warehouse'].browse(warehouse_id)
            if not warehouse.exists():
                return Response("Entrepôt introuvable", status=404)
            
            # Récupérer tous les inventaires de cet entrepôt
            inventories = request.env['stockex.stock.inventory'].search([
                ('warehouse_id', '=', warehouse_id),
                ('state', '=', 'done')
            ], order='date desc')
            
            # Créer le classeur Excel
            wb = Workbook()
            ws = wb.active
            ws.title = f"Inv_{warehouse.name[:25]}"
            
            # En-tête
            ws['A1'] = f"RAPPORT D'INVENTAIRE - {warehouse.name}"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A2'] = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
            ws['A2'].font = Font(italic=True)
            
            # Colonnes
            headers = ['N° Inv', 'Date', 'Produit', 'Catégorie', 'Emplacement', 'Qté Théo', 'Qté Réelle', 'Écart Qté', 'Prix Unit', 'Valeur Écart']
            row = 4
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            # Données
            row += 1
            for inv in inventories:
                for line in inv.line_ids:
                    ws.cell(row=row, column=1, value=inv.name)
                    ws.cell(row=row, column=2, value=inv.date.strftime('%d/%m/%Y') if inv.date else '')
                    ws.cell(row=row, column=3, value=line.product_id.name)
                    ws.cell(row=row, column=4, value=line.product_id.categ_id.name if line.product_id.categ_id else '')
                    ws.cell(row=row, column=5, value=line.location_id.name)
                    ws.cell(row=row, column=6, value=line.theoretical_qty or 0)
                    ws.cell(row=row, column=7, value=line.product_qty or 0)
                    ws.cell(row=row, column=8, value=line.difference or 0)
                    ws.cell(row=row, column=9, value=line.standard_price or 0)
                    ws.cell(row=row, column=10, value=(line.difference or 0) * (line.standard_price or 0))
                    row += 1
            
            # Ajuster largeur colonnes
            for col in range(1, 11):
                ws.column_dimensions[chr(64 + col)].width = 15
            
            # Générer fichier
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            filename = f"Inventaire_{warehouse.name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
            
            return Response(
                output.read(),
                headers=[
                    ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                    ('Content-Disposition', f'attachment; filename="{filename}"')
                ]
            )
        except Exception as e:
            _logger.error(f"Erreur export Excel entrepôt {warehouse_id}: {e}")
            return Response(f"Erreur lors de l'export: {str(e)}", status=500)


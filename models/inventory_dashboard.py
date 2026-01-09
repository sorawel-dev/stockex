# -*- coding: utf-8 -*-

from odoo import models, fields, api, tools
from odoo.tools import date_utils
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import logging

_logger = logging.getLogger(__name__)


class InventoryDashboard(models.Model):
    _name = 'stockex.inventory.dashboard'
    _description = 'Dashboard Inventaire Stock-INV'
    _rec_name = 'id'
    
    @api.model
    def get_dashboard_data(self, period='30d', valuation_method='standard', warehouse_ids=None, region_ids=None):
        """
        Récupère toutes les données du dashboard pour la période spécifiée.
        
        :param period: '30d', 'ytd', '12m', 'all'
        :param valuation_method: 'standard' ou 'economic'
        :param warehouse_ids: Liste d'IDs entrepôts (None = tous)
        :param region_ids: Liste d'IDs régions (None = toutes)
        :return: dict avec toutes les métriques dashboard
        """
        date_from, date_to = self._get_period_dates(period)
        
        # Construire le domaine de filtrage
        domain = [('state', '=', 'done')]
        if date_from:
            domain.append(('date', '>=', date_from))
        if date_to:
            domain.append(('date', '<=', date_to))
        if warehouse_ids:
            domain.append(('warehouse_id', 'in', warehouse_ids))
        if region_ids:
            domain.append(('warehouse_id.eneo_region_id', 'in', region_ids))
        
        return {
            'period': period,
            'date_from': date_from.strftime('%Y-%m-%d') if date_from else None,
            'date_to': date_to.strftime('%Y-%m-%d') if date_to else None,
            'valuation_method': valuation_method,
            'kpis': self._compute_kpis(domain, valuation_method),
            'charts': self._compute_charts(domain, valuation_method),
            'warehouse_table': self._compute_warehouse_table(domain, valuation_method),
            'top_variances': self._compute_top_variances(domain, valuation_method, limit=10),
        }
    
    @api.model
    def _get_period_dates(self, period):
        """Calcule la plage de dates pour la période donnée."""
        today = fields.Date.context_today(self)
        
        if period == '30d':
            date_from = today - timedelta(days=30)
            date_to = today
        elif period == 'ytd':
            date_from = date_utils.start_of(today, 'year')
            date_to = today
        elif period == '12m':
            date_from = today - relativedelta(months=12)
            date_to = today
        elif period == 'all':
            date_from = None
            date_to = None
        else:
            date_from = today - timedelta(days=30)
            date_to = today
        
        return date_from, date_to
    
    @api.model
    def _preload_product_prices(self, inventories, valuation_method):
        """🚀 OPTIMISATION: Précharge tous les prix produits en masse.
        
        Cette méthode évite le problème N+1 en chargeant tous les prix
        nécessaires en une seule requête par table.
        
        Args:
            inventories: recordset d'inventaires
            valuation_method: 'standard' ou 'economic'
            
        Returns:
            dict: {product_id: price}
        """
        if not inventories:
            return {}
        
        # Extraire tous les product_ids uniques
        product_ids = set()
        for inv in inventories:
            for line in inv.line_ids:
                product_ids.add(line.product_id.id)
        
        if not product_ids:
            return {}
        
        product_ids = list(product_ids)
        price_cache = {}
        
        # Récupérer la société et devise depuis le premier inventaire
        first_inv = inventories[0]
        company = first_inv.company_id
        company_currency = company.currency_id
        conv_date = first_inv.date or fields.Date.today()
        
        # 1) Charger tous les standard_price en une fois
        products = self.env['product.product'].browse(product_ids)
        for product in products:
            std_price = product.standard_price or 0.0
            # Gérer JSONB Odoo 19
            if isinstance(std_price, dict):
                company_key = str(company.id)
                std_price = std_price.get(company_key, 0.0) or std_price.get('1', 0.0) or 0.0
            price_cache[product.id] = std_price
        
        # 2) Si méthode économique, charger les valuation layers ou stock moves
        if valuation_method == 'economic':
            # Essayer d'abord avec valuation layers (plus fiable) si disponible
            try:
                ValuationLayer = self.env['stock.valuation.layer']
                layers = ValuationLayer.search([
                    ('product_id', 'in', product_ids),
                    ('company_id', '=', company.id)
                ], order='product_id, create_date desc')
                
                # Garder seulement la dernière couche par produit
                seen_products = set()
                for layer in layers:
                    if layer.product_id.id not in seen_products:
                        seen_products.add(layer.product_id.id)
                        unit_cost = getattr(layer, 'unit_cost', 0.0) or 0.0
                        if not unit_cost:
                            qty = getattr(layer, 'quantity', 0.0) or 0.0
                            val = getattr(layer, 'value', 0.0) or 0.0
                            unit_cost = (val / qty) if qty else 0.0
                        
                        if unit_cost > 0:
                            price_cache[layer.product_id.id] = unit_cost
            except KeyError:
                # Module stock_account non installé, passer directement aux stock moves
                _logger.info("⚠️ Module stock_account non installé, utilisation des stock.move pour valorisation économique")
                pass
            
            # Compléter avec stock moves pour produits sans layers
            missing_product_ids = [pid for pid in product_ids if price_cache.get(pid, 0) == 0]
            if missing_product_ids:
                StockMove = self.env['stock.move']
                # Rechercher tous les derniers mouvements en une requête
                moves = StockMove.search([
                    ('product_id', 'in', missing_product_ids),
                    ('company_id', '=', company.id),
                    ('state', '=', 'done'),
                    ('price_unit', '>', 0)
                ], order='product_id, date desc')
                
                # Garder seulement le dernier mouvement par produit
                seen_products = set()
                for move in moves:
                    if move.product_id.id not in seen_products:
                        seen_products.add(move.product_id.id)
                        price_cache[move.product_id.id] = move.price_unit
        
        _logger.info("🚀 Dashboard: %d prix préchargés en cache (méthode: %s)", len(price_cache), valuation_method)
        return price_cache
    
    @api.model
    def _compute_kpis(self, domain, valuation_method):
        """Calcule les KPI principaux."""
        Inventory = self.env['stockex.stock.inventory']
        inventories = Inventory.search(domain)
        
        if not inventories:
            return {
                'total_inventories': 0,
                'total_lines': 0,
                'total_products': 0,
                'value_inventoried': 0,
                'value_theoretical': 0,
                'value_variance': 0,
                'variance_rate': 0,
                'lines_with_variance': 0,
                'accuracy_rate': 0,
                'products_unclassified': 0,
            }
        
        # 🚀 OPTIMISATION: Précharger tous les prix en une seule fois
        price_cache = self._preload_product_prices(inventories, valuation_method)
        
        # Calcul avec la méthode de valorisation choisie
        total_value_real = 0
        total_value_theo = 0
        total_lines = 0
        lines_with_variance = 0
        products_set = set()
        
        for inv in inventories:
            for line in inv.line_ids:
                products_set.add(line.product_id.id)
                total_lines += 1
                
                # 🚀 Utiliser le cache au lieu de requêtes répétées
                price = price_cache.get(line.product_id.id, 0.0)
                
                total_value_real += line.product_qty * price
                total_value_theo += line.theoretical_qty * price
                
                if line.difference != 0:
                    lines_with_variance += 1
        
        value_variance = total_value_real - total_value_theo
        variance_rate = (abs(value_variance) / total_value_theo * 100) if total_value_theo > 0 else 0
        accuracy_rate = ((total_lines - lines_with_variance) / total_lines * 100) if total_lines > 0 else 0
        
        # Produits non classés (sans catégorie)
        products_unclassified = self.env['product.product'].search_count([
            ('id', 'in', list(products_set)),
            ('categ_id', '=', False)
        ])
        
        return {
            'total_inventories': len(inventories),
            'total_lines': total_lines,
            'total_products': len(products_set),
            'value_inventoried': total_value_real,
            'value_theoretical': total_value_theo,
            'value_variance': value_variance,
            'variance_rate': variance_rate,
            'lines_with_variance': lines_with_variance,
            'accuracy_rate': accuracy_rate,
            'products_unclassified': products_unclassified,
        }
    
    @api.model
    def _compute_charts(self, domain, valuation_method):
        """🚀 Calcule les données pour les graphiques (optimisé avec cache)."""
        Inventory = self.env['stockex.stock.inventory']
        inventories = Inventory.search(domain)
        
        # 🚀 Précharger les prix une seule fois pour tous les graphiques
        price_cache = self._preload_product_prices(inventories, valuation_method)
        
        return {
            'evolution': self._chart_evolution(inventories, valuation_method, price_cache),
            'warehouse_performance': self._chart_warehouse_performance(inventories, valuation_method, price_cache),
            'category_distribution': self._chart_category_distribution(inventories, valuation_method, price_cache),
            'region_heatmap': self._chart_region_heatmap(inventories, valuation_method, price_cache),
            'warehouse_distribution': self._chart_warehouse_distribution(inventories, valuation_method, price_cache),
            'variance_trend': self._chart_variance_trend(inventories, valuation_method, price_cache),
        }
    
    @api.model
    def _chart_evolution(self, inventories, valuation_method, price_cache):
        """🚀 Graphique : Évolution temporelle (optimisé)."""
        data_by_month = {}
        for inv in inventories:
            month_key = inv.date.strftime('%Y-%m')
            if month_key not in data_by_month:
                data_by_month[month_key] = {'value': 0, 'label': inv.date.strftime('%B %Y')}
            
            for line in inv.line_ids:
                price = price_cache.get(line.product_id.id, 0.0)
                data_by_month[month_key]['value'] += line.product_qty * price
        
        labels = [v['label'] for v in data_by_month.values()]
        values = [v['value'] for v in data_by_month.values()]
        
        return {
            'labels': labels,
            'datasets': [{
                'label': 'Valeur Inventoriée',
                'data': values,
                'borderColor': '#667eea',
                'backgroundColor': 'rgba(102, 126, 234, 0.1)',
                'tension': 0.4,
            }]
        }
    
    @api.model
    def _chart_warehouse_performance(self, inventories, valuation_method, price_cache):
        """🚀 Graphique : Écarts par entrepôt avec distinction positif/négatif (optimisé)."""
        warehouse_data = {}
        for inv in inventories:
            wh_name = inv.warehouse_id.name if inv.warehouse_id else 'Non défini'
            if wh_name not in warehouse_data:
                warehouse_data[wh_name] = {'variance_positive': 0, 'variance_negative': 0}
            
            for line in inv.line_ids:
                price = price_cache.get(line.product_id.id, 0.0)
                variance_value = line.difference * price
                
                if variance_value > 0:
                    warehouse_data[wh_name]['variance_positive'] += variance_value
                else:
                    warehouse_data[wh_name]['variance_negative'] += abs(variance_value)
        
        # Trier par écart total (positif + négatif)
        sorted_warehouses = sorted(
            warehouse_data.items(),
            key=lambda x: x[1]['variance_positive'] + x[1]['variance_negative'],
            reverse=True
        )
        
        labels = [wh[0] for wh in sorted_warehouses]
        variance_positive = [wh[1]['variance_positive'] for wh in sorted_warehouses]
        variance_negative = [wh[1]['variance_negative'] for wh in sorted_warehouses]
        
        return {
            'labels': labels,
            'datasets': [
                {
                    'label': '✅ Surplus (Écarts Positifs)',
                    'data': variance_positive,
                    'backgroundColor': '#10b981',  # Vert
                },
                {
                    'label': '⚠️ Manques (Écarts Négatifs)',
                    'data': variance_negative,
                    'backgroundColor': '#ef4444',  # Rouge
                }
            ]
        }
    
    @api.model
    def _chart_category_distribution(self, inventories, valuation_method, price_cache):
        """🚀 Graphique : Répartition par catégorie produit (optimisé)."""
        category_data = {}
        for inv in inventories:
            for line in inv.line_ids:
                categ_name = line.product_id.categ_id.name if line.product_id.categ_id else 'Non classé'
                if categ_name not in category_data:
                    category_data[categ_name] = 0
                
                price = price_cache.get(line.product_id.id, 0.0)
                category_data[categ_name] += line.product_qty * price
        
        # Trier et limiter aux top 10
        sorted_categories = sorted(category_data.items(), key=lambda x: x[1], reverse=True)[:10]
        
        labels = [cat[0] for cat in sorted_categories]
        values = [cat[1] for cat in sorted_categories]
        
        # Palette de couleurs
        colors = [
            '#667eea', '#764ba2', '#f093fb', '#4facfe',
            '#43e97b', '#fa709a', '#fee140', '#30cfd0',
            '#a8edea', '#fed6e3'
        ]
        
        return {
            'labels': labels,
            'datasets': [{
                'label': 'Valeur Inventoriée',
                'data': values,
                'backgroundColor': colors[:len(labels)],
            }]
        }
    
    @api.model
    def _chart_region_heatmap(self, inventories, valuation_method, price_cache):
        """🔥 Graphique : Heatmap par région électrique avec dégradé de couleurs (optimisé)."""
        region_data = {}
        for inv in inventories:
            region_name = inv.warehouse_id.eneo_region_id.name if inv.warehouse_id and inv.warehouse_id.eneo_region_id else 'Non défini'
            if region_name not in region_data:
                region_data[region_name] = {'variance': 0, 'count': 0}
            
            for line in inv.line_ids:
                price = price_cache.get(line.product_id.id, 0.0)
                variance_value = line.difference * price
                region_data[region_name]['variance'] += abs(variance_value)
                region_data[region_name]['count'] += 1
        
        labels = list(region_data.keys())
        values = [data['variance'] for data in region_data.values()]
        counts = [data['count'] for data in region_data.values()]
        
        # 🌈 Générer un dégradé de couleurs selon l'intensité des écarts
        if values:
            max_variance = max(values) if values else 1
            min_variance = min(values) if values else 0
            
            colors = []
            for value in values:
                # Normaliser entre 0 et 1
                if max_variance > min_variance:
                    intensity = (value - min_variance) / (max_variance - min_variance)
                else:
                    intensity = 0.5
                
                # Dégradé du vert (faible écart) au rouge (fort écart)
                if intensity < 0.33:
                    # Vert à jaune
                    r = int(76 + (255 - 76) * (intensity / 0.33))
                    g = int(175 + (235 - 175) * (intensity / 0.33))
                    b = int(80 + (59 - 80) * (intensity / 0.33))
                elif intensity < 0.66:
                    # Jaune à orange
                    adj_intensity = (intensity - 0.33) / 0.33
                    r = int(255)
                    g = int(235 - (68 * adj_intensity))
                    b = int(59 - (59 * adj_intensity))
                else:
                    # Orange à rouge
                    adj_intensity = (intensity - 0.66) / 0.34
                    r = int(255 - (22 * adj_intensity))
                    g = int(167 - (99 * adj_intensity))
                    b = int(0)
                
                colors.append(f'rgb({r}, {g}, {b})')
        else:
            colors = ['#4caf50']  # Vert par défaut si pas de données
        
        return {
            'labels': labels,
            'datasets': [{
                'label': 'Écart Total (FCFA)',
                'data': values,
                'backgroundColor': colors,  # 🌈 Heatmap colors
            }],
            'counts': counts,
        }
    
    @api.model
    def _chart_warehouse_distribution(self, inventories, valuation_method, price_cache):
        """🚀 Graphique : Répartition valeur par entrepôt (optimisé)."""
        warehouse_data = {}
        for inv in inventories:
            wh_id = inv.warehouse_id.id if inv.warehouse_id else False
            wh_name = inv.warehouse_id.name if inv.warehouse_id else 'Non défini'
            if wh_id not in warehouse_data:
                warehouse_data[wh_id] = {
                    'name': wh_name,
                    'value_theoretical': 0,
                    'value_inventoried': 0,
                    'variance_value': 0
                }
            
            for line in inv.line_ids:
                price = price_cache.get(line.product_id.id, 0.0)
                
                warehouse_data[wh_id]['value_theoretical'] += line.theoretical_qty * price
                warehouse_data[wh_id]['value_inventoried'] += line.product_qty * price
                warehouse_data[wh_id]['variance_value'] += line.difference * price
        
        labels = [data['name'] for data in warehouse_data.values()]
        warehouse_ids = list(warehouse_data.keys())
        value_theo_data = [data['value_theoretical'] for data in warehouse_data.values()]
        value_real_data = [data['value_inventoried'] for data in warehouse_data.values()]
        value_variance_data = [abs(data['variance_value']) for data in warehouse_data.values()]
        
        return {
            'labels': labels,
            'warehouse_ids': warehouse_ids,  # 🎯 Ajouter les IDs pour le clic
            'datasets': [
                {
                    'label': 'Valeur Théorique (FCFA)',
                    'data': value_theo_data,
                    'backgroundColor': '#8b5cf6',
                },
                {
                    'label': 'Valeur Inventoriée (FCFA)',
                    'data': value_real_data,
                    'backgroundColor': '#10b981',
                },
                {
                    'label': 'Écart Valeur (FCFA)',
                    'data': value_variance_data,
                    'backgroundColor': '#ef4444',
                }
            ]
        }
    
    @api.model
    def _chart_variance_trend(self, inventories, valuation_method, price_cache):
        """🚀 Graphique : Tendance des écarts dans le temps (optimisé)."""
        data_by_month = {}
        for inv in inventories:
            month_key = inv.date.strftime('%Y-%m')
            if month_key not in data_by_month:
                data_by_month[month_key] = {
                    'label': inv.date.strftime('%B %Y'),
                    'variance': 0,
                    'lines': 0,
                    'lines_variance': 0
                }
            
            for line in inv.line_ids:
                data_by_month[month_key]['lines'] += 1
                if line.difference != 0:
                    data_by_month[month_key]['lines_variance'] += 1
                    price = price_cache.get(line.product_id.id, 0.0)
                    data_by_month[month_key]['variance'] += abs(line.difference * price)
        
        labels = [v['label'] for v in data_by_month.values()]
        variance_values = [v['variance'] for v in data_by_month.values()]
        variance_rates = [
            (v['lines_variance'] / v['lines'] * 100) if v['lines'] > 0 else 0
            for v in data_by_month.values()
        ]
        
        return {
            'labels': labels,
            'datasets': [
                {
                    'label': 'Écart Total (FCFA)',
                    'data': variance_values,
                    'borderColor': '#ef4444',
                    'backgroundColor': 'rgba(239, 68, 68, 0.1)',
                    'yAxisID': 'y',
                    'type': 'line',
                    'tension': 0.4,
                },
                {
                    'label': 'Taux d\'écart (%)',
                    'data': variance_rates,
                    'borderColor': '#f59e0b',
                    'backgroundColor': 'rgba(245, 158, 11, 0.1)',
                    'yAxisID': 'y1',
                    'type': 'line',
                    'tension': 0.4,
                }
            ]
        }
    
    @api.model
    def _compute_warehouse_table(self, domain, valuation_method):
        """🚀 Tableau : Répartition détaillée par entrepôt (optimisé)."""
        Inventory = self.env['stockex.stock.inventory']
        inventories = Inventory.search(domain)
        
        # 🚀 Précharger les prix
        price_cache = self._preload_product_prices(inventories, valuation_method)
        
        warehouse_data = {}
        for inv in inventories:
            wh_id = inv.warehouse_id.id if inv.warehouse_id else False
            wh_name = inv.warehouse_id.name if inv.warehouse_id else 'Non défini'
            
            if wh_id not in warehouse_data:
                warehouse_data[wh_id] = {
                    'id': wh_id,
                    'name': wh_name,
                    'qty_theo': 0,
                    'qty_real': 0,
                    'qty_variance': 0,
                    'value_theo': 0,
                    'value_real': 0,
                    'value_variance': 0,
                    'lines': 0,
                    'lines_variance': 0,
                }
            
            for line in inv.line_ids:
                price = price_cache.get(line.product_id.id, 0.0)
                
                warehouse_data[wh_id]['qty_theo'] += line.theoretical_qty
                warehouse_data[wh_id]['qty_real'] += line.product_qty
                warehouse_data[wh_id]['qty_variance'] += line.difference
                warehouse_data[wh_id]['value_theo'] += line.theoretical_qty * price
                warehouse_data[wh_id]['value_real'] += line.product_qty * price
                warehouse_data[wh_id]['value_variance'] += line.difference * price
                warehouse_data[wh_id]['lines'] += 1
                
                if line.difference != 0:
                    warehouse_data[wh_id]['lines_variance'] += 1
        
        # Ajouter le statut basé sur le taux d'écart
        rows = []
        for data in warehouse_data.values():
            variance_rate = (data['lines_variance'] / data['lines'] * 100) if data['lines'] > 0 else 0
            
            if variance_rate < 5:
                status = 'excellent'
                status_label = '✅ Excellent'
            elif variance_rate < 10:
                status = 'good'
                status_label = '✓ Bon'
            elif variance_rate < 20:
                status = 'warning'
                status_label = '⚠ Attention'
            else:
                status = 'critical'
                status_label = '🔴 Critique'
            
            data['status'] = status
            data['status_label'] = status_label
            data['variance_rate'] = variance_rate
            rows.append(data)
        
        # Trier par valeur d'écart décroissante
        rows.sort(key=lambda x: abs(x['value_variance']), reverse=True)
        
        return rows
    
    @api.model
    def _compute_top_variances(self, domain, valuation_method, limit=10):
        """🚀 Top N des plus gros écarts (optimisé)."""
        Inventory = self.env['stockex.stock.inventory']
        inventories = Inventory.search(domain)
        
        # 🚀 Précharger les prix
        price_cache = self._preload_product_prices(inventories, valuation_method)
        
        variances = []
        for inv in inventories:
            for line in inv.line_ids:
                if line.difference == 0:
                    continue
                
                price = price_cache.get(line.product_id.id, 0.0)
                variance_value = line.difference * price
                
                variances.append({
                    'product_name': line.product_id.name,
                    'product_code': line.product_id.default_code or '',
                    'warehouse': inv.warehouse_id.name if inv.warehouse_id else 'N/A',
                    'qty_theo': line.theoretical_qty,
                    'qty_real': line.product_qty,
                    'qty_variance': line.difference,
                    'value_variance': variance_value,
                    'inventory_ref': inv.name,
                    'inventory_date': inv.date.strftime('%d/%m/%Y'),
                })
        
        # Trier par valeur absolue d'écart
        variances.sort(key=lambda x: abs(x['value_variance']), reverse=True)
        
        return variances[:limit]
    
    @api.model
    def export_warehouse_table_excel(self, period='30d', valuation_method='standard', warehouse_ids=None, region_ids=None):
        """Export Excel du tableau entrepôts."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            import base64
            from io import BytesIO
        except ImportError:
            return {'error': 'La bibliothèque openpyxl est requise pour l\'export Excel.'}
        
        # Récupérer les données
        date_from, date_to = self._get_period_dates(period)
        domain = [('state', '=', 'done')]
        if date_from:
            domain.append(('date', '>=', date_from))
        if date_to:
            domain.append(('date', '<=', date_to))
        if warehouse_ids:
            domain.append(('warehouse_id', 'in', warehouse_ids))
        if region_ids:
            domain.append(('warehouse_id.region_id', 'in', region_ids))
        
        table_data = self._compute_warehouse_table(domain, valuation_method)
        
        # Créer le workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Répartition Entrepôts"
        
        # Styles
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # En-tête
        ws['A1'] = f"RÉPARTITION PAR ENTREPÔT - Dashboard Inventaire StockEx"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:I1')
        
        ws['A2'] = f"Période: {period} | Valorisation: {valuation_method.upper()}"
        ws['A2'].font = Font(size=10)
        ws.merge_cells('A2:I2')
        
        # Headers
        headers = [
            'Entrepôt', 'Qté Théo', 'Qté Réelle', 'Écart Qté',
            'Valeur Théo (FCFA)', 'Valeur Réelle (FCFA)', 'Écart Valeur (FCFA)',
            'Taux Écart (%)', 'Statut'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Données
        row_num = 5
        for data in table_data:
            ws.cell(row=row_num, column=1, value=data['name'])
            ws.cell(row=row_num, column=2, value=data['qty_theo'])
            ws.cell(row=row_num, column=3, value=data['qty_real'])
            ws.cell(row=row_num, column=4, value=data['qty_variance'])
            ws.cell(row=row_num, column=5, value=data['value_theo'])
            ws.cell(row=row_num, column=6, value=data['value_real'])
            ws.cell(row=row_num, column=7, value=data['value_variance'])
            ws.cell(row=row_num, column=8, value=round(data['variance_rate'], 2))
            ws.cell(row=row_num, column=9, value=data['status_label'])
            
            # Bordures
            for col in range(1, 10):
                ws.cell(row=row_num, column=col).border = border
            
            row_num += 1
        
        # Ajuster largeur colonnes
        ws.column_dimensions['A'].width = 25
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col].width = 18
        ws.column_dimensions['I'].width = 15
        
        # Générer fichier
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        file_data = base64.b64encode(output.read())
        
        filename = f"Repartition_Entrepots_{period}_{valuation_method}_{fields.Date.today().strftime('%Y%m%d')}.xlsx"
        
        # Créer l'attachement
        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'type': 'binary',
            'datas': file_data,
            'res_model': 'stockex.inventory.dashboard',
            'res_id': 0,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

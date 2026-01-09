# -*- coding: utf-8 -*-

import logging

from odoo import models, fields, api
from odoo.exceptions import UserError
from markupsafe import Markup

_logger = logging.getLogger(__name__)

class ProductTemplate(models.Model):
    _inherit = 'product.template'
    brand = fields.Char(string='Marque')
    has_serial = fields.Boolean(string="A un N° série", help="Indique si l'article possède un numéro de série")


class ProductProduct(models.Model):
    _inherit = 'product.product'
    
    stock_value = fields.Monetary(
        string='Coût total',
        compute='_compute_stock_value',
        currency_field='currency_id',
        store=False,
        help='Valeur totale du stock = Quantité disponible × Coût unitaire'
    )
    
    @api.depends('qty_available', 'standard_price')
    def _compute_stock_value(self):
        """Calcule la valeur totale du stock (quantité × coût unitaire)."""
        for product in self:
            product.stock_value = (product.qty_available or 0.0) * (product.standard_price or 0.0)


class StockWarehouse(models.Model):
    """Héritage de stock.warehouse pour ajouter une hiérarchie d'entrepôts"""
    _inherit = 'stock.warehouse'
    
    # Redéfinition du champ code natif pour augmenter la taille à 6 caractères
    code = fields.Char(
        string='Diminutif',
        size=6,
        required=True,
        help='Diminutif de l\'entrepôt (6 caractères maximum)'
    )
    
    # Code entrepôt (ancien "nom court")
    warehouse_code = fields.Char(
        string='Code Entrepôt',
        size=10,
        index=True,
        help='Code unique de l\'entrepôt (ex: WH-ABJ-001, WH-YOP-002)'
    )
    
    parent_id = fields.Many2one(
        comodel_name='stock.warehouse',
        string='Entrepôt Parent',
        ondelete='restrict',
        index=True,
        help='Entrepôt parent pour créer une hiérarchie d\'entrepôts'
    )
    
    parent_name_upper = fields.Char(
        string='Magasin Parent',
        compute='_compute_parent_name_upper',
        store=False,
        help='Nom de l\'entrepôt parent en MAJUSCULES'
    )
    
    child_ids = fields.One2many(
        comodel_name='stock.warehouse',
        inverse_name='parent_id',
        string='Entrepôts Enfants',
        help='Entrepôts dépendant de cet entrepôt'
    )
    child_count = fields.Integer(
        string='Nombre d\'enfants',
        compute='_compute_child_count',
        store=True
    )
    
    warehouse_type = fields.Selection(
        selection=[
            ('production', 'Production'),
            ('distribution', 'Distribution'),
            ('commercialisation', 'Commercialisation'),
        ],
        string='Type de Magasin',
        required=True,
        default='distribution',
        index=True,
        help='Type de magasin : Production, Distribution ou Commercialisation'
    )
    
    # Champ color pour compatibilité avec les vues kanban Odoo standard
    color = fields.Integer(
        string='Couleur',
        default=0,
        help='Couleur pour l\'affichage dans les vues kanban'
    )
    
    # Champs de géolocalisation
    latitude = fields.Float(
        string='Latitude',
        digits=(10, 7),
        help='Latitude GPS de l\'entrepôt (ex: 5.3599517)'
    )
    longitude = fields.Float(
        string='Longitude',
        digits=(10, 7),
        help='Longitude GPS de l\'entrepôt (ex: -4.0082563)'
    )
    coordinates = fields.Char(
        string='Coordonnées GPS',
        compute='_compute_coordinates',
        store=False,
        help='Coordonnées au format "Latitude, Longitude"'
    )
    google_maps_url = fields.Char(
        string='Lien Google Maps',
        compute='_compute_google_maps_url',
        store=False
    )
    
    # Informations de contact
    address = fields.Text(string='Adresse complète')
    city = fields.Char(string='Ville')
    phone = fields.Char(string='Téléphone')
    email = fields.Char(string='Email')
    
    # Région électrique ENEO
    eneo_region_id = fields.Many2one(
        'stockex.eneo.region',
        string='Région Électrique ENEO',
        index=True,
        help='Région électrique ENEO à laquelle appartient cet entrepôt'
    )
    
    eneo_region_code = fields.Char(
        related='eneo_region_id.code',
        string='Code Région ENEO',
        store=True,
        readonly=True
    )
    
    eneo_network = fields.Selection(
        related='eneo_region_id.network',
        string='Réseau ENEO',
        store=True,
        readonly=True
    )
    
    # Smart Buttons & valeurs
    currency_id = fields.Many2one(
        'res.currency',
        string='Devise',
        related='company_id.currency_id',
        store=True,
        readonly=True,
    )
    stock_value_fcfa = fields.Monetary(
        string='Valeur du stock (FCFA)',
        currency_field='currency_id',
        compute='_compute_stock_value_fcfa',
        store=False,
    )
    inventory_variance_fcfa = fields.Monetary(
        string='Écart Inventaire',
        currency_field='currency_id',
        compute='_compute_inventory_variance_fcfa',
        store=False,
    )
    move_count = fields.Integer(
        string='Mouvements',
        compute='_compute_move_count',
        store=False,
    )
    quant_count = fields.Integer(
        string='Articles en stock',
        compute='_compute_quant_count',
        store=False,
    )
    product_ref_count = fields.Integer(
        string='Références d\'articles',
        compute='_compute_product_ref_count',
        store=False,
        help='Nombre de références de produits uniques en stock'
    )

    location_count = fields.Integer(
        string='Emplacements internes',
        compute='_compute_location_count',
        store=False,
    )
    
    def _compute_quant_count(self):
        for wh in self:
            root = wh.view_location_id.id if wh.view_location_id else False
            domain = [('location_id', 'child_of', root), ('quantity', '>', 0)] if root else [('id', '=', 0)]
            wh.quant_count = self.env['stock.quant'].search_count(domain)
    
    def _compute_location_count(self):
        for wh in self:
            root = wh.view_location_id.id if wh.view_location_id else False
            if root:
                wh.location_count = self.env['stock.location'].search_count([('id', 'child_of', root), ('usage', '=', 'internal')])
            else:
                wh.location_count = 0
    
    def _compute_product_ref_count(self):
        """Calcule le nombre de références (produits uniques) en stock."""
        for wh in self:
            root = wh.view_location_id.id if wh.view_location_id else False
            if root:
                # Compter les produits uniques avec quantité > 0
                quants = self.env['stock.quant'].search([('location_id', 'child_of', root), ('quantity', '>', 0)])
                wh.product_ref_count = len(quants.mapped('product_id'))
            else:
                wh.product_ref_count = 0
    
    def _compute_move_count(self):
        for wh in self:
            root = wh.view_location_id.id if wh.view_location_id else False
            domain = ['|', ('location_id', 'child_of', root), ('location_dest_id', 'child_of', root)] if root else [('id', '=', 0)]
            domain = domain + [('state', '=', 'done')]
            wh.move_count = self.env['stock.move'].search_count(domain)
    
    def _compute_inventory_variance_fcfa(self):
        for wh in self:
            invs = self.env['stockex.stock.inventory'].search([('warehouse_id', '=', wh.id), ('state', '=', 'done')])
            wh.inventory_variance_fcfa = sum(invs.mapped('total_value_difference'))

    def _compute_stock_value_fcfa(self):
        for wh in self:
            total = 0.0
            root = wh.view_location_id.id if wh.view_location_id else False
            quants = self.env['stock.quant'].search([('location_id', 'child_of', root), ('quantity', '>', 0)]) if root else []
            ICP = self.env['ir.config_parameter'].sudo()
            rule = ICP.get_param('stockex.valuation_rule', 'standard')
            
            # Odoo 19: stock.valuation.layer n'existe plus, utiliser stock_valuation_layer_ids depuis stock.move
            for q in quants:
                price = 0.0
                if rule == 'economic':
                    # Chercher les couches de valorisation via les mouvements de stock
                    try:
                        moves = self.env['stock.move'].search([
                            ('product_id', '=', q.product_id.id),
                            ('company_id', '=', wh.company_id.id),
                            ('state', '=', 'done')
                        ], limit=1, order='date desc')
                        if moves and hasattr(moves, 'price_unit'):
                            price = moves.price_unit or q.product_id.standard_price or 0.0
                        else:
                            price = q.product_id.standard_price or 0.0
                    except Exception:
                        # Fallback sur le prix standard
                        price = q.product_id.standard_price or 0.0
                else:
                    price = q.product_id.standard_price or 0.0
                total += (q.quantity or 0.0) * price
            wh.stock_value_fcfa = total
    
    def action_open_stock_value(self):
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'name': 'Valorisation du stock',
            'res_model': 'product.product',
            'view_mode': 'list,form',
            'views': [(self.env.ref('stockex.product_product_tree_stock_valuation').id, 'list')],
            'domain': [('qty_available', '!=', 0)],
            'context': {
                'search_default_real_stock_available': 1,
                'location': self.view_location_id.id,
                'group_by': 'categ_id',
            },
        }
    
    def action_open_moves(self):
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'name': 'Mouvements de stock',
            'res_model': 'stock.move',
            'view_mode': 'list,form,pivot,graph',
            'domain': ['|', ('location_id', 'child_of', self.view_location_id.id), ('location_dest_id', 'child_of', self.view_location_id.id)],
            'context': {
                'search_default_done': 1,
            },
        }
    
    def action_open_quants(self):
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'name': 'Articles en stock',
            'res_model': 'stock.quant',
            'view_mode': 'list,form',
            'domain': [('location_id', 'child_of', self.view_location_id.id), ('quantity', '>', 0)],
            'context': {
                'search_default_productgroup': 1,
                'search_default_locationgroup': 1,
            },
        }
    
    def _generate_warehouse_code(self, name):
        """
        Génère un diminutif intelligent du nom de l'entrepôt.
        
        Args:
            name (str): Nom de l'entrepôt
            
        Returns:
            str: Code diminutif (max 6 caractères)
        
        Exemples:
            - "Abidjan" → "ABIDJA"
            - "Entrepôt Central" → "EC"
            - "Grand Bassam Site Nord" → "GBSN"
        """
        if not name:
            return 'WH'
        
        name = name.strip()
        words = name.split()
        
        if len(words) == 1:
            # Un seul mot : prendre les 6 premiers caractères
            code = name[:6].upper()
        else:
            # Plusieurs mots : prendre la première lettre de chaque mot (max 6)
            code = ''.join([word[0].upper() for word in words[:6] if word])
            # Si le code est trop court, compléter avec les premières lettres du premier mot
            if len(code) < 3 and words:
                code = (words[0][:3] + code).upper()[:6]
        
        return code
    
    @api.onchange('name')
    def _onchange_name_generate_code(self):
        """Génère automatiquement le code (diminutif) lors de la saisie du nom."""
        if self.name and not self.code:
            self.code = self._generate_warehouse_code(self.name)
    
    @api.model_create_multi
    def create(self, vals_list):
        """Génère automatiquement le code si non fourni lors de la création."""
        for vals in vals_list:
            if vals.get('name') and not vals.get('code'):
                vals['code'] = self._generate_warehouse_code(vals['name'])
        return super().create(vals_list)
    
    def write(self, vals):
        """Met à jour le code si le nom change et que le code n'est pas explicitement fourni."""
        if 'name' in vals and 'code' not in vals:
            for warehouse in self:
                # Regénérer le code si le nom change
                new_code = self._generate_warehouse_code(vals['name'])
                vals['code'] = new_code
        return super().write(vals)
    
    @api.depends('parent_id', 'parent_id.name')
    def _compute_parent_name_upper(self):
        """Calcule le nom du parent en MAJUSCULES."""
        for warehouse in self:
            if warehouse.parent_id:
                warehouse.parent_name_upper = warehouse.parent_id.name.upper()
            else:
                warehouse.parent_name_upper = False
    
    @api.depends('child_ids')
    def _compute_child_count(self):
        """Calcule le nombre d'entrepôts enfants."""
        for warehouse in self:
            warehouse.child_count = len(warehouse.child_ids)
    
    @api.depends('latitude', 'longitude')
    def _compute_coordinates(self):
        """Calcule les coordonnées GPS au format texte."""
        for warehouse in self:
            if warehouse.latitude and warehouse.longitude:
                warehouse.coordinates = f"{warehouse.latitude}, {warehouse.longitude}"
            else:
                warehouse.coordinates = False
    
    @api.depends('latitude', 'longitude')
    def _compute_google_maps_url(self):
        """Génère l'URL Google Maps."""
        for warehouse in self:
            if warehouse.latitude and warehouse.longitude:
                warehouse.google_maps_url = f"https://www.google.com/maps?q={warehouse.latitude},{warehouse.longitude}"
            else:
                warehouse.google_maps_url = False
    
    def action_open_map(self):
        """Ouvre Google Maps dans un nouvel onglet."""
        self.ensure_one()
        if not self.google_maps_url:
            raise UserError("Veuillez renseigner les coordonnées GPS de cet entrepôt.")
        return {
            'type': 'ir.actions.act_url',
            'url': self.google_maps_url,
            'target': 'new',
        }


class StockInventory(models.Model):
    _name = 'stockex.stock.inventory'
    _description = 'Inventaire de Stock'
    _inherit = ['mail.thread', 'mail.activity.mixin']
    _order = 'date desc, id desc'
    _rec_name = 'name'
    _rec_names_search = ['name']

    # Champ requis pour le widget badge
    color = fields.Integer(string='Couleur', default=0)

    name = fields.Char(
        string='Référence',
        required=True,
        default='Nouveau',
        index=True,
        copy=False,
        tracking=True
    )
    date = fields.Date(
        string='Date',
        required=True,
        default=fields.Date.today,
        index=True,
        tracking=True
    )
    state = fields.Selection(
        selection=[
            ('draft', 'Brouillon'),
            ('in_progress', 'En cours'),
            ('pending_approval', 'En attente d\'approbation'),
            ('approved', 'Approuvé'),
            ('done', 'Validé'),
            ('cancel', 'Annulé')
        ],
        string='État',
        default='draft',
        required=True,
        copy=False,
        tracking=True,
        index=True
    )
    is_initial_stock = fields.Boolean(
        string='Stock Initial',
        default=False,
        help='Cocher si c\'est un inventaire de stock initial (base vide). Les écarts ne seront pas comptabilisés dans les statistiques.',
        tracking=True
    )
    description = fields.Text(string='Notes')
    location_id = fields.Many2one(
        comodel_name='stock.location',
        string='Emplacement',
        index=True,
        tracking=True
    )
    warehouse_id = fields.Many2one(
        comodel_name='stock.warehouse',
        string='Entrepôt',
        index=True,
        tracking=True,
        help='Entrepôt principal de cet inventaire'
    )
    company_id = fields.Many2one(
        comodel_name='res.company',
        string='Société',
        required=True,
        default=lambda self: self.env.company,
        index=True
    )
    user_id = fields.Many2one(
        comodel_name='res.users',
        string='Responsable',
        required=True,
        default=lambda self: self.env.user,
        index=True,
        tracking=True
    )
    approver_id = fields.Many2one(
        comodel_name='res.users',
        string='Approbateur',
        tracking=True,
        help='Utilisateur qui a approuvé l\'inventaire'
    )
    approval_date = fields.Datetime(
        string='Date d\'approbation',
        readonly=True,
        tracking=True
    )
    validator_id = fields.Many2one(
        comodel_name='res.users',
        string='Validateur',
        tracking=True,
        help='Utilisateur qui a validé l\'inventaire'
    )
    validation_date = fields.Datetime(
        string='Date de validation',
        readonly=True,
        tracking=True
    )
    line_ids = fields.One2many(
        comodel_name='stockex.stock.inventory.line',
        inverse_name='inventory_id',
        string='Lignes'
    )
    account_move_ids = fields.One2many(
        comodel_name='account.move',
        inverse_name='stockex_inventory_id',
        string='Écritures Comptables',
        readonly=True,
        help='Écritures comptables générées par la validation de cet inventaire'
    )
    account_move_count = fields.Integer(
        string='Nombre d\'écritures',
        compute='_compute_account_move_count',
        store=True
    )
    
    # Mouvements de stock liés (traçabilité)
    stock_move_ids = fields.One2many(
        comodel_name='stock.move',
        inverse_name='stockex_inventory_id',
        string='Mouvements de Stock',
        readonly=True,
    )
    stock_move_count = fields.Integer(
        string='Nombre de mouvements',
        compute='_compute_stock_move_count',
        store=True
    )
    variance_count = fields.Integer(
        string='Nombre d\'écarts',
        compute='_compute_variance_count',
        store=True
    )
    
    # Synchronisation avec le stock natif Odoo
    sync_to_native = fields.Boolean(
        string='Synchronisé avec Odoo Natif',
        default=False,
        readonly=True,
        help='Indique si les quantités ont été synchronisées avec les quants Odoo natifs'
    )
    sync_date = fields.Datetime(
        string='Date de synchronisation',
        readonly=True,
        help='Date de la dernière synchronisation avec les quants Odoo'
    )
    
    # Totaux inventaire
    total_quantity_real = fields.Float(string='Quantité réelle totale', compute='_compute_totals', digits='Product Unit of Measure')
    total_quantity_theoretical = fields.Float(string='Quantité théorique totale', compute='_compute_totals', digits='Product Unit of Measure')
    total_quantity_difference = fields.Float(string='Écart de quantité total', compute='_compute_totals', digits='Product Unit of Measure')
    total_value_real = fields.Float(string='Valeur totale réelle', compute='_compute_totals', digits='Product Price')
    total_value_theoretical = fields.Float(string='Valeur totale théorique', compute='_compute_totals', digits='Product Price')
    total_value_difference = fields.Float(
        string='Valeur totale des écarts',
        compute='_compute_value_difference',
        digits='Product Price',
        help='Différence entre la valeur inventoriée et la valeur réelle du stock Odoo'
    )
    
    # Valeur réelle du stock Odoo (tous produits de l'emplacement/entrepôt)
    odoo_stock_value = fields.Float(
        string='Valeur Stock Odoo',
        compute='_compute_odoo_stock_value',
        digits='Product Price',
        help='Valeur totale du stock dans Odoo pour l\'emplacement/entrepôt de cet inventaire'
    )
    
    company_currency_id = fields.Many2one(comodel_name='res.currency', string='Devise', related='company_id.currency_id', store=False)
    
    # Affichages formatés FCFA (sans décimales)
    display_odoo_stock_value = fields.Char(string='Valeur Stock Odoo (affichage)', compute='_compute_display_values')
    display_total_value_real = fields.Char(string='Valeur réelle (affichage)', compute='_compute_display_values')
    display_total_value_difference = fields.Char(string='Écart de valeur (affichage)', compute='_compute_display_values')
    
    def _format_fcfa(self, value):
        try:
            n = int(round(value or 0))
        except Exception:
            n = 0
        s = f"{abs(n):,}".replace(',', ' ')
        return (('-' + s) if n < 0 else s) + ' FCFA'
    
    @api.depends('odoo_stock_value', 'total_value_real', 'total_value_difference')
    def _compute_display_values(self):
        for inv in self:
            inv.display_odoo_stock_value = self._format_fcfa(inv.odoo_stock_value)
            inv.display_total_value_real = self._format_fcfa(inv.total_value_real)
            inv.display_total_value_difference = self._format_fcfa(inv.total_value_difference)

    
    _name_company_uniq = models.UniqueIndex("(name, company_id)")
    
    @api.depends('account_move_ids')
    def _compute_account_move_count(self):
        """Calcule le nombre d'écritures comptables."""
        for inventory in self:
            inventory.account_move_count = len(inventory.account_move_ids)
    
    @api.depends('stock_move_ids')
    def _compute_stock_move_count(self):
        """Calcule le nombre de mouvements de stock."""
        for inventory in self:
            inventory.stock_move_count = len(inventory.stock_move_ids)
    
    @api.depends('line_ids.difference', 'line_ids.product_qty', 'line_ids.theoretical_qty')
    def _compute_variance_count(self):
        """Calcule le nombre de lignes avec écart."""
        for inventory in self:
            inventory.variance_count = len(inventory.line_ids.filtered(lambda l: (l.difference or 0.0) != 0.0))
    
    def action_view_stock_moves(self):
        """Ouvre la vue des mouvements de stock liés à l'inventaire."""
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'name': f"Mouvements - {self.name}",
            'res_model': 'stock.move',
            'view_mode': 'list,form',
            'domain': [('stockex_inventory_id', '=', self.id)],
            'context': {
                'search_default_done': 1,
                'group_by': 'product_id',
            },
        }
    
    def action_view_variances(self):
        """Ouvre la liste des écarts de cet inventaire."""
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'name': f"Écarts - {self.name}",
            'res_model': 'stockex.stock.inventory.line',
            'view_mode': 'list,form',
            'domain': [('inventory_id', '=', self.id), ('difference', '!=', 0.0)],
            'context': {
                'tree_view_ref': 'stockex.view_stockex_inventory_line_variance_list',
                'search_default_group_by_product': 1,
                'group_by': 'product_id',
            },
        }
    
    def action_view_account_moves(self):
        """Ouvre la vue des écritures comptables liées à l'inventaire."""
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'name': f"Écritures Comptables - {self.name}",
            'res_model': 'account.move',
            'view_mode': 'list,form',
            'domain': [('id', 'in', self.account_move_ids.ids)],
            'context': {
                'search_default_posted': 1,
                'allow_inventory_attachment': True,
                'default_res_model': self._name,
                'default_res_id': self.id,
                'default_type': 'url',
            },
        }
    
    
    def action_view_documents(self):
        """Ouvre la liste des pièces jointes (documents) liées à l'inventaire."""
        self.ensure_one()
        view_form = self.env.ref('stockex.view_ir_attachment_inventory_minimal', raise_if_not_found=False)
        view_kanban = self.env.ref('stockex.view_ir_attachment_inventory_kanban', raise_if_not_found=False)
        list_view = self.env.ref('stockex.view_ir_attachment_inventory_list', raise_if_not_found=False)
        action = {
            'type': 'ir.actions.act_window',
            'name': f"Documents - {self.name}",
            'res_model': 'ir.attachment',
            'view_mode': 'kanban,list,form',
            'domain': [('res_model', '=', self._name), ('res_id', '=', self.id)],
            'context': {
                'default_res_model': self._name,
                'default_res_id': self.id,
                'allow_inventory_attachment': True,
            },
        }
        views_list = []
        if view_kanban:
            views_list.append((view_kanban.id, 'kanban'))
        if list_view:
            views_list.append((list_view.id, 'list'))
        else:
            views_list.append((False, 'list'))
        if view_form:
            views_list.append((view_form.id, 'form'))
        action['views'] = views_list
        return action
    
    
    def action_attach_document(self):
        """Ouvre le formulaire d'attachement (URL par défaut) pour cet inventaire."""
        self.ensure_one()
        view = self.env.ref('stockex.view_ir_attachment_inventory_minimal', raise_if_not_found=False)
        action = {
            'type': 'ir.actions.act_window',
            'name': f"Joindre un document - {self.name}",
            'res_model': 'ir.attachment',
            'view_mode': 'form',
            'target': 'new',
            'context': {
                'allow_inventory_attachment': True,
                'default_res_model': self._name,
                'default_res_id': self.id,
            },
        }
        if view:
            action['views'] = [(view.id, 'form')]
        return action

class IrAttachment(models.Model):
    _inherit = 'ir.attachment'

    # Champs MinIO
    minio_bucket = fields.Char(string='MinIO Bucket', readonly=True)
    minio_object_name = fields.Char(string='MinIO Object Path', readonly=True)
    use_minio = fields.Boolean(string='Stocké sur MinIO', default=False, readonly=True)

    def init(self):
        # Crée les colonnes si elles n'existent pas encore (évite l'erreur avant upgrade)
        try:
            self.env.cr.execute("ALTER TABLE ir_attachment ADD COLUMN IF NOT EXISTS minio_bucket varchar")
            self.env.cr.execute("ALTER TABLE ir_attachment ADD COLUMN IF NOT EXISTS minio_object_name varchar")
            self.env.cr.execute("ALTER TABLE ir_attachment ADD COLUMN IF NOT EXISTS use_minio boolean DEFAULT false")
        except Exception as e:
            _logger.warning(f"Échec création colonnes MinIO: {e}")

    @api.onchange('datas', 'url')
    def _onchange_datas_url(self):
        """Génère automatiquement le titre lors de la sélection du fichier ou saisie de l'URL."""
        if self.env.context.get('default_res_model') == 'stockex.stock.inventory' and not self.name:
            from datetime import date
            today = date.today().strftime('%Y-%m-%d')
            if self.datas:
                # Fichier: utiliser 'document' comme base
                self.name = f"document-{today}"
            elif self.url:
                # URL: extraire le dernier segment ou utiliser 'lien'
                url_parts = self.url.strip('/').split('/')
                url_name = url_parts[-1] if url_parts else 'lien'
                self.name = f"{url_name}-{today}"
    
    @api.model
    def create(self, vals_list):
        """Surcharge pour empêcher l'ajout de pièces jointes via le chatter."""
        # Odoo 19: vals_list est toujours une liste
        if not isinstance(vals_list, list):
            vals_list = [vals_list]
        
        for vals in vals_list:
            # Interdire la création d'attachements via le chatter pour les inventaires,
            # sauf si l'action "Documents" ou "Joindre un document" l'autorise explicitement.
            if vals.get('res_model') == 'stockex.stock.inventory' and not self.env.context.get('allow_inventory_attachment'):
                raise UserError("Veuillez ajouter les documents via le bouton 'Documents' ou 'Joindre un document'.\nPréférez les liens (URL) plutôt que des fichiers lourds.")
            
            # Générer automatiquement le titre si non fourni
            if vals.get('res_model') == 'stockex.stock.inventory' and not vals.get('name'):
                from datetime import date
                today = date.today().strftime('%Y-%m-%d')
                if vals.get('datas'):
                    # Fichier: utiliser le nom fourni ou 'document'
                    filename = vals.get('name', 'document') if vals.get('name') else 'document'
                    vals['name'] = f"{filename}-{today}"
                elif vals.get('url'):
                    # URL: extraire le dernier segment ou utiliser 'lien'
                    url_parts = vals['url'].strip('/').split('/')
                    url_name = url_parts[-1] if url_parts else 'lien'
                    vals['name'] = f"{url_name}-{today}"
                else:
                    vals['name'] = f"document-{today}"
        
        return super().create(vals_list)
    
    def action_open_preview(self):
        """Ouvre l'aperçu du document dans un modal sur la même page."""
        self.ensure_one()
        
        # Pour les images et PDFs, ouvrir dans un wizard avec widget
        if self.mimetype and (self.mimetype.startswith('image/') or self.mimetype == 'application/pdf'):
            return {
                'type': 'ir.actions.act_window',
                'name': self.name or 'Aperçu',
                'res_model': 'ir.attachment',
                'res_id': self.id,
                'view_mode': 'form',
                'view_id': self.env.ref('stockex.view_ir_attachment_preview_form').id,
                'target': 'new',
                'context': {'preview_mode': True},
            }
        
        # Pour les URLs, ouvrir dans un nouvel onglet
        if self.type == 'url' and self.url:
            return {
                'type': 'ir.actions.act_url',
                'url': self.url,
                'target': 'new',
            }
        
        # Pour les autres fichiers, télécharger
        return {
            'type': 'ir.actions.act_url',
            'url': f"/web/content/ir.attachment/{self.id}/datas?download=true",
            'target': 'self',
        }
    
    def action_download(self):
        self.ensure_one()
        if self.type == 'url' and self.url:
            return {
                'type': 'ir.actions.act_url',
                'url': self.url,
                'target': 'new',
            }
        return {
            'type': 'ir.actions.act_url',
            'url': f"/web/content/ir.attachment/{self.id}/datas?download=true",
            'target': 'self',
        }

# Fin de la classe IrAttachment

class StockInventoryExtension(models.Model):
    """Extension de StockInventory avec méthodes de valorisation."""
    _inherit = 'stockex.stock.inventory'

    def _get_product_valuation_price(self, product, method=None):
        """Retourne le prix de valorisation unitaire d'un produit.
        
        Args:
            product: recordset product.product
            method: 'standard' ou 'economic' (si None, utilise la config système)
        
        Priorité des sources:
        1) stock.valuation.layer (référence comptable fiable, convertie en devise société)
        2) stock.move.price_unit (dernier mouvement réalisé) si méthode économique
        3) product.standard_price (fallback)
        
        Applique éventuellement la décote selon rotation si activée.
        """
        self.ensure_one()
        
        if not product:
            return 0.0
        
        # Cache désactivé (les recordsets Odoo n'acceptent pas d'attributs dynamiques)
        
        ICP = self.env['ir.config_parameter'].sudo()
        # Utiliser la méthode passée en paramètre, sinon la config système
        rule = method if method else ICP.get_param('stockex.valuation_rule', 'standard')
        company = self.company_id
        company_currency = company.currency_id
        # Utiliser la date d'inventaire pour la conversion, sinon aujourd'hui
        conv_date = self.date or fields.Date.today()
        
        base_price = 0.0
        
        # 1) Source prioritaire: stock.valuation.layer
        try:
            ValuationLayer = self.env['stock.valuation.layer']
            layer = ValuationLayer.search([
                ('product_id', '=', product.id),
                ('company_id', '=', company.id),
            ], order='create_date desc', limit=1)
            if layer:
                # Déterminer le coût unitaire de la couche
                unit_cost = getattr(layer, 'unit_cost', 0.0) or 0.0
                if not unit_cost:
                    qty = getattr(layer, 'quantity', 0.0) or 0.0
                    val = getattr(layer, 'value', 0.0) or 0.0
                    unit_cost = (val / qty) if qty else 0.0
                # Conversion devise si nécessaire
                source_currency = getattr(layer, 'currency_id', company_currency) or company_currency
                if source_currency and source_currency != company_currency:
                    try:
                        base_price = source_currency._convert(unit_cost, company_currency, company, conv_date)
                    except Exception:
                        base_price = unit_cost
                else:
                    base_price = unit_cost
        except Exception:
            # Continuer vers les autres sources
            pass
        
        # 2) Règle économique: dernier stock.move.price_unit si rien obtenu
        if base_price <= 0.0 and rule == 'economic':
            try:
                moves = self.env['stock.move'].search([
                    ('product_id', '=', product.id),
                    ('company_id', '=', company.id),
                    ('state', '=', 'done')
                ], limit=1, order='date desc')
                if moves and hasattr(moves, 'price_unit') and moves.price_unit:
                    base_price = moves.price_unit
            except Exception:
                pass
        
        # 3) Fallback: coût standard du produit
        if base_price <= 0.0:
            std_price = product.standard_price or 0.0
            # En Odoo 19, standard_price peut être un dict JSONB {"company_id": price}
            if isinstance(std_price, dict):
                # Récupérer le prix pour la société courante (clé = str(company_id))
                company_key = str(company.id)
                std_price = std_price.get(company_key, 0.0) or std_price.get('1', 0.0) or 0.0
            base_price = std_price
        
        # Appliquer décote rotation si activée
        apply_depreciation = ICP.get_param('stockex.apply_depreciation', 'False') == 'True'
        if apply_depreciation and base_price > 0:
            depreciation_coef = self._get_depreciation_coefficient(product)
            base_price = base_price * depreciation_coef
        
        # Garde-fou
        if base_price < 0 or base_price is None:
            base_price = 0.0
        
        # (cache désactivé)
        return base_price
    
    def _get_depreciation_coefficient(self, product):
        """Retourne le coefficient de décote selon la rotation du produit.
        
        Args:
            product: recordset product.product
            
        Returns:
            float: Coefficient de décote (1.0 = pas de décote, 0.6 = 40%, 0.0 = 100%)
            
        Catégories:
        - Stock actif: Mouvement dans les N derniers jours → Coefficient 1.0 (0% décote)
        - Rotation lente: Mouvement entre N et M jours → Coefficient 0.6 (40% décote)
        - Stock mort: Aucun mouvement depuis plus de M jours → Coefficient 0.0 (100% décote)
        """
        self.ensure_one()
        
        if not product:
            return 1.0
        
        # Récupérer les paramètres de décote
        ICP = self.env['ir.config_parameter'].sudo()
        active_days = int(ICP.get_param('stockex.depreciation_active_days', '365'))
        slow_days = int(ICP.get_param('stockex.depreciation_slow_days', '1095'))
        slow_rate = float(ICP.get_param('stockex.depreciation_slow_rate', '40.0'))
        dead_rate = float(ICP.get_param('stockex.depreciation_dead_rate', '100.0'))
        
        # Chercher le dernier mouvement du produit (sortie ou entrée)
        StockMove = self.env['stock.move']
        last_move = StockMove.search([
            ('product_id', '=', product.id),
            ('state', '=', 'done'),
        ], limit=1, order='date desc')
        
        if not last_move:
            # Aucun mouvement = stock mort (décote maximale)
            return 1.0 - (dead_rate / 100.0)
        
        # Calculer le nombre de jours depuis le dernier mouvement
        from datetime import datetime
        
        # Convertir last_move.date en date si c'est un datetime
        if isinstance(last_move.date, datetime):
            last_move_date = last_move.date.date()
        else:
            last_move_date = last_move.date
        
        now = datetime.now().date()
        days_since_last_move = (now - last_move_date).days
        
        # Appliquer les règles de décote
        if days_since_last_move <= active_days:
            # Stock actif: pas de décote
            return 1.0
        
        elif days_since_last_move <= slow_days:
            # Rotation lente: décote partielle
            return 1.0 - (slow_rate / 100.0)
        
        else:
            # Stock mort: décote maximale
            return 1.0 - (dead_rate / 100.0)
    
    @api.depends('line_ids.product_qty','line_ids.theoretical_qty','line_ids.difference','line_ids.product_id')
    def _compute_totals(self):
        """Calcule les totaux (quantités et valeur) de l'inventaire.
        
        Pour l'écart de quantité, on utilise la somme des écarts des lignes
        (qui prend en compte la règle : si stock initial et qte_theo=0 alors écart=0)
        
        ⚠️ VALORISATION: Utilise la méthode de valorisation Stockex (_get_product_valuation_price)
        qui calcule le prix depuis product.standard_price (source unique de vérité)
        """
        for inv in self:
            qty_real = sum(inv.line_ids.mapped('product_qty'))
            qty_theo = sum(inv.line_ids.mapped('theoretical_qty'))
            # Utiliser la somme des écarts calculés (qui respecte la règle stock initial)
            qty_diff = sum(inv.line_ids.mapped('difference'))
            total_val_real = 0.0
            total_val_theo = 0.0
            
            for line in inv.line_ids:
                # Utiliser la méthode de valorisation Stockex (source unique: product.standard_price)
                price = inv._get_product_valuation_price(line.product_id)
                
                total_val_real += (line.product_qty or 0.0) * price
                total_val_theo += (line.theoretical_qty or 0.0) * price
            
            inv.total_quantity_real = qty_real
            inv.total_quantity_theoretical = qty_theo
            inv.total_quantity_difference = qty_diff
            inv.total_value_real = total_val_real
            inv.total_value_theoretical = total_val_theo
    
    @api.depends('total_value_real', 'total_value_theoretical')
    def _compute_value_difference(self):
        """Calcule l'écart de valeur entre valeur réelle et théorique inventoriée.
        
        total_value_difference = total_value_real - total_value_theoretical
        
        Cet écart reflète la différence de valeur sur les produits inventoriés uniquement.
        Écart positif = surplus, Écart négatif = manquant
        """
        for inv in self:
            inv.total_value_difference = inv.total_value_real - inv.total_value_theoretical
    
    @api.depends('location_id', 'warehouse_id', 'company_id')
    def _compute_odoo_stock_value(self):
        """Calcule la valeur totale réelle du stock dans Odoo.
        
        Cette valeur représente la somme de tous les stock.quants
        dans l'emplacement/entrepôt de l'inventaire.
        Cela permet de comparer la valeur inventoriée avec la valeur réelle du stock Odoo.
        
        IMPORTANT: Utilise la règle de valorisation Stockex configurée:
        - Règle 1 (standard): product.standard_price
        - Règle 2 (economic): dernier mouvement de stock (coût économique réel)
        
        Configuration: Inventaire > Configuration > Paramètres > Règle de valorisation
        """
        StockQuant = self.env['stock.quant']
        
        for inv in self:
            odoo_value = 0.0
            
            # Déterminer les emplacements à inclure
            location_ids = []
            if inv.location_id:
                # Emplacement spécifique + ses enfants
                location_ids.append(inv.location_id.id)
                location_ids.extend(inv.location_id.child_ids.ids)
            elif inv.warehouse_id:
                # Tous les emplacements de l'entrepôt (stock internal)
                stock_location = inv.warehouse_id.lot_stock_id
                if stock_location:
                    location_ids.append(stock_location.id)
                    location_ids.extend(stock_location.child_ids.ids)
            
            if not location_ids:
                inv.odoo_stock_value = 0.0
                continue
            
            # Récupérer tous les quants pour ces emplacements
            domain = [
                ('location_id', 'in', location_ids),
                ('company_id', '=', inv.company_id.id),
            ]
            
            quants = StockQuant.search(domain)
            
            # Calculer la valeur totale selon la règle de valorisation Stockex
            for quant in quants:
                available_qty = quant.quantity - quant.reserved_quantity
                
                if available_qty <= 0:
                    continue
                
                # Utiliser la méthode de valorisation Stockex
                product_price = inv._get_product_valuation_price(quant.product_id)
                odoo_value += available_qty * product_price
            
            inv.odoo_stock_value = odoo_value
            
            _logger.info(
                f"📊 Inventaire {inv.name}: Valeur Stock Odoo = {odoo_value:.2f} | "
                f"Valeur Inventoriée = {inv.total_value_real:.2f} | "
                f"Écart = {inv.total_value_real - odoo_value:.2f}"
            )

    def unlink(self):
        """Empêcher la suppression des inventaires validés."""
        for inventory in self:
            if inventory.state == 'done':
                raise UserError(
                    f"🚫 Impossible de supprimer l'inventaire '{inventory.name}'.\n\n"
                    f"L'inventaire a été validé le {inventory.validation_date.strftime('%d/%m/%Y à %H:%M') if inventory.validation_date else 'N/A'}.\n"
                    f"Les stocks Odoo ont déjà été mis à jour.\n\n"
                    f"❌ Un inventaire validé ne peut jamais être supprimé pour des raisons de traçabilité et d'audit.\n\n"
                    f"💡 Si vous devez corriger des erreurs :\n"
                    f"   - Créez un nouvel inventaire correctif\n"
                    f"   - Documentez les changements dans les notes"
                )
            elif inventory.state == 'approved':
                raise UserError(
                    f"🚫 Impossible de supprimer l'inventaire '{inventory.name}'.\n\n"
                    f"L'inventaire a été approuvé le {inventory.approval_date.strftime('%d/%m/%Y à %H:%M') if inventory.approval_date else 'N/A'}.\n"
                    f"Vous devez d'abord le rejeter avant de pouvoir le supprimer."
                )
        return super(StockInventory, self).unlink()
    
    def _notify_telegram(self, text):
        ICP = self.env['ir.config_parameter'].sudo()
        enabled = ICP.get_param('stockex.notify_by_telegram') or ''
        if str(enabled).lower() in ('false', '0', '', 'none'):
            return
        token = ICP.get_param('stockex.telegram_bot_token') or ''
        chats = ICP.get_param('stockex.telegram_chat_ids') or ''
        if not token or not chats:
            return
        import requests
        for chat_id in [c.strip() for c in chats.split(',') if c.strip()]:
            try:
                requests.post(
                    f'https://api.telegram.org/bot{token}/sendMessage',
                    data={'chat_id': chat_id, 'text': text}
                )
            except Exception as e:
                _logger.error(f"Telegram notification error: {e}")

    def action_start(self):
        """Démarre l'inventaire."""
        if not self.line_ids:
            raise UserError("Vous devez ajouter au moins une ligne avant de démarrer l'inventaire.")
        return self.write({'state': 'in_progress'})
    
    def action_request_approval(self):
        """Demande l'approbation de l'inventaire."""
        if not self.line_ids:
            raise UserError("Impossible de demander l'approbation d'un inventaire sans lignes.")
        
        # Créer une activité pour le manager
        manager = self.user_id.parent_id or self.env.ref('base.user_admin')
        self.activity_schedule(
            'mail.mail_activity_data_todo',
            user_id=manager.id,
            summary=f"Approbation d'inventaire : {self.name}",
            note=f"Merci d'approuver l'inventaire {self.name} du {self.date}"
        )
        
        return self.write({'state': 'pending_approval'})
    
    def action_approve(self):
        """Approuve l'inventaire et synchronise automatiquement avec Odoo natif."""
        self.write({
            'state': 'approved',
            'approver_id': self.env.user.id,
            'approval_date': fields.Datetime.now()
        })
        
        # Marquer l'activité comme terminée
        activity = self.activity_ids.filtered(lambda a: a.user_id == self.env.user)
        if activity:
            activity.action_done()
        
        # ✅ Synchronisation automatique vers Odoo natif
        try:
            _logger.info(f"🔄 Déclenchement synchronisation automatique pour {self.name}")
            self.sync_to_native_inventory()
        except Exception as e:
            # Ne pas bloquer l'approbation si la synchro échoue
            _logger.warning(f"⚠️ Échec synchronisation automatique pour {self.name}: {e}")
            self.message_post(
                body=Markup(f"""
                <p style="color: #856404;">⚠️ <strong>Avertissement</strong></p>
                <p>L'inventaire a été approuvé, mais la synchronisation automatique a échoué.</p>
                <p><em>Erreur : {str(e)}</em></p>
                <p>Vous pouvez relancer la synchronisation manuellement via le bouton '🔄 Synchro vers Odoo Natif'.</p>
                """),
                message_type='notification',
                subtype_xmlid='mail.mt_note'
            )
        
        return True
    
    def action_reject(self):
        """Rejette l'inventaire et le remet en brouillon."""
        # Marquer l'activité comme annulée
        activity = self.activity_ids.filtered(lambda a: a.user_id == self.env.user)
        if activity:
            activity.unlink()
        
        return self.write({'state': 'draft'})
    
    def action_validate(self):
        """Valide l'inventaire et met à jour les stocks Odoo."""
        import threading
        
        for inventory in self:
            if not inventory.line_ids:
                raise UserError("Impossible de valider un inventaire sans lignes.")
            
            total_lines = len(inventory.line_ids)
            
            # Pour les gros inventaires (> 500 lignes), utiliser traitement asynchrone
            if total_lines > 500:
                _logger.info(f"🚀 Gros inventaire ({total_lines} lignes) → Traitement en thread séparé")
                
                # Marquer comme validé immédiatement
                inventory.write({
                    'state': 'done',
                    'validator_id': self.env.user.id,
                    'validation_date': fields.Datetime.now()
                })
                self.env.cr.commit()
                inventory._notify_telegram(f"✅ Inventaire {inventory.name} validé ({total_lines} lignes). Mise à jour en arrière-plan.")
                
                # Lancer le traitement dans un thread séparé
                thread = threading.Thread(
                    target=inventory._update_odoo_stock_async,
                    args=(inventory.id, self.env.cr.dbname)
                )
                thread.daemon = True
                thread.start()
                
                # Message utilisateur
                inventory.message_post(
                    body=Markup(f"⏳ Mise à jour de {total_lines} lignes de stock en cours en arrière-plan..."),
                    message_type='notification'
                )
                
                return {
                    'type': 'ir.actions.client',
                    'tag': 'display_notification',
                    'params': {
                        'title': '✅ Inventaire validé',
                        'message': f'La mise à jour de {total_lines} lignes se fait en arrière-plan. Vous pouvez continuer à travailler.',
                        'type': 'success',
                        'sticky': True,
                    }
                }
            else:
                # Pour petits inventaires (≤ 500 lignes), traitement immédiat
                if inventory.state != 'approved':
                    raise UserError("Cet inventaire doit être approuvé avant validation.")
                moves = inventory._update_odoo_stock()
                inventory.write({
                    'state': 'done',
                    'validator_id': self.env.user.id,
                    'validation_date': fields.Datetime.now()
                })
                if moves:
                    # Les mouvements de stock générés sont déjà liés via stockex_inventory_id
                    # Mais il faut aussi lier les écritures comptables si elles existent
                    account_moves = moves.mapped('account_move_id').filtered(lambda m: m)
                    if account_moves:
                        inventory.account_move_ids = [(4, move.id) for move in account_moves]
                inventory._notify_telegram(f"✅ Inventaire {inventory.name} validé ({total_lines} lignes). Stocks mis à jour.")
                return True
    
    def _update_odoo_stock(self):
        """Met à jour les stocks Odoo avec l'API native (stock.move + _update_available_quantity)."""
        self.ensure_one()
        
        StockMove = self.env['stock.move']
        StockQuant = self.env['stock.quant']
        moves_created = self.env['stock.move']
        adjusted_count = 0
        errors = []
        skipped_no_data = 0
        skipped_bad_location = 0
        skipped_no_difference = 0
        
        # Emplacements virtuels pour ajustements
        inventory_loc = self.env.ref('stock.location_inventory', raise_if_not_found=False)
        if not inventory_loc:
            # Créer l'emplacement d'inventaire virtuel s'il n'existe pas
            inventory_loc = self.env['stock.location'].create({
                'name': 'Inventory adjustment',
                'usage': 'inventory',
                'company_id': self.company_id.id,
            })
        
        total_lines = len(self.line_ids)
        batch_size = 50
        
        _logger.info(f"🚀 [NATIF] Début mise à jour stocks pour {self.name} - {total_lines} lignes")
        
        for batch_num, i in enumerate(range(0, total_lines, batch_size), 1):
            batch_lines = self.line_ids[i:i + batch_size]
            batch_start = i + 1
            batch_end = min(i + batch_size, total_lines)
            
            _logger.info(f"📦 Lot {batch_num}: lignes {batch_start}-{batch_end}")
            
            for line in batch_lines:
                try:
                    if not line.product_id or not line.location_id:
                        skipped_no_data += 1
                        continue
                    
                    # Vérifier que l'emplacement est de type interne
                    if line.location_id.usage != 'internal':
                        skipped_bad_location += 1
                        errors.append(f"Emplacement '{line.location_id.name}' non interne")
                        continue
                    
                    # Calculer la différence à ajuster
                    difference = line.product_qty - line.theoretical_qty
                    
                    if difference == 0:
                        skipped_no_difference += 1
                        continue
                    
                    # Créer un stock.move pour l'ajustement (API native)
                    move_vals = {
                        'name': f'Inventaire {self.name} - {line.product_id.display_name}',
                        'product_id': line.product_id.id,
                        'product_uom': line.product_id.uom_id.id,
                        'product_uom_qty': abs(difference),
                        'company_id': self.company_id.id,
                        'date': self.date or fields.Datetime.now(),
                        'origin': self.name,
                        'reference': f'Ajustement inventaire {self.name}',
                        'stockex_inventory_id': self.id,
                        'stockex_inventory_line_id': line.id,
                    }
                    
                    # Si différence > 0 : entrée (depuis inventory vers location)
                    # Si différence < 0 : sortie (depuis location vers inventory)
                    if difference > 0:
                        move_vals.update({
                            'location_id': inventory_loc.id,
                            'location_dest_id': line.location_id.id,
                        })
                    else:
                        move_vals.update({
                            'location_id': line.location_id.id,
                            'location_dest_id': inventory_loc.id,
                        })
                    
                    # Créer et valider le mouvement
                    move = StockMove.create(move_vals)
                    move._action_confirm()
                    move._action_assign()
                    move._action_done()
                    
                    moves_created |= move
                    adjusted_count += 1
                    
                    _logger.debug(f"✅ Ajustement {difference:+.2f} pour {line.product_id.default_code} @ {line.location_id.name}")
                    
                except Exception as e:
                    error_msg = f"{line.product_id.default_code or line.product_id.name} @ {line.location_id.name}: {str(e)}"
                    errors.append(error_msg)
                    _logger.error(f"❌ Erreur: {error_msg}")
            
            # Commit après chaque lot
            self.env.cr.commit()
            progress_pct = (batch_end / total_lines) * 100
            _logger.info(f"✅ Lot {batch_num} terminé: {adjusted_count} mouvements ({progress_pct:.1f}%)")
        
        # Statistiques détaillées
        stats = f"""
📊 Statistiques (API Native):
- Total lignes: {total_lines}
- ✅ Mouvements créés: {adjusted_count}
- ⏭️ Ignorées (pas de différence): {skipped_no_difference}
- ⚠️ Ignorées (emplacement non interne): {skipped_bad_location}
- ⚠️ Ignorées (sans données): {skipped_no_data}
- ❌ Erreurs: {len(errors)}
"""
        
        # Message de confirmation
        message = f"✅ Stocks mis à jour via API native : {adjusted_count} mouvements créés sur {total_lines} lignes"
        message += stats
        
        if errors:
            message += f"\n\n⚠️ Détails des erreurs ({len(errors)}):\n" + "\n".join(errors[:20])
        
        _logger.info(message)
        
        # Poster un message dans le chatter
        try:
            self.message_post(
                body=Markup(message),
                message_type='notification',
                subtype_xmlid='mail.mt_note'
            )
        except Exception as msg_error:
            _logger.warning(f"⚠️ Impossible de poster le message dans le chatter: {msg_error}")
        
        return moves_created
    
    @staticmethod
    def _update_odoo_stock_async(inventory_id, dbname):
        """Version asynchrone de _update_odoo_stock pour gros inventaires (utilise un nouveau curseur)."""
        from odoo import api, SUPERUSER_ID
        from odoo.modules.registry import Registry
        
        registry = None
        try:
            # Créer un nouveau curseur et registry
            registry = Registry(dbname)
            with registry.cursor() as new_cr:
                env = api.Environment(new_cr, SUPERUSER_ID, {})
                inventory = env['stockex.stock.inventory'].browse(inventory_id)
                
                _logger.info(f"🚀 Début traitement asynchrone pour inventaire {inventory.name}")
                inventory._update_odoo_stock()
                new_cr.commit()
                _logger.info(f"✅ Traitement asynchrone terminé pour inventaire {inventory.name}")
                
        except Exception as e:
            _logger.error(f"❌ Erreur traitement asynchrone: {e}", exc_info=True)
            # Essayer de poster l'erreur dans le chatter
            if registry:
                try:
                    with registry.cursor() as err_cr:
                        err_env = api.Environment(err_cr, SUPERUSER_ID, {})
                        inventory = err_env['stockex.stock.inventory'].browse(inventory_id)
                        inventory.message_post(
                            body=Markup(f"❌ Erreur lors de la mise à jour des stocks: {str(e)}"),
                            message_type='notification',
                            subtype_xmlid='mail.mt_note'
                        )
                        err_cr.commit()
                except Exception as msg_err:
                    _logger.error(f"❌ Impossible de poster l'erreur: {msg_err}")
    
    def action_draft(self):
        """Remet l'inventaire en brouillon."""
        return self.write({'state': 'draft'})
    
    def action_open_cancel_wizard(self):
        """Ouvre le wizard de confirmation d'annulation pour un ou plusieurs inventaires."""
        # Si plusieurs inventaires sélectionnés, traiter le premier (ou afficher un message d'erreur)
        if len(self) > 1:
            raise UserError(
                "Vous ne pouvez annuler qu'un seul inventaire à la fois.\n"
                "Veuillez sélectionner un seul inventaire et réessayer."
            )
        
        # Vérifier que l'inventaire peut être annulé
        if self.state != 'done':
            raise UserError(
                f"L'inventaire '{self.name}' ne peut pas être annulé.\n"
                f"Seuls les inventaires validés (état: Validé) peuvent être annulés.\n"
                f"État actuel: {dict(self._fields['state'].selection).get(self.state)}"
            )
        
        # Créer le wizard avec l'inventaire sélectionné
        wizard = self.env['stockex.cancel.inventory.wizard'].create({
            'inventory_id': self.id,
        })
        
        return {
            'name': 'Confirmer l\'Annulation',
            'type': 'ir.actions.act_window',
            'res_model': 'stockex.cancel.inventory.wizard',
            'res_id': wizard.id,
            'view_mode': 'form',
            'target': 'new',
        }
    
    def action_cancel(self):
        """Annule l'inventaire validé en inversant les ajustements de stock et en supprimant les écritures comptables."""
        for inventory in self:
            if inventory.state != 'done':
                raise UserError("Seuls les inventaires validés peuvent être annulés.")
            
            _logger.info(f"🔄 Début de l'annulation de l'inventaire {inventory.name}")
            # Tentative d'annulation native via les mouvements liés
            moves_to_cancel = inventory.stock_move_ids.filtered(lambda m: m.state == 'done')
            if moves_to_cancel:
                _logger.info(f"📦 Annulation de {len(moves_to_cancel)} mouvements liés")
                try:
                    moves_to_cancel._action_cancel()
                    inventory.write({'state': 'cancel', 'validator_id': False, 'validation_date': False})
                    inventory.message_post(
                        body=Markup(f"❌ Inventaire annulé - {len(moves_to_cancel)} mouvements inversés"),
                        message_type='notification'
                    )
                    continue
                except Exception as cancel_err:
                    _logger.warning(f"⚠️ Annulation via mouvements a échoué, fallback quants: {cancel_err}")
            
            quants_adjusted = 0
            account_moves_count = 0
            errors = []
            
            # 1. INVERSER LES AJUSTEMENTS DE STOCK DANS LES QUANTS
            # Au lieu de chercher des stock.move, on inverse directement les quants
            StockQuant = self.env['stock.quant']
            
            for line in inventory.line_ids:
                try:
                    if not line.product_id or not line.location_id:
                        continue
                    
                    # Vérifier que l'emplacement est de type interne
                    if line.location_id.usage != 'internal':
                        continue
                    
                    # Trouver le quant
                    quant = StockQuant.search([
                        ('product_id', '=', line.product_id.id),
                        ('location_id', '=', line.location_id.id),
                        ('company_id', '=', inventory.company_id.id),
                    ], limit=1)
                    
                    if quant:
                        # Calculer la quantité à restaurer (quantité avant inventaire)
                        # product_qty = quantité après inventaire (dans l'inventaire)
                        # theoretical_qty = quantité avant inventaire (théorique)
                        qty_to_restore = line.theoretical_qty
                        
                        # Restaurer la quantité d'avant inventaire
                        quant.inventory_quantity = qty_to_restore
                        quant.inventory_quantity_set = True
                        quant.action_apply_inventory()
                        
                        quants_adjusted += 1
                        _logger.info(
                            f"✅ Quant inversé: {line.product_id.default_code} @ {line.location_id.name}: "
                            f"{line.product_qty} → {qty_to_restore}"
                        )
                    else:
                        # Si le quant n'existe plus et que la quantité théorique était > 0, le recréer
                        if line.theoretical_qty > 0:
                            new_quant = StockQuant.create({
                                'product_id': line.product_id.id,
                                'location_id': line.location_id.id,
                                'company_id': inventory.company_id.id,
                                'inventory_quantity': line.theoretical_qty,
                                'inventory_quantity_set': True,
                            })
                            new_quant.action_apply_inventory()
                            quants_adjusted += 1
                            _logger.info(
                                f"✅ Quant recréé: {line.product_id.default_code} @ {line.location_id.name}: "
                                f"0 → {line.theoretical_qty}"
                            )
                            
                except Exception as e:
                    error_msg = f"Produit {line.product_id.default_code} @ {line.location_id.name}: {str(e)}"
                    errors.append(error_msg)
                    _logger.error(f"❌ Erreur inversion ligne: {error_msg}")
            
            _logger.info(f"📦 {quants_adjusted} quant(s) inversé(s)")
            
            # 2. SUPPRIMER LES ÉCRITURES COMPTABLES LIÉES
            account_moves = inventory.move_ids if hasattr(inventory, 'move_ids') else self.env['account.move']
            
            # Rechercher aussi par référence si aucune écriture liée
            if not account_moves:
                account_moves = self.env['account.move'].search([
                    ('ref', 'ilike', inventory.name),
                    ('move_type', '=', 'entry'),
                ])
            
            if account_moves:
                _logger.info(f"📒 Trouvé {len(account_moves)} écriture(s) comptable(s) à supprimer")
                account_moves_count = len(account_moves)
                
                for account_move in account_moves:
                    try:
                        move_name = account_move.name
                        move_id = account_move.id
                        
                        # Annuler l'écriture si elle est validée
                        if account_move.state == 'posted':
                            account_move.button_draft()
                            _logger.info(f"✅ Écriture {move_name} remise en brouillon")
                        
                        # Supprimer l'écriture avec force
                        try:
                            account_move.with_context(force_delete=True).unlink()
                            _logger.info(f"🗑️ Écriture {move_name} supprimée")
                        except Exception as unlink_error:
                            # Si unlink échoue, forcer la suppression via SQL
                            _logger.warning(f"⚠️ unlink() a échoué pour {move_name}, utilisation de SQL: {str(unlink_error)}")
                            self.env.cr.execute("DELETE FROM account_move_line WHERE move_id = %s", (move_id,))
                            self.env.cr.execute("DELETE FROM account_move WHERE id = %s", (move_id,))
                            _logger.info(f"🗑️ Écriture {move_name} (ID:{move_id}) supprimée via SQL")
                    except Exception as e:
                        error_msg = f"Écriture {account_move.name}: {str(e)}"
                        errors.append(error_msg)
                        _logger.error(f"❌ Erreur suppression écriture: {error_msg}")
            
            # 3. Annuler l'inventaire
            inventory.write({'state': 'cancel'})
            
            # 4. Message dans le chatter
            message_body = f"""
            <div style="padding: 20px; background: #f8d7da; border-left: 5px solid #dc3545; border-radius: 5px;">
                <h3 style="color: #721c24; margin-top: 0;">❌ Inventaire annulé par {self.env.user.name}</h3>
                <hr style="border-color: #dc3545;"/>
                <p><strong>🔄 Ajustements inversés :</strong></p>
                <ul>
                    <li>📦 {quants_adjusted} quant(s) restauré(s) à leur état avant inventaire</li>
                    <li>📒 {account_moves_count} écriture(s) comptable(s) supprimée(s)</li>
                </ul>
            """
            
            if errors:
                message_body += f"""
                <hr style="border-color: #ffc107;"/>
                <p><strong>⚠️ Erreurs ({len(errors)}) :</strong></p>
                <ul style="color: #856404; font-size: 12px;">
                """
                for error in errors[:10]:
                    message_body += f"<li>{error}</li>"
                if len(errors) > 10:
                    message_body += f"<li>... et {len(errors) - 10} autre(s) erreur(s)</li>"
                message_body += "</ul>"
            
            message_body += "</div>"
            
            inventory.message_post(
                body=Markup(message_body),
                message_type='comment',
                subtype_xmlid='mail.mt_note'
            )
            
            _logger.info(
                f"✅ Inventaire {inventory.name} annulé avec succès: "
                f"{quants_adjusted} quants inversés, {account_moves_count} écritures supprimées"
            )
        
        return True
    
    def name_get(self):
        """Affichage personnalisé du nom."""
        result = []
        for record in self:
            name = f"{record.name} - {record.date}"
            if record.location_id:
                name += f" ({record.location_id.name})"
            result.append((record.id, name))
        return result
    
    @api.model_create_multi
    def create(self, vals_list):
        """Génère automatiquement la référence si elle n'est pas fournie."""
        for vals in vals_list:
            if vals.get('name', 'Nouveau') == 'Nouveau' or not vals.get('name'):
                vals['name'] = self.env['ir.sequence'].next_by_code('stockex.stock.inventory') or 'Nouveau'
        return super().create(vals_list)
    
    def sync_to_native_inventory(self):
        """Synchronise l'inventaire StockEx vers les quants natifs Odoo.
        
        Met à jour directement les stock.quant avec les quantités comptées dans StockEx,
        en respectant les emplacements (Odoo 19+).
        """
        self.ensure_one()
        
        if not self.line_ids:
            raise UserError("⚠️ Impossible de synchroniser un inventaire sans lignes.")
        
        StockQuant = self.env['stock.quant']
        synced_count = 0
        errors = []
        
        _logger.info(f"🔄 Début synchronisation StockEx → Quants Odoo pour {self.name}")
        
        for line in self.line_ids:
            if not line.product_id or not line.location_id:
                continue
            
            try:
                # Chercher le quant correspondant
                quant = StockQuant.search([
                    ('product_id', '=', line.product_id.id),
                    ('location_id', '=', line.location_id.id),
                    ('company_id', '=', self.company_id.id),
                ], limit=1)
                
                if quant:
                    # Mettre à jour la quantité inventoriee
                    quant.inventory_quantity = line.product_qty
                    quant.inventory_quantity_set = True
                    quant.inventory_diff_quantity = line.product_qty - quant.quantity
                    
                    _logger.debug(
                        f"✅ Quant mis à jour: {line.product_id.default_code} @ {line.location_id.name}: "
                        f"{quant.quantity} → {line.product_qty}"
                    )
                else:
                    # Créer un nouveau quant si nécessaire et quantité > 0
                    if line.product_qty > 0:
                        quant = StockQuant.create({
                            'product_id': line.product_id.id,
                            'location_id': line.location_id.id,
                            'company_id': self.company_id.id,
                            'inventory_quantity': line.product_qty,
                            'inventory_quantity_set': True,
                            'inventory_diff_quantity': line.product_qty,
                        })
                        _logger.debug(
                            f"✅ Quant créé: {line.product_id.default_code} @ {line.location_id.name}: "
                            f"0 → {line.product_qty}"
                        )
                    else:
                        continue
                
                synced_count += 1
                
            except Exception as e:
                error_msg = f"{line.product_id.default_code} @ {line.location_id.name}: {str(e)}"
                errors.append(error_msg)
                _logger.error(f"❌ Erreur synchronisation: {error_msg}")
        
        # Marquer comme synchronisé
        self.write({
            'sync_to_native': True,
            'sync_date': fields.Datetime.now(),
        })
        
        # Message de confirmation
        message = f"""
        <div style="padding: 15px; background: #d1ecf1; border-left: 4px solid #0c5460; border-radius: 5px;">
            <h4 style="color: #0c5460; margin-top: 0;">🔄 Synchronisation vers Stock Natif Odoo</h4>
            <hr style="border-color: #bee5eb;"/>
            <ul style="margin: 10px 0;">
                <li><strong>✅ Quants synchronisés :</strong> {synced_count} ligne(s)</li>
                <li><strong>📍 Emplacement :</strong> {self.location_id.complete_name if self.location_id else 'Tous'}</li>
                <li><strong>🕒 Date :</strong> {self.sync_date.strftime('%d/%m/%Y à %H:%M')}</li>
            </ul>
            <p style="margin: 10px 0; font-style: italic; color: #0c5460;">
                Les quantités comptées dans StockEx sont maintenant dans les quants Odoo natifs.
                Vous pouvez appliquer l'ajustement depuis <strong>Inventaire > Opérations > Ajustements d'inventaire</strong>.
            </p>
        """
        
        if errors:
            message += f"""
            <hr style="border-color: #ffc107;"/>
            <p style="color: #856404;"><strong>⚠️ Erreurs ({len(errors)}) :</strong></p>
            <ul style="font-size: 11px; color: #856404;">
            """
            for error in errors[:10]:
                message += f"<li>{error}</li>"
            if len(errors) > 10:
                message += f"<li>... et {len(errors) - 10} autre(s) erreur(s)</li>"
            message += "</ul>"
        
        message += "</div>"
        
        self.message_post(
            body=Markup(message),
            message_type='notification',
            subtype_xmlid='mail.mt_note'
        )
        
        _logger.info(f"✅ Synchronisation terminée: {synced_count} quants mis à jour")
        
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': '✅ Synchronisation réussie',
                'message': f'{synced_count} quantité(s) synchronisée(s) vers les quants Odoo natifs',
                'type': 'success',
                'sticky': False,
            }
        }
    
    def action_export_excel(self):
        """Exporter l'inventaire en Excel."""
        self.ensure_one()
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            import base64
            from io import BytesIO
        except ImportError:
            raise UserError("La bibliothèque openpyxl est requise pour l'export Excel.")
        
        # Créer le workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventaire"
        
        # Styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # En-tête du document
        ws['A1'] = f"INVENTAIRE DE STOCK - {self.name}"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:J1')
        
        ws['A2'] = f"Date: {self.date}"
        ws['A3'] = f"Responsable: {self.user_id.name}"
        ws['A4'] = f"Société: {self.company_id.name}"
        ws['A5'] = f"État: {dict(self._fields['state'].selection).get(self.state)}"
        
        # En-têtes colonnes (ligne 7)
        headers = [
            'Produit',
            'Référence',
            'Catégorie',
            'Emplacement',
            'Qté Théorique',
            'Qté Réelle',
            'Écart',
            'Prix Standard (FCFA)',
            'Valeur Écart (FCFA)'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Données
        row_num = 8
        total_theoretical = 0
        total_real = 0
        total_difference_value = 0
        
        for line in self.line_ids:
            ws.cell(row=row_num, column=1, value=line.product_id.name or '')
            ws.cell(row=row_num, column=2, value=line.product_id.default_code or '')
            ws.cell(row=row_num, column=3, value=line.product_id.categ_id.name or '')
            ws.cell(row=row_num, column=4, value=line.location_id.display_name or '')
            ws.cell(row=row_num, column=5, value=line.theoretical_qty)
            ws.cell(row=row_num, column=6, value=line.product_qty)
            ws.cell(row=row_num, column=7, value=line.difference)
            ws.cell(row=row_num, column=8, value=line.standard_price)
            
            diff_value = line.difference * line.standard_price
            ws.cell(row=row_num, column=9, value=diff_value)
            
            # Bordures
            for col in range(1, 10):
                ws.cell(row=row_num, column=col).border = border
            
            # Colorer les écarts
            diff_cell = ws.cell(row=row_num, column=7)
            if line.difference > 0:
                diff_cell.font = Font(color="008000", bold=True)  # Vert
            elif line.difference < 0:
                diff_cell.font = Font(color="FF0000", bold=True)  # Rouge
            
            total_theoretical += line.theoretical_qty
            total_real += line.product_qty
            total_difference_value += diff_value
            row_num += 1
        
        # Ligne totaux
        row_num += 1
        ws.cell(row=row_num, column=1, value="TOTAUX").font = Font(bold=True)
        ws.cell(row=row_num, column=5, value=total_theoretical).font = Font(bold=True)
        ws.cell(row=row_num, column=6, value=total_real).font = Font(bold=True)
        ws.cell(row=row_num, column=7, value=total_real - total_theoretical).font = Font(bold=True)
        ws.cell(row=row_num, column=9, value=total_difference_value).font = Font(bold=True)
        
        for col in range(1, 10):
            ws.cell(row=row_num, column=col).fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        # Ajuster largeur colonnes
        column_widths = [30, 15, 20, 35, 15, 15, 12, 15, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Sauvegarder en mémoire
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Créer l'attachement
        attachment = self.env['ir.attachment'].create({
            'name': f'Inventaire_{self.name.replace("/", "_")}.xlsx',
            'type': 'binary',
            'datas': base64.b64encode(output.read()),
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }
    
    def action_print_pdf(self):
        """Imprimer l'inventaire en PDF."""
        self.ensure_one()
        return self.env.ref('stockex.action_report_inventory').report_action(self)
    
    def action_print_variance_list(self):
        """Imprimer la liste des écarts (produits avec difference != 0)."""
        self.ensure_one()
        
        # Filtrer les lignes avec des écarts
        variance_lines = self.line_ids.filtered(lambda l: (l.difference or 0.0) != 0.0)
        
        if not variance_lines:
            raise UserError("Aucun écart trouvé dans cet inventaire.")
        
        # Créer un rapport personnalisé avec uniquement les écarts
        return {
            'type': 'ir.actions.report',
            'report_type': 'qweb-pdf',
            'report_name': 'stockex.report_inventory_variance_list',
            'report_file': 'stockex.report_inventory_variance_list',
            'context': {
                'active_model': 'stockex.stock.inventory',
                'active_ids': [self.id],
                'active_id': self.id,
                'variance_only': True,  # Indicateur pour le template
            },
            'data': {
                'model': 'stockex.stock.inventory',
                'ids': [self.id],
                'id': self.id,
                'docs': self,
                'variance_lines': variance_lines.read(),
            }
        }
    
    def action_refresh_theoretical_qty(self):
        """Recalcule les quantités théoriques ET les prix unitaires depuis le stock Odoo actuel."""
        self.ensure_one()
        _logger.info(f"🔄 Recalcul des quantités théoriques ET prix pour inventaire {self.name}")
        
        if not self.line_ids:
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': 'Aucune ligne',
                    'message': 'Cet inventaire ne contient aucune ligne.',
                    'type': 'warning',
                    'sticky': False,
                }
            }
        
        # Forcer le recalcul pour chaque ligne
        updated_count = 0
        price_updated = 0
        for line in self.line_ids:
            if not line.product_id or not line.location_id:
                continue
            
            # Récupérer la quantité depuis stock.quant
            quants = self.env['stock.quant'].search([
                ('product_id', '=', line.product_id.id),
                ('location_id', '=', line.location_id.id),
            ])
            
            qty_available = sum(quants.mapped('quantity')) - sum(quants.mapped('reserved_quantity'))
            
            # Calculer la différence
            difference = line.product_qty - qty_available
            
            # Forcer l'écriture directe (bypass du compute)
            self.env.cr.execute("""
                UPDATE stockex_stock_inventory_line 
                SET theoretical_qty = %s, difference = %s
                WHERE id = %s
            """, (qty_available, difference, line.id))
            
            if qty_available > 0:
                updated_count += 1
            
            _logger.info(
                f"📦 Ligne {line.id}: {line.product_id.name} → "
                f"Théo: {qty_available}, Réel: {line.product_qty}, Écart: {difference}"
            )
        
        # Invalider le cache pour forcer le rechargement
        self.line_ids.invalidate_recordset(['theoretical_qty', 'difference', 'difference_display', 'standard_price'])
        
        # Compter les résultats
        lines_with_qty = len([l for l in self.line_ids if l.theoretical_qty > 0])
        
        message = f"✅ Quantités théoriques recalculées\n"
        message += f"📊 {lines_with_qty} ligne(s) avec stock > 0\n"
        message += f"📦 {len(self.line_ids) - lines_with_qty} ligne(s) avec stock = 0"
        
        _logger.info(f"✅ Recalcul terminé: {lines_with_qty}/{len(self.line_ids)} lignes avec stock")
        
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': 'Recalcul terminé',
                'message': message,
                'type': 'success',
                'sticky': False,
            }
        }
    
    @api.model
    def _send_inventory_reminders(self):
        """Envoie des rappels pour les inventaires en cours depuis plus de 7 jours."""
        from datetime import datetime, timedelta
        
        cutoff_date = datetime.now() - timedelta(days=7)
        old_inventories = self.search([
            ('state', 'in', ['in_progress', 'pending_approval']),
            ('create_date', '<=', cutoff_date)
        ])
        
        for inventory in old_inventories:
            # Créer une activité de rappel
            inventory.activity_schedule(
                'mail.mail_activity_data_todo',
                user_id=inventory.user_id.id,
                summary=f"Rappel: Inventaire en attente - {inventory.name}",
                note=f"L'inventaire {inventory.name} est en cours depuis plus de 7 jours. Merci de le finaliser.",
                date_deadline=fields.Date.today()
            )
            
            _logger.info(f"Rappel envoyé pour inventaire {inventory.name}")
        
        return True


class StockInventoryLine(models.Model):
    _name = 'stockex.stock.inventory.line'
    _description = 'Ligne d\'inventaire'
    _order = 'product_id, id'
    _rec_name = 'product_id'
    
    # Champ requis pour le widget badge
    color = fields.Integer(string='Couleur', default=0)
    
    inventory_id = fields.Many2one(
        comodel_name='stockex.stock.inventory',
        string='Inventaire',
        required=True,
        ondelete='cascade',
        index=True
    )
    product_id = fields.Many2one(
        comodel_name='product.product',
        string='Produit',
        required=True,
        index=True
    )
    product_barcode = fields.Char(
        string='Code-barres',
        related='product_id.barcode',
        readonly=True,
        help='Code-barres du produit pour scan mobile'
    )
    product_categ_id = fields.Many2one(
        comodel_name='product.category',
        string='Catégorie',
        related='product_id.categ_id',
        readonly=True,
        store=True
    )
    scanned_barcode = fields.Char(
        string='Code-barres scanné',
        help='Code-barres scanné pour recherche rapide de produit'
    )
    theoretical_qty = fields.Float(
        string='Quantité théorique',
        digits='Product Unit of Measure',
        readonly=True,
        compute='_compute_theoretical_qty',
        store=True,
        compute_sudo=True,
        inverse='_inverse_theoretical_qty',  # Permettre de forcer la valeur
    )
    product_qty = fields.Float(
        string='Quantité réelle',
        digits='Product Unit of Measure',
        default=0.0
    )
    difference = fields.Float(
        string='Différence',
        compute='_compute_difference',
        store=True,
        digits='Product Unit of Measure',
        readonly=True
    )
    difference_display = fields.Html(
        string='Écart',
        compute='_compute_difference_display',
        readonly=True
    )
    location_id = fields.Many2one(
        comodel_name='stock.location',
        string='Emplacement',
        index=True
    )
    standard_price = fields.Float(
        string='Prix unitaire',
        compute='_compute_standard_price',
        digits='Product Price',
        store=False,
        help='Prix unitaire du produit (calculé depuis product.standard_price)'
    )
    product_uom_id = fields.Many2one(
        comodel_name='uom.uom',
        string='Unité de Mesure',
        related='product_id.uom_id',
        readonly=True
    )
    currency_id = fields.Many2one(
        comodel_name='res.currency',
        string='Devise',
        related='inventory_id.company_id.currency_id',
        store=True,
        readonly=True,
    )
    line_value = fields.Monetary(
        string='Valeur de ligne',
        currency_field='currency_id',
        compute='_compute_line_value',
        store=True,
        help="Valeur inventoriée = Quantité réelle × Prix unitaire",
    )
    difference_value = fields.Monetary(
        string='Valeur de l’écart',
        currency_field='currency_id',
        compute='_compute_difference_value',
        store=True,
        help="Valeur de l’écart = Écart × Prix unitaire",
    )
    inventory_date = fields.Date(
        string='Date inventaire',
        related='inventory_id.date',
        store=True,
        readonly=True,
    )
    warehouse_id = fields.Many2one(
        comodel_name='stock.warehouse',
        string='Entrepôt',
        related='inventory_id.warehouse_id',
        store=True,
        readonly=True,
    )
    
    # Pièces jointes et photos
    image_1 = fields.Binary(
        string='Photo 1',
        attachment=True,
        help='Première photo du produit compté'
    )
    image_2 = fields.Binary(
        string='Photo 2',
        attachment=True,
        help='Deuxième photo du produit compté'
    )
    image_3 = fields.Binary(
        string='Photo 3',
        attachment=True,
        help='Troisième photo du produit compté'
    )
    
    # Champs supplémentaires pour le comptage terrain
    product_brand = fields.Char(
        string='Marque',
        related='product_id.product_tmpl_id.brand',
        store=True,
        readonly=True,
        help='Marque du produit (depuis la fiche article)'
    )
    product_serial = fields.Char(
        string='N° série',
        help='Numéro de série saisi lors du comptage'
    )
    product_has_serial = fields.Boolean(
        string="A un N° série",
        related='product_id.product_tmpl_id.has_serial',
        store=True,
        readonly=True,
        help="Indique si l'article possède un numéro de série (depuis la fiche article)"
    )
    
    note = fields.Text(
        string='Remarques',
        help='Notes ou observations sur cette ligne d\'inventaire'
    )
    
    @api.model_create_multi
    def create(self, vals_list):
        """Auto-remplit location_id depuis l'inventaire parent et product_serial avec le code produit si non fourni."""
        for vals in vals_list:
            # Si location_id n'est pas fourni mais inventory_id oui, récupérer l'emplacement de l'inventaire
            if not vals.get('location_id') and vals.get('inventory_id'):
                inventory = self.env['stockex.stock.inventory'].browse(vals['inventory_id'])
                if inventory.location_id:
                    vals['location_id'] = inventory.location_id.id
            
            # Si product_serial n'est pas fourni mais product_id oui, remplir avec le code produit
            if not vals.get('product_serial') and vals.get('product_id'):
                product = self.env['product.product'].browse(vals['product_id'])
                if product.default_code:
                    vals['product_serial'] = product.default_code
        
        return super().create(vals_list)
    
    @api.depends('product_id')
    def _compute_standard_price(self):
        """Calcule le prix unitaire depuis product.standard_price (source unique de vérité)."""
        for line in self:
            if line.product_id:
                # Utiliser la méthode de valorisation de l'inventaire
                if line.inventory_id:
                    line.standard_price = line.inventory_id._get_product_valuation_price(line.product_id)
                else:
                    # Fallback si pas d'inventaire (cas rare)
                    line.standard_price = line.product_id.standard_price or 0.0
            else:
                line.standard_price = 0.0
    
    @api.depends('product_id', 'location_id')
    def _compute_theoretical_qty(self):
        """Calcule la quantité théorique depuis le stock Odoo.
        
        Cherche dans l'emplacement exact ET ses enfants pour plus de flexibilité.
        """
        # Préparer l'ensemble des lignes pertinentes (avec produit et emplacement)
        lines = self.filtered(lambda l: l.product_id and l.location_id)

        # Initialiser les lignes sans données requises à 0
        for line in (self - lines):
            line.theoretical_qty = 0.0

        if not lines:
            return

        product_ids = list(set(lines.mapped('product_id').ids))
        
        # Récupérer tous les emplacements + leurs enfants
        location_ids = set()
        for loc in lines.mapped('location_id'):
            # Ajouter l'emplacement lui-même
            location_ids.add(loc.id)
            # Ajouter tous ses enfants (child_ids est récursif dans Odoo)
            location_ids.update(loc.child_ids.ids)
        
        location_ids = list(location_ids)
        
        # Récupérer les company_ids des lignes
        company_ids = list(set(lines.mapped('inventory_id.company_id').ids))
        if not company_ids:
            company_ids = [self.env.company.id]

        _logger.info(
            f"🔍 Calcul theoretical_qty pour {len(lines)} lignes, "
            f"{len(product_ids)} produits, {len(location_ids)} emplacements (incluant enfants), "
            f"companies: {company_ids}"
        )

        # Agréger en une seule requête SQL AVEC FILTRE COMPANY ET EMPLACEMENTS ENFANTS
        groups = self.env['stock.quant'].read_group(
            domain=[
                ('product_id', 'in', product_ids),
                ('location_id', 'in', location_ids),  # ✅ Inclut les enfants
                ('company_id', 'in', company_ids),
            ],
            fields=['quantity:sum', 'reserved_quantity:sum'],
            groupby=['product_id', 'location_id'],
        )

        _logger.info(f"📊 {len(groups)} groupes de stock.quant trouvés")

        # Construire un mapping (product_id, location_id) -> quantity
        # Inclure les quantités des emplacements enfants
        qty_map = {}  # (product_id, location_id) -> qty
        
        for g in groups:
            if 'product_id' in g and 'location_id' in g:
                prod_id = g['product_id'][0] if g['product_id'] else None
                loc_id = g['location_id'][0] if g['location_id'] else None
                if prod_id and loc_id:
                    qty = g.get('quantity_sum', 0.0) or 0.0
                    reserved = g.get('reserved_quantity_sum', 0.0) or 0.0
                    available = qty - reserved
                    
                    # Ajouter à toutes les lignes dont le location_id correspond
                    # (soit exact, soit parent de loc_id)
                    for line in lines:
                        if line.product_id.id == prod_id:
                            # Vérifier si loc_id est l'emplacement de la ligne ou un enfant
                            if (loc_id == line.location_id.id or 
                                loc_id in line.location_id.child_ids.ids):
                                key = (line.product_id.id, line.location_id.id)
                                qty_map[key] = qty_map.get(key, 0.0) + available

        # Appliquer les quantités
        updated_count = 0
        for line in lines:
            theo_qty = qty_map.get((line.product_id.id, line.location_id.id), 0.0)
            line.theoretical_qty = theo_qty
            if theo_qty > 0:
                updated_count += 1
        
        if updated_count > 0:
            _logger.info(f"✅ {updated_count} lignes avec quantité théorique > 0")
        else:
            _logger.warning(
                f"⚠️ Aucune ligne avec stock trouvée ! "
                f"Vérifiez que les emplacements correspondent."
            )

    def _inverse_theoretical_qty(self):
        """Méthode inverse pour permettre de forcer la valeur theoretical_qty (pour stock initial)."""
        # Ne rien faire - la valeur est déjà écrite par le create/write
        pass
    
    def action_inspect_image(self):
        """Ouvre une fenêtre pour inspecter les images de la ligne."""
        self.ensure_one()
        view = self.env.ref('stockex.view_stockex_inventory_line_form', raise_if_not_found=False)
        action = {
            'type': 'ir.actions.act_window',
            'name': 'Inspecter la ligne',
            'res_model': 'stockex.stock.inventory.line',
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
        }
        if view:
            action['views'] = [(view.id, 'form')]
        return action
    
    @api.depends('theoretical_qty', 'product_qty', 'inventory_id.is_initial_stock')
    def _compute_difference(self):
        """Calcule la différence entre la quantité réelle et théorique."""
        for line in self:
            line.difference = (line.product_qty or 0.0) - (line.theoretical_qty or 0.0)
    
    @api.depends('difference')
    def _compute_difference_display(self):
        """Affiche la différence en couleur selon le signe."""
        for line in self:
            diff = line.difference
            if diff < 0:
                color = 'red'
                icon = '⚠️'
            elif diff > 0:
                color = 'green'
                icon = '✓'
            else:
                color = 'gray'
                icon = '='
            line.difference_display = f'<span style="color: {color}; font-weight: bold;">{icon} {diff:,.2f}</span>'
    
    @api.onchange('scanned_barcode')
    def _onchange_scanned_barcode(self):
        """Recherche le produit par code-barres scanné."""
        if self.scanned_barcode:
            product = self.env['product.product'].search([
                ('barcode', '=', self.scanned_barcode)
            ], limit=1)
            if product:
                self.product_id = product.id
                self.scanned_barcode = False  # Reset après scan
            else:
                return {
                    'warning': {
                        'title': 'Code-barres non trouvé',
                        'message': f"Aucun produit trouvé avec le code-barres '{self.scanned_barcode}'"
                    }
                }
    
    @api.onchange('product_id', 'location_id')
    def _onchange_product_location(self):
        """Remplit automatiquement la quantité théorique et le numéro de série."""
        if self.product_id:
            # Remplir le numéro de série avec le code produit
            if not self.product_serial and self.product_id.default_code:
                self.product_serial = self.product_id.default_code
            
            # Le prix unitaire est maintenant calculé automatiquement via _compute_standard_price
            
            # Récupérer la quantité théorique si l'emplacement est défini
            if self.location_id:
                quant = self.env['stock.quant'].search([
                    ('product_id', '=', self.product_id.id),
                    ('location_id', '=', self.location_id.id),
                    ('company_id', '=', self.inventory_id.company_id.id if self.inventory_id else self.env.company.id),
                ], limit=1)
                
                if quant:
                    theoretical_qty = quant.quantity - quant.reserved_quantity
                    self.theoretical_qty = theoretical_qty
                    _logger.info(
                        f"✅ Auto-rempli: {self.product_id.default_code} @ {self.location_id.name}: "
                        f"Qté théo={theoretical_qty}"
                    )
                else:
                    self.theoretical_qty = 0.0
    
    @api.depends('product_qty', 'standard_price')
    def _compute_line_value(self):
        for line in self:
            qty = line.product_qty or 0.0
            price = line.standard_price or 0.0
            line.line_value = qty * price

    @api.depends('difference', 'standard_price')
    def _compute_difference_value(self):
        for line in self:
            diff = line.difference or 0.0
            price = line.standard_price or 0.0
            line.difference_value = diff * price

    @api.constrains('product_id', 'inventory_id')
    def _check_product_uniqueness(self):
        """Vérifie qu'un produit n'apparaît qu'une seule fois par inventaire."""
        for line in self:
            domain = [
                ('inventory_id', '=', line.inventory_id.id),
                ('product_id', '=', line.product_id.id),
                ('id', '!=', line.id)
            ]
            if line.location_id:
                domain.append(('location_id', '=', line.location_id.id))
            if self.search_count(domain) > 0:
                raise UserError(
                    f"Le produit '{line.product_id.display_name}' est déjà présent dans cet inventaire "
                    f"pour cet emplacement."
                )



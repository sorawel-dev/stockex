#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script pour construire stock_initial_converti.xlsx complet
en fusionnant stock_initial.xlsx (produits/catégories) 
avec val_stock_brut.xlsx (quantités/prix/emplacements)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

print("=" * 70)
print(" CONSTRUCTION STOCK INITIAL COMPLET")
print("=" * 70)
print()

# ===== CHARGEMENT FICHIER PRODUITS =====
print("[1/6] Chargement fichier produits (stock_initial.xlsx)...")
wb_produits = openpyxl.load_workbook('c:/apps/stockex/docs/stock_initial.xlsx', data_only=True)
ws_produits = wb_produits.active

print(f"✅ Fichier produits chargé : {ws_produits.max_row - 1} lignes")
print()

# ===== CHARGEMENT FICHIER STOCKS =====
print("[2/6] Chargement fichier stocks (val_stock_brut.xlsx)...")
wb_stocks = openpyxl.load_workbook('c:/apps/stockex/docs/val_stock_brut.xlsx', data_only=True)
ws_stocks = wb_stocks.active

print(f"✅ Fichier stocks chargé : {ws_stocks.max_row - 1} lignes")
print()

# ===== CONSTRUCTION INDEX STOCKS PAR CODE PRODUIT =====
print("[3/6] Indexation des stocks par code produit...")

# Structure: stock_data[code_produit] = [(emplacement, quantite, prix), ...]
stock_data = defaultdict(list)

for row_idx in range(2, ws_stocks.max_row + 1):
    code_produit = ws_stocks.cell(row_idx, 5).value  # Col 5 = product_default_code
    emplacement = ws_stocks.cell(row_idx, 4).value   # Col 4 = wharehouse
    quantite = ws_stocks.cell(row_idx, 8).value      # Col 8 = quantity
    prix = ws_stocks.cell(row_idx, 9).value          # Col 9 = standard_price
    
    if code_produit and quantite:
        # Nettoyer les valeurs
        try:
            quantite = float(quantite) if quantite else 0
            prix = float(prix) if prix else 0
            emplacement = str(emplacement).strip() if emplacement else "Stock Principal"
            
            stock_data[str(code_produit).strip()].append({
                'emplacement': emplacement,
                'quantite': quantite,
                'prix': prix
            })
        except (ValueError, TypeError):
            pass

print(f"✅ {len(stock_data)} produits indexés avec stock")
print()

# ===== CRÉATION NOUVEAU FICHIER =====
print("[4/6] Création du fichier de destination...")
wb_dest = openpyxl.Workbook()
ws_dest = wb_dest.active
ws_dest.title = 'Stock Initial'

# En-têtes
headers = ['CODE PRODUIT', 'NOM PRODUIT', 'CATEGORIE', 'CODE CATEGORIE', 'EMPLACEMENT', 'QUANTITE', 'PRIX UNITAIRE']

for col, header in enumerate(headers, 1):
    cell = ws_dest.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color='FFFFFF', size=12)
    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

ws_dest.row_dimensions[1].height = 25
print("✅ En-têtes créés")
print()

# ===== FUSION DES DONNÉES =====
print("[5/6] Fusion des données produits + stocks...")

lignes_creees = 0
produits_sans_stock = 0
produits_avec_stock = 0
total_lignes_stock = 0

current_row = 2

for row_idx in range(2, ws_produits.max_row + 1):
    # Données produit du fichier stock_initial.xlsx
    # Col 1=N°, 2=CODE CAT, 3=CATEGORIE, 4=CODE PRODUIT, 5=PRODUIT, 6=UDM
    code_categorie = ws_produits.cell(row_idx, 2).value
    categorie = ws_produits.cell(row_idx, 3).value
    code_produit = ws_produits.cell(row_idx, 4).value
    nom_produit = ws_produits.cell(row_idx, 5).value
    
    if not code_produit:
        continue
    
    code_produit = str(code_produit).strip()
    
    # Chercher les stocks pour ce produit
    stocks_produit = stock_data.get(code_produit, [])
    
    if stocks_produit:
        # Produit avec stock : créer une ligne par emplacement
        produits_avec_stock += 1
        for stock in stocks_produit:
            # CODE PRODUIT
            ws_dest.cell(current_row, 1, code_produit).alignment = Alignment(horizontal='left')
            
            # NOM PRODUIT
            ws_dest.cell(current_row, 2, nom_produit).alignment = Alignment(horizontal='left')
            
            # CATEGORIE
            ws_dest.cell(current_row, 3, categorie).alignment = Alignment(horizontal='left')
            
            # CODE CATEGORIE
            ws_dest.cell(current_row, 4, code_categorie).alignment = Alignment(horizontal='left')
            
            # EMPLACEMENT
            ws_dest.cell(current_row, 5, stock['emplacement']).alignment = Alignment(horizontal='left')
            
            # QUANTITE
            cell = ws_dest.cell(current_row, 6, stock['quantite'])
            cell.alignment = Alignment(horizontal='right')
            cell.number_format = '0.00'
            
            # PRIX UNITAIRE
            cell = ws_dest.cell(current_row, 7, stock['prix'])
            cell.alignment = Alignment(horizontal='right')
            cell.number_format = '#,##0.00'
            
            # Bordures
            for col in range(1, 8):
                ws_dest.cell(current_row, col).border = Border(
                    left=Side(style='thin', color='D3D3D3'),
                    right=Side(style='thin', color='D3D3D3'),
                    top=Side(style='thin', color='D3D3D3'),
                    bottom=Side(style='thin', color='D3D3D3')
                )
            
            current_row += 1
            lignes_creees += 1
            total_lignes_stock += 1
    else:
        # Produit sans stock : créer une ligne avec valeurs par défaut
        produits_sans_stock += 1
        
        # CODE PRODUIT
        ws_dest.cell(current_row, 1, code_produit).alignment = Alignment(horizontal='left')
        
        # NOM PRODUIT
        ws_dest.cell(current_row, 2, nom_produit).alignment = Alignment(horizontal='left')
        
        # CATEGORIE
        ws_dest.cell(current_row, 3, categorie).alignment = Alignment(horizontal='left')
        
        # CODE CATEGORIE
        ws_dest.cell(current_row, 4, code_categorie).alignment = Alignment(horizontal='left')
        
        # ENTREPOT
        ws_dest.cell(current_row, 5, "Stock Principal").alignment = Alignment(horizontal='left')
        
        # QUANTITE (0 - fond jaune)
        cell = ws_dest.cell(current_row, 6, 0)
        cell.alignment = Alignment(horizontal='right')
        cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        
        # PRIX UNITAIRE (0 - fond jaune)
        cell = ws_dest.cell(current_row, 7, 0)
        cell.alignment = Alignment(horizontal='right')
        cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        
        # Bordures
        for col in range(1, 8):
            ws_dest.cell(current_row, col).border = Border(
                left=Side(style='thin', color='D3D3D3'),
                right=Side(style='thin', color='D3D3D3'),
                top=Side(style='thin', color='D3D3D3'),
                bottom=Side(style='thin', color='D3D3D3')
            )
        
        current_row += 1
        lignes_creees += 1
    
    if lignes_creees % 500 == 0:
        print(f"  {lignes_creees} lignes créées...")

print(f"✅ {lignes_creees} lignes créées")
print(f"   • Produits avec stock : {produits_avec_stock} ({total_lignes_stock} lignes)")
print(f"   • Produits sans stock : {produits_sans_stock}")
print()

# ===== LARGEUR COLONNES =====
print("[6/6] Ajustement des colonnes et sauvegarde...")
ws_dest.column_dimensions['A'].width = 18  # CODE PRODUIT
ws_dest.column_dimensions['B'].width = 40  # NOM PRODUIT
ws_dest.column_dimensions['C'].width = 35  # CATEGORIE
ws_dest.column_dimensions['D'].width = 18  # CODE CATEGORIE
ws_dest.column_dimensions['E'].width = 25  # EMPLACEMENT
ws_dest.column_dimensions['F'].width = 12  # QUANTITE
ws_dest.column_dimensions['G'].width = 15  # PRIX UNITAIRE

# Sauvegarde
output_file = 'c:/apps/stockex/docs/stock_initial_COMPLET.xlsx'
wb_dest.save(output_file)

print(f"✅ Fichier sauvegardé : {output_file}")
print()

# ===== RÉSUMÉ =====
print("=" * 70)
print(" CONSTRUCTION TERMINÉE")
print("=" * 70)
print()
print(f"📁 Fichiers sources :")
print(f"   • stock_initial.xlsx (produits/catégories)")
print(f"   • val_stock_brut.xlsx (stocks/prix/emplacements)")
print()
print(f"📁 Fichier destination : {output_file}")
print()
print(f"📊 Statistiques :")
print(f"   • Total lignes créées       : {lignes_creees}")
print(f"   • Produits avec stock       : {produits_avec_stock}")
print(f"   • Lignes avec stock         : {total_lignes_stock}")
print(f"   • Produits sans stock       : {produits_sans_stock} (quantité/prix = 0)")
print()
print("✅ Avantages :")
print("   • Quantités et prix récupérés depuis val_stock_brut.xlsx")
print("   • Emplacements récupérés (multi-emplacements supportés)")
print("   • Produits sans stock marqués (fond jaune) pour remplissage manuel")
print()
print("📋 Format du fichier :")
for i, h in enumerate(headers, 1):
    print(f"   {i}. {h}")
print()
print("🚀 Prochaine étape :")
print("   1. Vérifiez le fichier converti")
print("   2. Complétez les produits sans stock (fond jaune) si nécessaire")
print("   3. Importez via : Gestion d'Inventaire → Import → 📦 Stock Initial")
print()

# Fermer les fichiers
wb_produits.close()
wb_stocks.close()
wb_dest.close()

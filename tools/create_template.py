#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script pour créer le template Excel Stock Initial
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Créer le workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Stock Initial'

# ===== TITRE =====
ws['A1'] = 'TEMPLATE STOCK INITIAL - STOCKEX'
ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
ws['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.merge_cells('A1:F1')
ws.row_dimensions[1].height = 30

# ===== INSTRUCTIONS =====
ws['A3'] = 'Instructions:'
ws['A3'].font = Font(bold=True, size=12)

ws['A4'] = '1. Remplissez les colonnes ci-dessous avec vos données'
ws['A5'] = '2. CODE PRODUIT doit être unique (ex: ASP001, PAR002)'
ws['A6'] = '3. QUANTITE et PRIX UNITAIRE doivent être des nombres'
ws['A7'] = '4. Sauvegardez ce fichier et importez-le via: Gestion d\'Inventaire → Import → 📦 Stock Initial'

# ===== EN-TÊTES =====
headers = ['CODE PRODUIT', 'NOM PRODUIT', 'CATEGORIE', 'CODE CATEGORIE', 'ENTREPOT', 'QUANTITE', 'PRIX UNITAIRE']

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=9, column=col, value=header)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

ws.row_dimensions[9].height = 25

# ===== EXEMPLES =====
examples = [
    ['ASP001', 'Aspirine 500mg', 'Médicaments', 'MED001', 'Stock Principal', 1000, 500],
    ['PAR001', 'Paracétamol 1g', 'Médicaments', 'MED001', 'Stock Principal', 850, 450],
    ['AMX001', 'Amoxicilline 500mg', 'Antibiotiques', 'ATB001', 'Stock Principal', 500, 1200],
    ['IBU001', 'Ibuproène 400mg', 'Anti-inflammatoires', 'ANT001', 'Stock Principal', 750, 350],
    ['VIT001', 'Vitamine C 1000mg', 'Vitamines', 'VIT001', 'Stock Principal', 2000, 200],
]

for row_idx, example in enumerate(examples, 10):
    for col_idx, value in enumerate(example, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(horizontal='left' if col_idx <= 5 else 'right')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Fond gris pour les exemples
        if col_idx <= 5:
            cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')

# ===== LARGEUR COLONNES =====
for col in range(1, 8):
    ws.column_dimensions[get_column_letter(col)].width = 20

# ===== SÉPARATEUR =====
ws['A16'] = 'AJOUTEZ VOS PRODUITS CI-DESSOUS'
ws['A16'].font = Font(bold=True, italic=True, color='808080')
ws['A16'].alignment = Alignment(horizontal='center')
ws.merge_cells('A16:G16')

# ===== LIGNES VIDES (100 lignes) =====
for row in range(17, 117):
    for col in range(1, 8):
        cell = ws.cell(row=row, column=col)
        cell.border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )

# ===== SAUVEGARDER =====
output_path = 'c:/apps/stockex/template_stock_initial.xlsx'
wb.save(output_path)

print('✅ Template créé avec succès!')
print(f'📁 Fichier: {output_path}')
print('')
print('📋 Colonnes:')
for i, h in enumerate(headers, 1):
    print(f'  {i}. {h}')
print('')
print('💡 5 exemples de produits inclus')
print('💡 100 lignes vides prêtes à remplir')

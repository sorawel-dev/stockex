#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script pour convertir stock_initial.xlsx au format template StockInv
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

print("=" * 60)
print(" CONVERSION STOCK_INITIAL.XLSX")
print("=" * 60)
print()

# Charger le fichier source
print("[1/5] Chargement du fichier source...")
source_file = 'c:/apps/stockex/docs/stock_initial.xlsx'
wb_source = openpyxl.load_workbook(source_file, data_only=True)
ws_source = wb_source.active

print(f"✅ Fichier chargé : {ws_source.max_row} lignes")
print()

# Créer le nouveau workbook
print("[2/5] Création du nouveau fichier...")
wb_dest = openpyxl.Workbook()
ws_dest = wb_dest.active
ws_dest.title = 'Stock Initial'

# ===== EN-TÊTES DU NOUVEAU FORMAT =====
headers = ['CODE PRODUIT', 'NOM PRODUIT', 'CATEGORIE', 'CODE CATEGORIE', 'EMPLACEMENT', 'QUANTITE', 'PRIX UNITAIRE']

# Style des en-têtes
for col, header in enumerate(headers, 1):
    cell = ws_dest.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color='FFFFFF', size=12)
    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

ws_dest.row_dimensions[1].height = 25

print("✅ En-têtes créés")
print()

# ===== CONVERSION DES DONNÉES =====
print("[3/5] Conversion des données...")

# Mapping des colonnes source (N°=0, CODE CATEGORIE=1, CATEGORIE=2, CODE PRODUIT=3, PRODUIT=4, UDM=5)
# vers destination (CODE PRODUIT=0, NOM PRODUIT=1, CATEGORIE=2, CODE CATEGORIE=3, EMPLACEMENT=4, QUANTITE=5, PRIX=6)

converted_count = 0
skipped_count = 0
emplacement_default = "Stock Principal"

for row_idx in range(2, ws_source.max_row + 1):
    row_data = [cell.value for cell in ws_source[row_idx]]
    
    # Extraire les données source
    code_categorie = row_data[1] if len(row_data) > 1 else ''
    categorie = row_data[2] if len(row_data) > 2 else ''
    code_produit = row_data[3] if len(row_data) > 3 else ''
    produit = row_data[4] if len(row_data) > 4 else ''
    
    # Vérifier que le code produit existe
    if not code_produit:
        skipped_count += 1
        continue
    
    # Écrire dans le nouveau format
    dest_row = converted_count + 2  # +2 car ligne 1 = en-têtes
    
    # CODE PRODUIT
    cell = ws_dest.cell(row=dest_row, column=1, value=code_produit)
    cell.alignment = Alignment(horizontal='left')
    
    # NOM PRODUIT
    cell = ws_dest.cell(row=dest_row, column=2, value=produit)
    cell.alignment = Alignment(horizontal='left')
    
    # CATEGORIE
    cell = ws_dest.cell(row=dest_row, column=3, value=categorie)
    cell.alignment = Alignment(horizontal='left')
    
    # CODE CATEGORIE
    cell = ws_dest.cell(row=dest_row, column=4, value=code_categorie)
    cell.alignment = Alignment(horizontal='left')
    
    # EMPLACEMENT
    cell = ws_dest.cell(row=dest_row, column=5, value=emplacement_default)
    cell.alignment = Alignment(horizontal='left')
    
    # QUANTITE (vide - à remplir manuellement)
    cell = ws_dest.cell(row=dest_row, column=6, value=0)
    cell.alignment = Alignment(horizontal='right')
    cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')  # Jaune clair
    
    # PRIX UNITAIRE (vide - à remplir manuellement)
    cell = ws_dest.cell(row=dest_row, column=7, value=0)
    cell.alignment = Alignment(horizontal='right')
    cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')  # Jaune clair
    
    # Bordures pour toute la ligne
    for col in range(1, 8):
        ws_dest.cell(row=dest_row, column=col).border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )
    
    converted_count += 1
    
    if converted_count % 500 == 0:
        print(f"  {converted_count} lignes converties...")

print(f"✅ {converted_count} lignes converties")
print(f"⚠️  {skipped_count} lignes ignorées (code produit manquant)")
print()

# ===== LARGEUR COLONNES =====
print("[4/5] Ajustement des largeurs de colonnes...")
ws_dest.column_dimensions['A'].width = 18  # CODE PRODUIT
ws_dest.column_dimensions['B'].width = 35  # NOM PRODUIT
ws_dest.column_dimensions['C'].width = 30  # CATEGORIE
ws_dest.column_dimensions['D'].width = 18  # CODE CATEGORIE
ws_dest.column_dimensions['E'].width = 20  # EMPLACEMENT
ws_dest.column_dimensions['F'].width = 12  # QUANTITE
ws_dest.column_dimensions['G'].width = 15  # PRIX UNITAIRE

print("✅ Largeurs ajustées")
print()

# ===== SAUVEGARDE =====
print("[5/5] Sauvegarde du fichier...")
output_file = 'c:/apps/stockex/docs/stock_initial_converti.xlsx'
wb_dest.save(output_file)

print(f"✅ Fichier sauvegardé : {output_file}")
print()

# ===== RÉSUMÉ =====
print("=" * 60)
print(" CONVERSION TERMINÉE")
print("=" * 60)
print()
print(f"📁 Fichier source    : {source_file}")
print(f"📁 Fichier destination : {output_file}")
print()
print(f"📊 Statistiques :")
print(f"   • Lignes converties  : {converted_count}")
print(f"   • Lignes ignorées    : {skipped_count}")
print(f"   • Total              : {ws_source.max_row - 1}")
print()
print("⚠️  IMPORTANT :")
print("   Les colonnes QUANTITE et PRIX UNITAIRE sont à 0 (fond jaune)")
print("   Vous devez les remplir manuellement avec vos quantités et prix")
print()
print("📋 Format du fichier :")
for i, h in enumerate(headers, 1):
    print(f"   {i}. {h}")
print()
print("🚀 Prochaine étape :")
print("   1. Ouvrez le fichier converti")
print("   2. Remplissez les colonnes QUANTITE et PRIX UNITAIRE")
print("   3. Importez via : Gestion d'Inventaire → Import → 📦 Stock Initial")
print()

# Fermer les fichiers
wb_source.close()
wb_dest.close()
